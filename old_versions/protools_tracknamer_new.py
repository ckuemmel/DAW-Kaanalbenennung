import argparse
import csv
import sys
import time
from pathlib import Path
import threading

try:
    from pynput import keyboard
    from pynput.keyboard import Key, Controller
except Exception as e:
    print("Fehler: Das Paket 'pynput' ist nicht installiert oder konnte nicht geladen werden.")
    print("Installiere es mit: pip install pynput")
    sys.exit(1)

def send_next_track(kb: Controller):
    """Sendet den Befehl zum Wechseln zur nächsten Spur"""
    try:
        print("DEBUG: Sende Strg+Rechts")
        # Stelle sicher, dass keine Tasten hängenbleiben
        kb.release(Key.ctrl)
        kb.release(Key.right)
        time.sleep(0.2)
        
        # Ein einzelner, präziser Strg+Rechts
        kb.press(Key.ctrl)
        time.sleep(0.1)
        kb.press(Key.right)
        time.sleep(0.1)
        kb.release(Key.right)
        time.sleep(0.1)
        kb.release(Key.ctrl)
        print("DEBUG: Strg+Rechts gesendet")
    except Exception as e:
        print(f"DEBUG: Fehler beim Senden von Strg+Rechts: {e}")

class NameTyper:
    def __init__(self, names, args):
        self.names = names
        self.args = args
        self.idx = 0
        self.kb = Controller()
        self.paused = False
        self.auto_running = False
        self.stop_requested = False
        self.worker_thread = None

    def select_all(self):
        try:
            if self.args.auto_next == 'mac':
                self.kb.press(Key.cmd)
                self.kb.press('a')
                self.kb.release('a')
                self.kb.release(Key.cmd)
            else:
                self.kb.press(Key.ctrl)
                self.kb.press('a')
                self.kb.release('a')
                self.kb.release(Key.ctrl)
        except Exception:
            pass

    def type_next_name(self):
        if self.idx >= len(self.names):
            print("Fertig: Alle Namen wurden verwendet.")
            return

        name, channel, mic = self.names[self.idx]
        self.idx += 1

        # Formatiere den Namen mit Unterstrichen
        formatted_name = name
        if self.args.kanal_spalte and channel:
            formatted_name = f"{channel}_{formatted_name}"
        if self.args.mikrofon_spalte and mic:
            formatted_name = f"{formatted_name}_{mic}"

        # Warte und tippe
        time.sleep(self.args.delay)
        self.select_all()
        time.sleep(0.2)
        self.kb.type(formatted_name)
        print(f"Getippt: {formatted_name}")

        # Navigation zur nächsten Spur
        if self.args.auto_next:
            time.sleep(self.args.next_delay)
            if self.args.auto_next == 'windows':
                send_next_track(self.kb)
            elif self.args.auto_next == 'mac':
                self.kb.press(Key.cmd)
                self.kb.press(Key.right)
                self.kb.release(Key.right)
                self.kb.release(Key.cmd)
            time.sleep(self.args.next_delay)
        elif self.args.enter:
            self.kb.press(Key.enter)
            self.kb.release(Key.enter)

    def on_key(self, key):
        try:
            if key == Key.f9:
                if self.auto_running:
                    self.stop_requested = True
                    if self.worker_thread and self.worker_thread.is_alive():
                        self.worker_thread.join(timeout=1)
                    self.auto_running = False
                self.paused = not self.paused
                print(f"Pause: {self.paused}")
            elif key == Key.f7:
                if self.idx > 0:
                    self.idx -= 1
                print(f"Zurück. Nächster Index: {self.idx}")
            elif key == Key.f8 and not self.paused:
                if self.args.auto_run and not self.auto_running:
                    # Starte automatischen Modus
                    self.stop_requested = False
                    self.auto_running = True
                    print("Auto-Run gestartet. ESC zum Abbrechen, F9 für Pause.")
                    
                    while self.idx < len(self.names) and not self.stop_requested and not self.paused:
                        self.type_next_name()
                        if self.stop_requested or self.paused:
                            break
                        time.sleep(self.args.delay)
                    
                    self.auto_running = False
                    if self.idx >= len(self.names):
                        print("Fertig: Alle Namen wurden verwendet.")
                else:
                    # Einzelner Name
                    self.type_next_name()
            elif key == Key.esc:
                self.stop_requested = True
                if self.worker_thread and self.worker_thread.is_alive():
                    self.worker_thread.join(timeout=1)
                print("Beendet.")
                return False
        except Exception as e:
            print(f"Fehler beim Senden der Tasten: {e}")

    def run(self):
        with keyboard.Listener(on_press=self.on_key) as listener:
            listener.join()

def main():
    parser = argparse.ArgumentParser(
        description="Überträgt Spur-Namen aus CSV/XLSX in Pro Tools per Hotkey (F8).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument('datei', type=str, help='Pfad zur CSV oder XLSX Datei')
    parser.add_argument('--spalte', type=int, default=3, help='Spaltenindex (0-basiert)')
    parser.add_argument('--kanal-spalte', type=int, default=1, help='Spaltenindex für Kanalnummern')
    parser.add_argument('--mikrofon-spalte', type=int, default=4, help='Spaltenindex für Mikrofone')
    parser.add_argument('--header-row', type=int, default=8, help='Headerzeile (1-basiert)')
    parser.add_argument('--delay', type=float, default=1.0, help='Verzögerung vor dem Tippen (Sekunden)')
    parser.add_argument('--next-delay', type=float, default=2.0, help='Verzögerung vor/nach Spurwechsel (Sekunden)')
    parser.add_argument('--auto-next', choices=['windows', 'mac'], help='Nach Eingabe zur nächsten Spur')
    parser.add_argument('--auto-run', action='store_true', help='Automatisch alle Namen übertragen (Start mit F8)')
    args = parser.parse_args()

    # Namen aus Excel laden
    try:
        import openpyxl
    except ImportError:
        print("Fehler: openpyxl ist nicht installiert. Bitte installieren Sie es mit:")
        print("pip install openpyxl")
        sys.exit(1)

    # Excel-Datei laden
    wb = openpyxl.load_workbook(args.datei, read_only=True, data_only=True)
    ws = wb.active
    
    # Namen einlesen
    names = []
    for r in range(args.header_row, ws.max_row + 1):
        name_cell = ws.cell(row=r, column=args.spalte + 1)
        channel_cell = ws.cell(row=r, column=args.kanal_spalte + 1)
        mic_cell = ws.cell(row=r, column=args.mikrofon_spalte + 1)
        
        if name_cell.value:
            name = str(name_cell.value).strip()
            channel = str(channel_cell.value).strip() if channel_cell.value else None
            mic = str(mic_cell.value).strip() if mic_cell.value else None
            if name:  # Nur hinzufügen wenn ein Name vorhanden ist
                names.append((name, channel, mic))

    if not names:
        print("Keine Namen in der Excel-Datei gefunden!")
        sys.exit(1)

    print(f"DEBUG: {len(names)} Namen geladen")
    print("DEBUG: Erste 3 Namen zur Kontrolle:")
    for i, (name, channel, mic) in enumerate(names[:3]):
        print(f"DEBUG: {i + 1}: {channel} {name} {mic}")

    typer = NameTyper(names, args)
    typer.run()

if __name__ == '__main__':
    main()
