#!/usr/bin/env python3
"""
TrackNamer GUI - OHNE TreeView, mit Text-Widget und separaten Checkboxen
"""
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import openpyxl

class TextBasedTrackNamerGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Pro Tools TrackNamer - Text-Basiert")
        self.geometry("1000x800")
        
        self.data = []
        self.checkboxes = []
        
        self.create_widgets()
        self.add_test_data()  # Sofort Test-Daten anzeigen
    
    def create_widgets(self):
        # Header
        header = tk.Label(self, text="🎵 Pro Tools Track Namer", font=("Arial", 16, "bold"))
        header.pack(pady=10)
        
        # Excel laden
        load_frame = tk.Frame(self)
        load_frame.pack(pady=5)
        tk.Button(load_frame, text="📁 Excel laden", command=self.load_excel, bg="lightblue", font=("Arial", 12)).pack(side=tk.LEFT, padx=5)
        
        # Layout-Auswahl
        self.layout_var = tk.StringVar(value="auto")
        layout_frame = tk.Frame(load_frame)
        layout_frame.pack(side=tk.LEFT, padx=20)
        tk.Label(layout_frame, text="Layout:").pack(side=tk.LEFT)
        
        for text, value in [("Auto", "auto"), ("A(C,D)", "layout_a"), ("B(D,E)", "layout_b")]:
            tk.Radiobutton(layout_frame, text=text, variable=self.layout_var, value=value).pack(side=tk.LEFT)
        
        # Status
        self.status_label = tk.Label(self, text="TEST-DATEN geladen - Excel laden für echte Daten", fg="green", font=("Arial", 10))
        self.status_label.pack(pady=5)
        
        # Haupt-Frame mit Scroll
        main_frame = tk.Frame(self)
        main_frame.pack(expand=True, fill="both", padx=20, pady=10)
        
        # Canvas + Scrollbar für Checkboxen
        self.canvas = tk.Canvas(main_frame, bg="white")
        self.scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Frame inside Canvas
        self.data_frame = tk.Frame(self.canvas, bg="white")
        self.canvas.create_window((0, 0), window=self.data_frame, anchor="nw")
        
        # Buttons
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="✓ Alle", command=self.select_all, bg="lightgreen").pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="✗ Keine", command=self.deselect_all, bg="lightcoral").pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="🎵 Export", command=self.export_names, bg="gold", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=10)
        
        # Export-Pfad
        path_frame = tk.Frame(self)
        path_frame.pack(pady=5)
        tk.Label(path_frame, text="Speichern als:").pack(side=tk.LEFT)
        self.export_path = tk.StringVar(value="tracknamen.txt")
        tk.Entry(path_frame, textvariable=self.export_path, width=40).pack(side=tk.LEFT, padx=5)
        tk.Button(path_frame, text="📁", command=self.choose_export_path).pack(side=tk.LEFT)
        
        # Update canvas scroll region when frame changes
        self.data_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
    
    def add_test_data(self):
        """Füge Test-Daten hinzu, damit sofort etwas sichtbar ist"""
        test_data = [
            [1, "Violine 1", "MK 4"],
            [2, "Violine 2", "MK 4"],
            [3, "Viola", "MK 4"],
            [4, "Cello", "MK 4"],
            [5, "Kontrabass", "MK 22"]
        ]
        
        self.data = test_data
        self.display_data()
    
    def display_data(self):
        """Zeige Daten mit Checkboxen an"""
        # Alte Widgets entfernen
        for widget in self.data_frame.winfo_children():
            widget.destroy()
        self.checkboxes = []
        
        # Header
        header_frame = tk.Frame(self.data_frame, bg="lightgray", relief="raised", bd=1)
        header_frame.pack(fill="x", padx=5, pady=2)
        
        tk.Label(header_frame, text="Auswahl", width=8, bg="lightgray", font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="Kanal", width=8, bg="lightgray", font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="Instrument", width=20, bg="lightgray", font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="Mikrofon", width=20, bg="lightgray", font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="Trackname", width=25, bg="lightgray", font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        
        # Datenzeilen
        for i, row in enumerate(self.data):
            row_frame = tk.Frame(self.data_frame, bg="white" if i % 2 == 0 else "#f0f0f0", relief="solid", bd=1)
            row_frame.pack(fill="x", padx=5, pady=1)
            
            # Checkbox
            var = tk.IntVar(value=1)
            cb = tk.Checkbutton(row_frame, variable=var, bg=row_frame["bg"])
            cb.pack(side=tk.LEFT, padx=5)
            self.checkboxes.append(var)
            
            # Daten
            kanal = str(int(row[0])) if row[0] else ""
            instrument = str(row[1]) if row[1] else ""
            mikrofon = str(row[2]) if row[2] else ""
            trackname = f"{kanal}_{instrument}_{mikrofon}"
            
            tk.Label(row_frame, text=kanal, width=8, anchor="w", bg=row_frame["bg"]).pack(side=tk.LEFT)
            tk.Label(row_frame, text=instrument, width=20, anchor="w", bg=row_frame["bg"]).pack(side=tk.LEFT)
            tk.Label(row_frame, text=mikrofon, width=20, anchor="w", bg=row_frame["bg"]).pack(side=tk.LEFT)
            tk.Label(row_frame, text=trackname, width=25, anchor="w", bg=row_frame["bg"], font=("Courier", 9)).pack(side=tk.LEFT)
        
        # Canvas aktualisieren
        self.canvas.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        print(f"Angezeigt: {len(self.data)} Zeilen")
    
    def load_excel(self):
        filetypes = [
            ("Excel files", ("*.xlsx", "*.xls", "*.xlsm")),
            ("All files", "*.*")
        ]
        filepath = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=filetypes)
        
        if not filepath:
            return
        
        try:
            self.status_label.config(text=f"Lade: {filepath}...", fg="blue")
            self.update()  # GUI aktualisieren
            
            # Excel laden
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Layout bestimmen
            layout = self.layout_var.get()
            if layout == "auto":
                header_d = sheet.cell(row=6, column=4).value
                detected_layout = "layout_b" if header_d and "instrument" in str(header_d).lower() else "layout_a"
            else:
                detected_layout = layout
            
            # Spalten-Indizes
            wanted_indices = [2, 3, 4] if detected_layout == "layout_a" else [2, 4, 5]
            
            # Daten laden
            self.data = []
            for row_num in range(7, min(sheet.max_row + 1, 200)):  # Max 200 Zeilen
                row_data = [sheet.cell(row=row_num, column=col_idx).value for col_idx in wanted_indices]
                
                if row_data[1]:  # Instrument vorhanden
                    self.data.append(row_data)
            
            workbook.close()
            
            # Anzeigen
            self.display_data()
            self.status_label.config(text=f"✓ {len(self.data)} Zeilen geladen ({detected_layout})", fg="green")
            
        except Exception as e:
            self.status_label.config(text=f"Fehler: {e}", fg="red")
            messagebox.showerror("Fehler", f"Excel konnte nicht geladen werden:\n{e}")
    
    def select_all(self):
        for var in self.checkboxes:
            var.set(1)
    
    def deselect_all(self):
        for var in self.checkboxes:
            var.set(0)
    
    def choose_export_path(self):
        path = filedialog.asksaveasfilename(
            title="Speicherort wählen",
            defaultextension=".txt",
            filetypes=[("Textdatei", "*.txt")]
        )
        if path:
            self.export_path.set(path)
    
    def export_names(self):
        selected_indices = [i for i, var in enumerate(self.checkboxes) if var.get()]
        
        if not selected_indices:
            messagebox.showinfo("Hinweis", "Bitte wähle mindestens eine Zeile aus.")
            return
        
        try:
            names = []
            for i in selected_indices:
                row = self.data[i]
                kanal = str(int(row[0])) if row[0] else ""
                instrument = str(row[1]) if row[1] else ""
                mikrofon = str(row[2]) if row[2] else ""
                name = f"{kanal}_{instrument}_{mikrofon}"
                names.append(name)
            
            out_path = self.export_path.get()
            if not out_path:
                messagebox.showinfo("Hinweis", "Bitte wähle einen Speicherpfad.")
                return
            
            with open(out_path, "w", encoding="utf-8") as f:
                for i, name in enumerate(names, 1):
                    f.write(f"Spur {i}  [{name}]\n")
            
            messagebox.showinfo("Erfolg", f"🎵 {len(names)} Tracknamen gespeichert!")
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Export fehlgeschlagen:\n{e}")

if __name__ == "__main__":
    app = TextBasedTrackNamerGUI()
    app.mainloop()