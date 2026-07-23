#!/usr/bin/env python3
"""
Web-basierter Track Namer - läuft im Browser statt Tkinter
"""
import os
from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify
import openpyxl
import tempfile
from werkzeug.utils import secure_filename
import webbrowser
import threading
import time

app = Flask(__name__)
app.secret_key = 'tracknamer_secret_key'

# Globale Variablen für Daten
current_data = []
current_layout = "auto"

@app.route('/')
def index():
    return render_template('index.html', data=current_data)

@app.route('/upload', methods=['POST'])
def upload_excel():
    global current_data, current_layout
    
    if 'excel_file' not in request.files:
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('index'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('index'))
    
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        flash('Nur Excel-Dateien (.xlsx, .xls, .xlsm) erlaubt', 'error')
        return redirect(url_for('index'))
    
    try:
        current_layout = request.form.get('layout', 'auto')
        
        # Temporäre Datei erstellen
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            
            # Excel laden, bevorzugt den Tab "Inputliste" nutzen
            workbook = openpyxl.load_workbook(tmp_file.name)
            sheet = workbook['Inputliste'] if 'Inputliste' in workbook.sheetnames else workbook.active

            inputliste_header = str(sheet.cell(row=7, column=2).value or "").strip().lower()
            is_inputliste = sheet.title.strip().lower() == "inputliste" or (
                "signal" in inputliste_header and "instrument" in inputliste_header
            )

            # Layout bestimmen
            if is_inputliste:
                detected_layout = "inputliste"
            elif current_layout == "auto":
                header_d = sheet.cell(row=6, column=4).value
                detected_layout = "layout_b" if header_d and "instrument" in str(header_d).lower() else "layout_a"
            else:
                detected_layout = current_layout

            # Spalten-Indizes und Startzeile
            if is_inputliste:
                wanted_indices = [1, 2, 3]  # A=Kanal, B=Signal/Instrument, C=Mikrofon
                row_start = 8
            else:
                wanted_indices = [2, 3, 4] if detected_layout == "layout_a" else [2, 4, 5]
                row_start = 7
            
            # Daten laden
            current_data = []
            for row_num in range(row_start, min(sheet.max_row + 1, 200)):
                row_data = [sheet.cell(row=row_num, column=col_idx).value for col_idx in wanted_indices]

                if is_inputliste:
                    kanal = str(row_data[0]).strip() if row_data[0] is not None else ""
                    if kanal:
                        try:
                            kanal_float = float(kanal.replace(",", "."))
                            kanal = str(int(kanal_float)) if kanal_float.is_integer() else kanal
                        except ValueError:
                            pass
                    instrument = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else ""
                    mikrofon = str(row_data[2]).strip() if len(row_data) > 2 and row_data[2] else ""
                else:
                    if not row_data[1]:
                        continue

                    kanal_text = str(row_data[0]).strip() if row_data[0] is not None else ""
                    if kanal_text:
                        try:
                            kanal_float = float(kanal_text.replace(",", "."))
                            kanal = str(int(kanal_float)) if kanal_float.is_integer() else kanal_text
                        except ValueError:
                            kanal = kanal_text
                    else:
                        kanal = ""

                    instrument = str(row_data[1]).strip() if row_data[1] else ""
                    mikrofon = str(row_data[2]).strip() if row_data[2] else ""

                if instrument:
                    track_parts = [part for part in [kanal, instrument, mikrofon] if part]
                    trackname = "_".join(track_parts)
                    
                    current_data.append({
                        'id': len(current_data),
                        'kanal': kanal,
                        'instrument': instrument,
                        'mikrofon': mikrofon,
                        'trackname': trackname,
                        'selected': True
                    })
            
            workbook.close()
            os.unlink(tmp_file.name)  # Temporäre Datei löschen
            
            flash(f'✅ {len(current_data)} Zeilen geladen ({detected_layout})', 'success')
            
    except Exception as e:
        flash(f'Fehler beim Laden: {e}', 'error')
    
    return redirect(url_for('index'))

@app.route('/export', methods=['POST'])
def export_tracks():
    try:
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            flash('Bitte mindestens eine Spur auswählen', 'error')
            return redirect(url_for('index'))
        
        # Tracknamen sammeln
        names = []
        for item in current_data:
            if item['id'] in selected_ids:
                names.append(item['trackname'])
        
        # Temporäre Export-Datei erstellen
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.txt', encoding='utf-8') as tmp_file:
            for i, name in enumerate(names, 1):
                tmp_file.write(f"Spur {i}  [{name}]\n")
            tmp_file_path = tmp_file.name
        
        flash(f'🎵 {len(names)} Tracknamen exportiert', 'success')
        return send_file(tmp_file_path, as_attachment=True, download_name='tracknamen.txt')
        
    except Exception as e:
        flash(f'Export-Fehler: {e}', 'error')
        return redirect(url_for('index'))

@app.route('/select_all')
def select_all():
    return jsonify({'action': 'select_all'})

@app.route('/select_none')
def select_none():
    return jsonify({'action': 'select_none'})

@app.route('/manual_instructions')
def manual_instructions():
    show_manual_instructions()
    return jsonify({'success': '📋 Manuelle Anleitung im Terminal angezeigt'})

@app.route('/protools_create_tracks', methods=['POST'])
def protools_create_tracks():
    """Pro Tools: Spuren erstellen - Browser-Button Version"""
    try:
        print("🔥 Browser-Button 'Spuren erstellen' geklickt!")
        
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            print("❌ Keine Spuren ausgewählt")
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        track_count = len(selected_ids)
        print(f"🎯 {track_count} Spuren werden erstellt (nur ausgewählte!)")
        
        # Zeit geben zum Wechseln zu Pro Tools
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
        import time
        time.sleep(3)
        
        # Korrekte Anzahl verwenden (nur ausgewählte Spuren)
        create_tracks_correct(track_count)
        
        return jsonify({
            'success': f'✅ {track_count} Spuren sollten erstellt sein! (Siehe Terminal für Details)'
        })
        
    except Exception as e:
        print(f"❌ Browser-Button Fehler: {e}")
        return jsonify({'error': f'Pro Tools Fehler: {e}'})

@app.route('/protools_name_tracks', methods=['POST'])
def protools_name_tracks():
    """Pro Tools: Spuren benennen - Browser-Button Version"""
    try:
        print("🔥 Browser-Button 'Spuren benennen' geklickt!")
        
        # Ausgewählte IDs sammeln
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            print("❌ Keine Spuren ausgewählt")
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        # Namensformat-Optionen sammeln
        include_channel = 'include_channel' in request.form
        include_instrument = 'include_instrument' in request.form  
        include_microphone = 'include_microphone' in request.form
        
        print(f"🎯 {len(selected_ids)} ausgewählte Spuren werden benannt")
        print(f"📝 Format: Kanal={include_channel}, Instrument={include_instrument}, Mikrofon={include_microphone}")
        
        # Zeit geben zum Wechseln zu Pro Tools
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
        print("💡 Markieren Sie die erste Spur in Pro Tools!")
        import time
        time.sleep(3)
        
        # Korrekte Funktion verwenden
        name_tracks_correct(selected_ids, include_channel, include_instrument, include_microphone)
        
        return jsonify({
            'success': f'✅ {len(selected_ids)} Spuren sollten benannt sein! (Siehe Terminal für Details)'
        })
        
    except Exception as e:
        print(f"❌ Browser-Button Fehler: {e}")
        return jsonify({'error': f'Pro Tools Fehler: {e}'})

# ==================== REAPER FUNCTIONS ====================

@app.route('/reaper_create_tracks', methods=['POST'])
def reaper_create_tracks():
    """Reaper: Spuren erstellen - Browser-Button Version"""
    try:
        print("🎧 DEBUG: Browser-Button 'REAPER Spuren erstellen' geklickt!")
        
        # Ausgewählte IDs sammeln
        selected_ids = request.form.getlist('selected_tracks')
        track_count = len(selected_ids)
        
        if track_count == 0:
            print("❌ Keine Spuren ausgewählt")
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        print(f"🎧 Erstelle {track_count} Spuren in Reaper...")
        
        # Zeit geben zum Wechseln zu Reaper
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Reaper...")
        import time
        time.sleep(3)
        
        # Reaper Spuren erstellen
        create_tracks_reaper(track_count)
        
        return jsonify({
            'success': f'✅ {track_count} Reaper Spuren sollten erstellt sein! (Siehe Terminal für Details)'
        })
        
    except Exception as e:
        print(f"❌ Reaper Browser-Button Fehler: {e}")
        return jsonify({'error': f'Reaper Fehler: {e}'})

@app.route('/reaper_name_tracks', methods=['POST'])
def reaper_name_tracks():
    """Reaper: Spuren benennen - Browser-Button Version"""
    try:
        print("🎧 Browser-Button 'Reaper Spuren benennen' geklickt!")
        
        # Ausgewählte IDs sammeln
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            print("❌ Keine Spuren ausgewählt")
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        # Namensformat-Optionen für Reaper aus Browser-Formular übernehmen
        include_channel = 'include_channel' in request.form
        include_instrument = 'include_instrument' in request.form  
        include_microphone = 'include_microphone' in request.form
        
        # DEBUG: Ausgabe der empfangenen Form-Daten
        print(f"🔍 DEBUG Form-Daten: {dict(request.form)}")
        print(f"🔍 DEBUG Checkbox-Status: Kanal={include_channel}, Instrument={include_instrument}, Mikrofon={include_microphone}")
        
        print(f"🎧 {len(selected_ids)} ausgewählte Spuren werden in Reaper benannt")
        print(f"📝 Format: Kanal={include_channel}, Instrument={include_instrument}, Mikrofon={include_microphone}")
        print(f"📝 Reaper-Format: 'Kanal Instrument Mikrofon' mit Leerzeichen getrennt")
        
        # Zeit geben zum Wechseln zu Reaper
        print("⏳ 8 Sekunden Zeit zum Wechseln zu Reaper...")
        print("💡 1. Wechseln Sie zu Reaper")
        print("💡 2. Öffnen Sie den Namens-Dialog der ersten Spur (Doppelklick)")
        print("💡 3. Die App tippt dann automatisch alle Namen mit Tab-Navigation...")
        import time
        time.sleep(8)
        
        # Reaper Spuren benennen
        name_tracks_reaper(selected_ids, include_channel, include_instrument, include_microphone)
        
        return jsonify({
            'success': f'✅ {len(selected_ids)} Reaper Spuren sollten benannt sein! (Siehe Terminal für Details)'
        })
        
    except Exception as e:
        print(f"❌ Reaper Browser-Button Fehler: {e}")
        return jsonify({'error': f'Reaper Fehler: {e}'})

# HTML Template erstellen
def create_template():
    template_dir = os.path.join(os.path.dirname(__file__), 'templates')
    os.makedirs(template_dir, exist_ok=True)
    
    html_content = '''<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🎵 Pro Tools Track Namer</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { text-align: center; margin-bottom: 30px; }
        .upload-section { background: #e3f2fd; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
        .layout-options { margin: 10px 0; }
        .layout-options input { margin-right: 5px; }
        .data-table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        .data-table th, .data-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        .data-table th { background-color: #f2f2f2; font-weight: bold; }
        .data-table tr:nth-child(even) { background-color: #f9f9f9; }
        .trackname { font-family: 'Courier New', monospace; font-weight: bold; color: #2196F3; }
        .buttons { margin: 20px 0; text-align: center; }
        .buttons button { margin: 0 10px; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; font-size: 14px; }
        .btn-primary { background-color: #2196F3; color: white; }
        .btn-success { background-color: #4CAF50; color: white; }
        .btn-warning { background-color: #FF9800; color: white; }
        .btn-danger { background-color: #f44336; color: white; }
        .flash-messages { margin: 20px 0; }
        .flash-success { background: #d4edda; color: #155724; padding: 10px; border-radius: 5px; border: 1px solid #c3e6cb; }
        .flash-error { background: #f8d7da; color: #721c24; padding: 10px; border-radius: 5px; border: 1px solid #f5c6cb; }
        .upload-btn { background-color: #2196F3; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }
        input[type="file"] { margin: 10px 0; }
        .stats { background: #e8f5e9; padding: 10px; border-radius: 5px; margin: 10px 0; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎵 Pro Tools Track Namer</h1>
            <p>Web-basierte Version - läuft im Browser</p>
        </div>
        
        <!-- Flash Messages -->
        <div class="flash-messages">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="flash-{{ 'success' if category == 'success' else 'error' }}">
                            {{ message }}
                        </div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
        </div>
        
        <!-- Upload Section -->
        <div class="upload-section">
            <h2>📁 Excel-Datei laden</h2>
            <form method="post" action="/upload" enctype="multipart/form-data">
                <input type="file" name="excel_file" accept=".xlsx,.xls,.xlsm" required>
                
                <div class="layout-options">
                    <strong>Layout:</strong><br>
                    <input type="radio" name="layout" value="auto" checked> Automatisch erkennen<br>
                    <input type="radio" name="layout" value="layout_a"> Layout A (Instrument=C, Mikrofon=D)<br>
                    <input type="radio" name="layout" value="layout_b"> Layout B (Instrument=D, Mikrofon=E)
                </div>
                
                <button type="submit" class="upload-btn">Excel laden</button>
            </form>
        </div>
        
        {% if data %}
        <div class="stats">
            <strong>📊 Geladene Daten:</strong> {{ data|length }} Zeilen | 
            <strong>Layout:</strong> {{ current_layout }}
        </div>
        
        <!-- Selection Buttons -->
        <div class="buttons">
            <button class="btn-success" onclick="selectAll()">✅ Alle auswählen</button>
            <button class="btn-warning" onclick="selectNone()">❌ Alle abwählen</button>
        </div>
        
        <!-- Data Table -->
        <form method="post" action="/export">
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Auswahl</th>
                        <th>Kanal</th>
                        <th>Instrument</th>
                        <th>Mikrofon</th>
                        <th>Trackname</th>
                    </tr>
                </thead>
                <tbody>
                    {% for row in data %}
                    <tr>
                        <td><input type="checkbox" name="selected_tracks" value="{{ row.id }}" {% if row.selected %}checked{% endif %}></td>
                        <td>{{ row.kanal }}</td>
                        <td>{{ row.instrument }}</td>
                        <td>{{ row.mikrofon }}</td>
                        <td class="trackname">{{ row.trackname }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
            
            <!-- Namensformat-Auswahl -->
            <div style="background: #e8f5e9; padding: 15px; border-radius: 8px; margin: 20px 0; border: 1px solid #4caf50;">
                <h3>🏷️ Namensformat wählen</h3>
                <div style="display: flex; gap: 20px; flex-wrap: wrap;">
                    <label><input type="checkbox" name="include_channel" checked> Spurnummer (z.B. "1_")</label>
                    <label><input type="checkbox" name="include_instrument" checked> Instrument (z.B. "Kick")</label>  
                    <label><input type="checkbox" name="include_microphone" checked> Mikrofon (z.B. "SM57")</label>
                </div>
                <p style="margin-top: 10px; font-size: 12px; color: #666;">
                    <strong>Beispiel:</strong> <span id="name-preview">1_Kick_SM57</span>
                </p>
            </div>
            
            <div class="buttons">
                <button type="submit" class="btn-primary">📄 Tracknamen als Textdatei exportieren</button>
            </div>
        </form>
        
        <!-- Pro Tools Integration -->
        <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0; border: 1px solid #ffeaa7;">
            <h3>🎵 Pro Tools Integration</h3>
            <p><strong>Workflow:</strong> Erst Spuren erstellen, dann benennen</p>
            
            <div class="buttons">
                <button class="btn-success" onclick="protoolsCreateTracks()" id="createBtn">
                    🏗️ Ctrl+Alt+C: Spuren erstellen
                </button>
                <button class="btn-warning" onclick="protoolsNameTracks()" id="nameBtn">
                    🏷️ Ctrl+Alt+N: Spuren benennen
                </button>
                <button class="btn-primary" onclick="showManualInstructions()" id="manualBtn">
                    📋 Ctrl+Alt+H: Anleitung
                </button>
            </div>
            
            <div id="protools-status" style="margin-top: 10px; padding: 10px; border-radius: 5px; display: none;"></div>
        </div>
        
        <!-- Reaper Integration -->
        <div style="background: #e3f2fd; padding: 15px; border-radius: 8px; margin: 20px 0; border: 1px solid #90caf9;">
            <h3>🎧 Reaper Integration</h3>
            <p><strong>Workflow:</strong> Erst Spuren erstellen (Cmd+T), dann benennen (F2+Tab)</p>
            
            <div class="buttons">
                <button class="btn-success" onclick="reaperCreateTracks()" id="reaperCreateBtn">
                    🏗️ Reaper - Spuren erstellen
                </button>
                <button class="btn-warning" onclick="reaperNameTracks()" id="reaperNameBtn">
                    🏷️ Reaper - Spuren benennen
                </button>
            </div>
            
            <div id="reaper-status" style="margin-top: 10px; padding: 10px; border-radius: 5px; display: none;"></div>
        </div>
        
        {% else %}
        <div class="stats">
            <strong>ℹ️ Anleitung:</strong> Laden Sie eine Excel-Datei (.xlsm, .xlsx, .xls), um zu beginnen.
        </div>
        {% endif %}
    </div>
    
    <script>
        let lastClickedTrackCheckbox = null;

        function setupShiftRangeSelection() {
            const checkboxes = Array.from(document.querySelectorAll('input[name="selected_tracks"]'));
            checkboxes.forEach(cb => {
                cb.addEventListener('click', event => {
                    if (!event.shiftKey || !lastClickedTrackCheckbox) {
                        lastClickedTrackCheckbox = cb;
                        return;
                    }

                    const start = checkboxes.indexOf(lastClickedTrackCheckbox);
                    const end = checkboxes.indexOf(cb);
                    if (start === -1 || end === -1) {
                        lastClickedTrackCheckbox = cb;
                        return;
                    }

                    const shouldCheck = cb.checked;
                    const from = Math.min(start, end);
                    const to = Math.max(start, end);

                    for (let i = from; i <= to; i++) {
                        checkboxes[i].checked = shouldCheck;
                    }

                    lastClickedTrackCheckbox = cb;
                });
            });
        }

        function selectAll() {
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]');
            checkboxes.forEach(cb => cb.checked = true);
        }
        
        function selectNone() {
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]');
            checkboxes.forEach(cb => cb.checked = false);
        }
        
        function getSelectedTracks() {
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]:checked');
            return Array.from(checkboxes).map(cb => cb.value);
        }
        
        function showProtoolsStatus(message, isError = false) {
            const status = document.getElementById('protools-status');
            status.style.display = 'block';
            status.className = isError ? 'flash-error' : 'flash-success';
            status.textContent = message;
        }
        
        function protoolsCreateTracks() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showProtoolsStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showProtoolsStatus('⏳ Erstelle ' + selected.length + ' Spuren in Pro Tools...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/protools_create_tracks', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showProtoolsStatus(data.error, true);
                } else {
                    showProtoolsStatus(data.success);
                }
            })
            .catch(error => {
                showProtoolsStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function protoolsNameTracks() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showProtoolsStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showProtoolsStatus('⏳ Benenne ' + selected.length + ' Spuren in Pro Tools...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            // Namensformat-Checkboxen hinzufügen
            const includeChannel = document.querySelector('input[name="include_channel"]').checked;
            const includeInstrument = document.querySelector('input[name="include_instrument"]').checked;
            const includeMicrophone = document.querySelector('input[name="include_microphone"]').checked;
            
            if (includeChannel) formData.append('include_channel', 'on');
            if (includeInstrument) formData.append('include_instrument', 'on');
            if (includeMicrophone) formData.append('include_microphone', 'on');
            
            fetch('/protools_name_tracks', {
                method: 'POST', 
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showProtoolsStatus(data.error, true);
                } else {
                    showProtoolsStatus(data.success);
                }
            })
            .catch(error => {
                showProtoolsStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function showManualInstructions() {
            showProtoolsStatus('📋 Manuelle Anleitung im Terminal angezeigt - schauen Sie ins Terminal!');
            
            // Einfacher GET-Request um F10-Funktion auszulösen
            fetch('/manual_instructions', { method: 'GET' })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        showProtoolsStatus(data.success);
                    }
                });
        }

        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', setupShiftRangeSelection);
        } else {
            setupShiftRangeSelection();
        }
        
        // ==================== REAPER JAVASCRIPT FUNCTIONS ====================
        
        function reaperCreateTracks() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showReaperStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showReaperStatus('⏳ Erstelle ' + selected.length + ' Spuren in Reaper...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            // Note: Format-Optionen werden automatisch serverseitig gelesen
            
            fetch('/reaper_create_tracks', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showReaperStatus(data.error, true);
                } else {
                    showReaperStatus(data.success);
                }
            })
            .catch(error => {
                showReaperStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function reaperNameTracks() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showReaperStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showReaperStatus('⏳ Benenne ' + selected.length + ' Spuren in Reaper...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            // Checkbox-Status für Reaper übernehmen
            const includeChannel = document.querySelector('input[name="include_channel"]').checked;
            const includeInstrument = document.querySelector('input[name="include_instrument"]').checked;
            const includeMicrophone = document.querySelector('input[name="include_microphone"]').checked;
            
            if (includeChannel) formData.append('include_channel', 'on');
            if (includeInstrument) formData.append('include_instrument', 'on');
            if (includeMicrophone) formData.append('include_microphone', 'on');
            
            fetch('/reaper_name_tracks', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showReaperStatus(data.error, true);
                } else {
                    showReaperStatus(data.success);
                }
            })
            .catch(error => {
                showReaperStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function showReaperStatus(message, isError = false) {
            const statusDiv = document.getElementById('reaper-status');
            statusDiv.style.display = 'block';
            statusDiv.textContent = message;
            statusDiv.className = isError ? 'status-error' : 'status-success';
            
            if (!isError) {
                setTimeout(() => {
                    statusDiv.style.display = 'none';
                }, 5000);
            }
        }
    </script>
</body>
</html>'''
    
    with open(os.path.join(template_dir, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(html_content)

def setup_global_hotkeys():
    """Globale F8/F9 Hotkeys einrichten"""
    try:
        from pynput import keyboard
        
        def on_hotkey_f8():
            """Ctrl+Alt+C: Spuren erstellen"""
            print("🔥 Ctrl+Alt+C gedrückt - Erstelle Spuren...")
            create_tracks_from_selection()
        
        def on_hotkey_f9():
            """Ctrl+Alt+N: Spuren benennen"""
            print("🔥 Ctrl+Alt+N gedrückt - Benenne Spuren...")
            name_tracks_from_selection()
            
        def on_hotkey_f7():
            """Ctrl+Alt+H: Manuelle Anleitung anzeigen"""
            print("🔥 Ctrl+Alt+H gedrückt - Manuelle Anleitung:")
            show_manual_instructions()
        
        # Globale Hotkeys registrieren - Verwende Ctrl+Alt+Kombinationen (sicher!)
        keyboard.GlobalHotKeys({
            '<ctrl>+<alt>+c': on_hotkey_f8,  # Ctrl+Alt+C für Create tracks
            '<ctrl>+<alt>+n': on_hotkey_f9,  # Ctrl+Alt+N für Name tracks  
            '<ctrl>+<alt>+h': on_hotkey_f7,  # Ctrl+Alt+H für Help/Manual
        }).start()
        
        print("🎹 Globale Hotkeys registriert (macOS & Pro Tools kompatibel):")
        print("   Ctrl+Alt+H (Anleitung), Ctrl+Alt+C (Spuren erstellen), Ctrl+Alt+N (Spuren benennen)")
        
    except Exception as e:
        print(f"⚠️ Hotkey-Registrierung fehlgeschlagen: {e}")

def create_tracks_correct(track_count):
    """Spuren erstellen - Korrekte Anzahl basierend auf Auswahl"""
    try:
        from pynput.keyboard import Key, Controller as KeyboardController
        import time
        
        keyboard = KeyboardController()
        
        print(f"🏗️ Erstelle {track_count} Spuren in Pro Tools...")
        print("🎯 Schnelle Methode: Dialog öffnen + Anzahl sofort eingeben")
        
        # Pro Tools New Track Dialog öffnen: Cmd+Shift+N
        keyboard.press(Key.cmd)
        keyboard.press(Key.shift)
        keyboard.press('n')
        keyboard.release('n')
        keyboard.release(Key.shift)
        keyboard.release(Key.cmd)
        
        # Minimal warten und sofort Anzahl eingeben
        time.sleep(0.6)  # Noch kürzer
        
        # Anzahl sofort eingeben (überschreibt Standard-Wert)
        track_str = str(track_count)
        print(f"   Eingabe sofort: {track_str}")
        
        # Alles markieren und überschreiben
        keyboard.press(Key.cmd)
        keyboard.press('a')
        keyboard.release('a')
        keyboard.release(Key.cmd)
        
        # Anzahl eingeben
        keyboard.type(track_str)
        
        # Sofort Enter drücken
        time.sleep(0.1)  # Sehr kurz
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        
        print(f"✅ {track_count} Spuren sollten sofort erstellt sein!")
        print("💡 Optimiert: Minimale Wartezeit, sofortige Eingabe")
        
    except Exception as e:
        print(f"❌ Fehler beim Erstellen der Spuren: {e}")

def create_tracks_from_selection():
    """Legacy-Funktion für Hotkeys - verwendet korrekte Anzahl"""
    if not current_data:
        print("❌ Keine Daten geladen")
        return
    
    # Nur ausgewählte Tracks zählen
    selected_count = sum(1 for item in current_data if item.get('selected', True))
    
    if selected_count == 0:
        print("❌ Keine Spuren ausgewählt")
        return
        
    create_tracks_correct(selected_count)

def name_tracks_correct(selected_ids, include_channel, include_instrument, include_microphone):
    """Spuren benennen - Korrekt mit ausgewählten IDs und Namensformat"""
    try:
        if not current_data:
            print("❌ Keine Daten geladen")
            return
        
        # Nur die wirklich ausgewählten Items holen
        selected_items = [item for item in current_data if item['id'] in selected_ids]
        
        if not selected_items:
            print("❌ Keine Spuren ausgewählt")
            return
        
        from pynput.keyboard import Key, Controller as KeyboardController
        import time
        
        keyboard = KeyboardController()
        
        print(f"🏷️ Benenne {len(selected_items)} Spuren in Pro Tools...")
        print("📋 ANLEITUNG:")
        print("   1. Öffne MANUELL den Namensgebungs-Dialog der ersten Spur")
        print("   2. Die App übernimmt dann automatisch!")
        print("⏳ 2 Sekunden Zeit zum Öffnen des ersten Dialogs...")
        time.sleep(2)
        
        # Jede Spur benennen mit korrektem Format
        for i, item in enumerate(selected_items):
            # Debug-Output für die ersten paar Spuren
            if i < 5:
                print(f"  🔍 Debug Spur {i+1}: Kanal='{item['kanal']}', Instrument='{item['instrument']}', Mikrofon='{item['mikrofon']}'")
            
            # Namen nach Auswahl zusammensetzen
            name_parts = []
            if include_channel and item['kanal']:
                name_parts.append(str(item['kanal']))
            if include_instrument and item['instrument']:
                name_parts.append(str(item['instrument']))
            if include_microphone and item['mikrofon']:
                name_parts.append(str(item['mikrofon']))
            
            # Namen zusammenbauen
            if name_parts:
                name = "_".join(name_parts)
            else:
                # Fallback: Wenn alle Felder leer sind oder nichts ausgewählt
                if item['instrument']:
                    name = str(item['instrument'])
                else:
                    name = f"Track_{i+1}"
            
            # Extra Debug für problematische Spuren
            if i >= 34 and i <= 36:
                print(f"  🚨 Problem-Debug Spur {i+1}: Resultat='{name}', Parts={name_parts}, Item={item}")
            
            print(f"  🎯 Spur {i+1}/{len(selected_items)}: {name}")
            
            # Namen eingeben (überschreibt aktuellen Inhalt)
            keyboard.press(Key.cmd)
            keyboard.press('a')  # Alles markieren
            keyboard.release('a')
            keyboard.release(Key.cmd)
            time.sleep(0.03)  # Ultra kurz
            
            # Neuen Namen eingeben
            keyboard.type(name)
            time.sleep(0.05)  # Ultra kurz
            
            # Zur nächsten Spur wechseln (außer bei der letzten)
            if i < len(selected_items) - 1:
                keyboard.press(Key.cmd)
                keyboard.press(Key.right)  # Cmd+→ zur nächsten Spur
                keyboard.release(Key.right)
                keyboard.release(Key.cmd)
                time.sleep(0.08)  # Ultra kurz
        
        # Enter am Ende drücken um letzte Spur zu bestätigen
        print("  🎯 Bestätige letzte Spur mit Enter...")
        time.sleep(0.1)
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        
        print(f"✅ Alle {len(selected_items)} Spuren benannt und bestätigt!")
        print("🎉 Namensgebung komplett abgeschlossen!")
        
    except Exception as e:
        print(f"❌ Fehler beim Benennen der Spuren: {e}")

def name_tracks_from_selection():
    """Legacy-Funktion für Hotkeys - verwendet Standard-Format"""
    if not current_data:
        print("❌ Keine Daten geladen")
        return
    
    # Alle ausgewählten IDs sammeln
    selected_ids = [item['id'] for item in current_data if item.get('selected', True)]
    
    # Standard-Format verwenden (alle Komponenten)
    name_tracks_correct(selected_ids, True, True, True)

def show_manual_instructions():
    """Zeige manuelle Anleitung für Pro Tools"""
    if not current_data:
        print("❌ Keine Daten geladen")
        return
    
    selected_count = sum(1 for item in current_data if item.get('selected', True))
    names = [item['trackname'] for item in current_data if item.get('selected', True)]
    
    print("\n" + "="*60)
    print("📋 MANUELLE PRO TOOLS ANLEITUNG")
    print("="*60)
    print(f"🎯 Zu erstellende Spuren: {selected_count}")
    print("\n🏗️ SPUREN ERSTELLEN:")
    print("1. Gehe zu: Track → New...")
    print("2. Audio Tracks: Mono")
    print(f"3. Anzahl: {selected_count}")
    print("4. Klicke 'Create'")
    print("\n🏷️ SPUREN BENENNEN:")
    for i, name in enumerate(names, 1):
        print(f"   Spur {i}: {name}")
    print("\n💡 Tipp: Erste Spur markieren, dann F9 für automatische Benennung")
    print("="*60 + "\n")

def open_browser():
    """Browser nach kurzer Wartezeit öffnen"""
    time.sleep(1.5)
    webbrowser.open('http://127.0.0.1:5003')

# ==================== REAPER BACKEND FUNCTIONS ====================

def create_tracks_reaper(track_count):
    """Erstelle Spuren in Reaper mit Cmd+T Shortcut"""
    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        print(f"🎧 Erstelle {track_count} Spuren in Reaper mit Cmd+T...")
        
        for i in range(track_count):
            print(f"  🔧 Erstelle Spur {i+1}/{track_count}")
            # Cmd+T für neue Spur in Reaper
            keyboard.press(Key.cmd)
            keyboard.press('t')
            keyboard.release('t')
            keyboard.release(Key.cmd)
            time.sleep(0.05)  # Sehr schnelle Spur-Erstellung
            
        print(f"✅ {track_count} Reaper Spuren erstellt!")
        
    except Exception as e:
        print(f"❌ Reaper Spur-Erstellung Fehler: {e}")

def name_tracks_reaper(selected_ids, include_channel, include_instrument, include_microphone):
    """Benenne Spuren in Reaper"""
    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        # Daten der ausgewählten Spuren sammeln
        selected_items = [current_data[i] for i in selected_ids]
        
        print(f"🎧 Benenne {len(selected_items)} Spuren in Reaper...")
        print("🎹 Layout: Name direkt eingeben → Tab zur nächsten Spur")
        
        for i, item in enumerate(selected_items):
            # Name zusammenstellen (Leerzeichen statt Underscore für Reaper)
            name_parts = []
            if include_channel and item.get('kanal'):
                name_parts.append(str(item['kanal']))
            if include_instrument and item.get('instrument'):
                name_parts.append(str(item['instrument']))
            if include_microphone and item.get('mikrofon'):
                name_parts.append(str(item['mikrofon']))
            
            # Namen zusammenbauen (mit Leerzeichen statt Underscore)
            if name_parts:
                name = " ".join(name_parts)
            else:
                # Fallback: Wenn alle Felder leer sind oder nichts ausgewählt
                if item.get('instrument'):
                    name = str(item['instrument'])
                else:
                    name = f"Track_{i+1}"
            
            print(f"  🎧 Spur {i+1}/{len(selected_items)}: '{name}'")
            
            # KEIN F2! Direkt Namen tippen (Namens-Dialog ist bereits offen)
            keyboard.type(name)
            time.sleep(0.05)  # Sehr schnelles Tippen
            
            # Tab zur nächsten Spur (außer bei der letzten)
            if i < len(selected_items) - 1:
                print("    🔄 Tab zur nächsten Spur...")
                keyboard.press(Key.tab)
                keyboard.release(Key.tab)
                time.sleep(0.05)  # Sehr schneller Tab-Wechsel
        
        print(f"✅ {len(selected_items)} Reaper Spuren benannt!")
        
    except Exception as e:
        print(f"❌ Reaper Spur-Benennung Fehler: {e}")

if __name__ == '__main__':
    # Template erstellen
    create_template()
    
    # Globale Hotkeys in separatem Thread starten
    threading.Thread(target=setup_global_hotkeys, daemon=True).start()
    
    # Browser in separatem Thread öffnen
    threading.Thread(target=open_browser, daemon=True).start()
    
    print("🎵 Multi-DAW Track Namer (Pro Tools + Reaper) startet...")
    print("🌐 Öffne Browser auf: http://127.0.0.1:5003")
    print("🎹 Hotkeys: Ctrl+Alt+H (Anleitung), Ctrl+Alt+C (Erstellen), Ctrl+Alt+N (Benennen)")
    print("⚠️  Zum Beenden: Strg+C drücken")
    
    # Flask-App starten
    app.run(debug=False, host='127.0.0.1', port=5003, use_reloader=False)