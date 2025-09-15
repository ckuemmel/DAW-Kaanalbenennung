#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

# Erstelle Test-Excel mit Layout A (Instrument in Spalte C)
def create_layout_a_test():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Micplan"
    
    # Header in Zeile 6
    sheet.cell(row=6, column=2, value="Kanal")      # B
    sheet.cell(row=6, column=3, value="Instrument") # C <- Instrument hier!
    sheet.cell(row=6, column=4, value="Mikrofon")   # D <- Mikrofon hier!
    sheet.cell(row=6, column=5, value="Stativ")     # E
    
    # Test-Daten ab Zeile 8
    test_data = [
        (1, "Flöte 1", "KM184"),
        (2, "Flöte 2", "KM184"), 
        (3, "Oboe", "KM184"),
        (4, "Klarinette 1", "SM57"),
        (5, "Klarinette 2", "SM57"),
        (6, "Trompete 1", "SM58"),
        (7, "Trompete 2", "SM58"),
    ]
    
    for i, (kanal, instrument, mikrofon) in enumerate(test_data, start=8):
        sheet.cell(row=i, column=2, value=kanal)       # B: Kanal
        sheet.cell(row=i, column=3, value=instrument)  # C: Instrument  
        sheet.cell(row=i, column=4, value=mikrofon)    # D: Mikrofon
    
    workbook.save("Data/test_layout_a.xlsx")
    print("Layout A Test-Datei erstellt: Data/test_layout_a.xlsx")

if __name__ == "__main__":
    create_layout_a_test()