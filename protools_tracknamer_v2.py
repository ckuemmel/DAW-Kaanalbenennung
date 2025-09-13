import argparse
import csv
import sys
import time
import threading
from pathlib import Path

try:
    from pynput import keyboard
    from pynput.keyboard import Key, Controller
except ImportError:
    print("Fehler: Das Paket 'pynput' ist nicht installiert.")
    print("Installiere es mit: pip install pynput")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Fehler: Das Paket 'openpyxl' ist nicht installiert.")
    print("Installiere es mit: pip install openpyxl")
    sys.exit(1)

class NameTyper:
    def __init__(self):
        self.keyboard = Controller()
        self.names = []
        self.idx = 0
        self.running = True
    
    def load_excel(self, filepath, name_col, channel_col, mic_col, header_row):
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Namen einlesen
        for r in range(header_row, ws.max_row + 1):
            name_cell = ws.cell(row=r, column=name_col + 1)
            channel_cell = ws.cell(row=r, column=channel_col + 1)
            mic_cell = ws.cell(row=r, column=mic_col + 1)
            
            if name_cell.value:
                name = str(name_cell.value).strip()
                channel = str(channel_cell.value).strip() if channel_cell.value else None
                mic = str(mic_cell.value).strip() if mic_cell.value else None
                if name:  # Nur hinzufügen wenn ein Name vorhanden ist
                    self.names.append((name, channel, mic))

        print(f"DEBUG: {len(self.names)} Namen geladen")
        print("DEBUG: Erste 3 Namen zur Kontrolle:")
        for i, (name, channel, mic) in enumerate(self.names[:3]):
            print(f"DEBUG: {i + 1}: {channel}_{name}_{mic}")
    
    def type_next(self, delay):
        if self.idx >= len(self.names):
            print("Fertig: Alle Namen wurden verwendet.")
            return False
            
        name, channel, mic = self.names[self.idx]
        self.idx += 1
        
        # Formatiere den Namen mit Unterstrichen
        formatted_name = name
        if channel:
            formatted_name = f"{channel}_{formatted_name}"
        if mic:
            formatted_name = f"{formatted_name}_{mic}"
        
        # Warte und tippe
        time.sleep(delay)
        self.keyboard.type(formatted_name)
        print(f"Getippt: {formatted_name}")
        return True
    
    def send_next_track(self, delay):
        """Sendet einen einzelnen Strg+Rechts Befehl"""
        time.sleep(delay)  # Warte nach dem Tippen
        print("DEBUG: Sende Strg+Rechts")
        
        # Stelle sicher, dass keine Taste hängenbleibt
        self.keyboard.release(Key.ctrl)
        self.keyboard.release(Key.right)
        time.sleep(0.2)
        
        # Sende Strg+Rechts
        self.keyboard.press(Key.ctrl)
        time.sleep(0.2)
        self.keyboard.press(Key.right)
        time.sleep(0.2)
        self.keyboard.release(Key.right)
        time.sleep(0.2)
        self.keyboard.release(Key.ctrl)
        time.sleep(delay)  # Warte nach dem Spurwechsel
    
    def on_key_press(self, key):
        try:
            if key == keyboard.Key.f8:
                if self.type_next(0.2):
                    self.send_next_track(0.2)
            elif key == keyboard.Key.f7 and self.idx > 0:
                self.idx -= 1
                print(f"Zurück zu Index {self.idx}")
            elif key == keyboard.Key.esc:
                print("Beendet.")
                self.running = False
                return False
        except Exception as e:
            print(f"Fehler: {e}")
    
    def run(self):
        with keyboard.Listener(on_press=self.on_key_press) as listener:
            listener.join()

def main():
    parser = argparse.ArgumentParser(description="ProTools Track Namer")
    parser.add_argument("excel_file", help="Pfad zur Excel-Datei")
    parser.add_argument("--spalte", type=int, default=3, help="Spalte mit Namen (0-basiert)")
    parser.add_argument("--kanal-spalte", type=int, default=1, help="Spalte mit Kanälen")
    parser.add_argument("--mikrofon-spalte", type=int, default=4, help="Spalte mit Mikrofonen")
    parser.add_argument("--header-row", type=int, default=8, help="Headerzeile (1-basiert)")
    args = parser.parse_args()
    
    typer = NameTyper()
    typer.load_excel(
        args.excel_file,
        args.spalte,
        args.kanal_spalte,
        args.mikrofon_spalte,
        args.header_row
    )
    
    print("\nBereit! Anleitung:")
    print("1) Erstelle/Ordne die Spuren in Pro Tools")
    print("2) Doppelklicke die erste Spurbezeichnung")
    print("3) Drücke F8 für jeden Namen (F7 für zurück, ESC zum Beenden)")
    
    typer.run()

if __name__ == "__main__":
    main()