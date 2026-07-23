#!/usr/bin/env python3
"""
Einfache, funktionsfähige Version - Multi-DAW Track Namer
"""

import time
import openpyxl
import tempfile
import threading
import webbrowser
from flask import Flask, render_template_string, request, redirect, url_for, flash, jsonify

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'

# Globale Variablen
current_data = []

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE, data=current_data)

@app.route('/upload', methods=['POST'])
def upload_file():
    global current_data
    if 'file' not in request.files or request.files['file'].filename == '':
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('index'))
    
    file = request.files['file']
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            file.save(tmp_file.name)
            workbook = openpyxl.load_workbook(tmp_file.name)
            sheet = workbook.active
            
            current_data = []
            for row_num in range(7, min(sheet.max_row + 1, 200)):
                kanal = sheet.cell(row=row_num, column=2).value
                instrument = sheet.cell(row=row_num, column=3).value  
                mikrofon = sheet.cell(row=row_num, column=4).value
                
                if instrument:
                    trackname = f"{kanal} {instrument} {mikrofon}" if mikrofon else f"{kanal} {instrument}"
                    current_data.append({
                        'id': len(current_data),
                        'kanal': kanal,
                        'instrument': instrument,
                        'mikrofon': mikrofon or '',
                        'trackname': trackname,
                        'selected': True
                    })
                    
        flash(f'✅ {len(current_data)} Spuren geladen', 'success')
    except Exception as e:
        flash(f'❌ Fehler: {e}', 'error')
    
    return redirect(url_for('index'))

@app.route('/create_tracks_reaper', methods=['POST'])
def create_tracks_reaper():
    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        selected_ids = request.form.getlist('selected_tracks')
        track_count = len(selected_ids)
        
        if track_count == 0:
            return jsonify({'error': 'Keine Spuren ausgewählt'})
        
        print(f"🎧 Erstelle {track_count} Spuren in Reaper...")
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Reaper...")
        time.sleep(3)
        
        for i in range(track_count):
            print(f"🔧 Erstelle Spur {i+1}/{track_count}")
            keyboard.press(Key.cmd)
            keyboard.press('t')
            keyboard.release('t')
            keyboard.release(Key.cmd)
            time.sleep(0.3)
            
        return jsonify({'success': f'✅ {track_count} Reaper Spuren erstellt!'})
        
    except Exception as e:
        return jsonify({'error': f'Reaper Fehler: {e}'})

@app.route('/name_tracks_reaper', methods=['POST'])
def name_tracks_reaper():
    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Keine Spuren ausgewählt'})
        
        print(f"🏷️ Benenne {len(selected_ids)} Spuren in Reaper...")
        print("⏳ WARTEN AUF SIE:")
        print("💡 1. Wechseln Sie zu Reaper")
        print("💡 2. Klicken Sie in das Namensfeld der ersten Spur") 
        print("💡 3. Die App wartet 8 Sekunden, dann übernimmt sie...")
        time.sleep(8)
        
        for i, track_id in enumerate(selected_ids):
            track_data = current_data[track_id]
            trackname = track_data['trackname']
            
            print(f"🏷️ JETZT: Schreibe '{trackname}' in Spur {i+1}")
            
            # F2 drücken um Namen zu editieren (falls nötig)
            keyboard.press(Key.f2)
            keyboard.release(Key.f2)
            time.sleep(0.1)
            
            # Alten Text löschen
            keyboard.press(Key.cmd)
            keyboard.press('a')
            keyboard.release('a')
            keyboard.release(Key.cmd)
            time.sleep(0.1)
            
            # Neuen Namen tippen
            keyboard.type(trackname)
            time.sleep(0.2)
            
            # Tab zur nächsten Spur
            if i < len(selected_ids) - 1:
                print("🔄 Tab zur nächsten Spur...")
                keyboard.press(Key.tab)
                keyboard.release(Key.tab)
                time.sleep(0.3)
        
        print(f"✅ FERTIG: {len(selected_ids)} Namen eingegeben")
        return jsonify({'success': f'✅ {len(selected_ids)} Reaper Spuren benannt!'})
        
    except Exception as e:
        return jsonify({'error': f'Reaper Fehler: {e}'})

# Identische Funktionen für ProTools
@app.route('/create_tracks_protools', methods=['POST'])
def create_tracks_protools():
    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        selected_ids = request.form.getlist('selected_tracks')
        track_count = len(selected_ids)
        
        if track_count == 0:
            return jsonify({'error': 'Keine Spuren ausgewählt'})
        
        print(f"🎵 Erstelle {track_count} Spuren in Pro Tools...")
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
        time.sleep(3)
        
        for i in range(track_count):
            print(f"🔧 Erstelle Spur {i+1}/{track_count}")
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(0.3)
            
        return jsonify({'success': f'✅ {track_count} Pro Tools Spuren erstellt!'})
        
    except Exception as e:
        return jsonify({'error': f'Pro Tools Fehler: {e}'})

@app.route('/name_tracks_protools', methods=['POST'])
def name_tracks_protools():
    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Keine Spuren ausgewählt'})
        
        print(f"🏷️ Benenne {len(selected_ids)} Spuren in Pro Tools...")
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
        print("💡 Markieren Sie die erste Spur in Pro Tools!")
        time.sleep(3)
        
        for i, track_id in enumerate(selected_ids):
            track_data = current_data[track_id]
            trackname = track_data['trackname']
            
            print(f"🏷️ Spur {i+1}: '{trackname}'")
            
            keyboard.type(trackname)
            time.sleep(0.2)
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(0.1)
            keyboard.press(Key.cmd)
            keyboard.press(Key.right)
            keyboard.release(Key.right)
            keyboard.release(Key.cmd)
            time.sleep(0.2)
        
        return jsonify({'success': f'✅ {len(selected_ids)} Pro Tools Spuren benannt!'})
        
    except Exception as e:
        return jsonify({'error': f'Pro Tools Fehler: {e}'})

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Multi-DAW Track Namer</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .container { max-width: 1200px; margin: 0 auto; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .buttons { margin: 20px 0; text-align: center; }
        .buttons button { margin: 10px; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }
        .btn-success { background: #28a745; color: white; }
        .btn-primary { background: #007bff; color: white; }
        .status-box { padding: 10px; margin: 10px 0; border-radius: 5px; background: #d4edda; color: #155724; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🎵 Multi-DAW Track Namer</h1>
        
        <!-- Excel Upload -->
        <form method="POST" action="/upload" enctype="multipart/form-data">
            <p>Excel-Datei auswählen:</p>
            <input type="file" name="file" accept=".xlsx,.xls" required>
            <button type="submit">📂 Datei hochladen</button>
        </form>
        
        {% if data %}
        <p>📊 <strong>{{ data|length }}</strong> Spuren geladen | <span id="selected-count">{{ data|length }}</span> ausgewählt</p>
        
        <!-- Auswahl-Buttons -->
        <div class="buttons">
            <button class="btn-primary" onclick="selectAll()">✅ Alle auswählen</button>
            <button class="btn-primary" onclick="selectNone()">❌ Alle abwählen</button>
        </div>
        
        <!-- Track-Tabelle -->
        <table>
            <thead>
                <tr><th>✓</th><th>Kanal</th><th>Instrument</th><th>Mikrofon</th><th>Track-Name</th></tr>
            </thead>
            <tbody>
                {% for item in data %}
                <tr>
                    <td><input type="checkbox" name="selected_tracks" value="{{ item.id }}" checked></td>
                    <td>{{ item.kanal }}</td>
                    <td>{{ item.instrument }}</td>
                    <td>{{ item.mikrofon }}</td>
                    <td style="font-weight: bold; color: #2196F3;">{{ item.trackname }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        
        <!-- DAW Buttons -->
        <div class="status-box" id="status-box">Bereit für Pro Tools und Reaper...</div>
        
        <h3>🎹 Pro Tools</h3>
        <div class="buttons">
            <button class="btn-success" onclick="createTracksProtools()">🏗️ Pro Tools - Spuren erstellen</button>
            <button class="btn-success" onclick="nameTracksProtools()">🏷️ Pro Tools - Spuren benennen</button>
        </div>
        
        <h3>🎧 Reaper</h3>
        <div class="buttons">
            <button class="btn-success" onclick="createTracksReaper()">🏗️ Reaper - Spuren erstellen</button>
            <button class="btn-success" onclick="nameTracksReaper()">🏷️ Reaper - Spuren benennen</button>
        </div>
        {% endif %}
    </div>

    <script>
        function updateSelectedCount() {
            const checked = document.querySelectorAll('input[name="selected_tracks"]:checked');
            document.getElementById('selected-count').textContent = checked.length;
        }
        
        function selectAll() {
            document.querySelectorAll('input[name="selected_tracks"]').forEach(cb => cb.checked = true);
            updateSelectedCount();
        }
        
        function selectNone() {
            document.querySelectorAll('input[name="selected_tracks"]').forEach(cb => cb.checked = false);
            updateSelectedCount();
        }
        
        function getSelectedTracks() {
            const checked = document.querySelectorAll('input[name="selected_tracks"]:checked');
            return Array.from(checked).map(cb => cb.value);
        }
        
        function showStatus(message, isError = false) {
            const statusBox = document.getElementById('status-box');
            statusBox.textContent = message;
            statusBox.style.backgroundColor = isError ? '#f8d7da' : '#d4edda';
            statusBox.style.color = isError ? '#721c24' : '#155724';
        }
        
        function createTracksReaper() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            showStatus('⏳ Erstelle ' + selected.length + ' Spuren in Reaper...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/create_tracks_reaper', {
                method: 'POST',
                body: formData
            }).then(response => response.json()).then(data => {
                showStatus(data.error || data.success, !!data.error);
            });
        }
        
        function nameTracksReaper() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            showStatus('⏳ Benenne ' + selected.length + ' Spuren in Reaper...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/name_tracks_reaper', {
                method: 'POST',
                body: formData
            }).then(response => response.json()).then(data => {
                showStatus(data.error || data.success, !!data.error);
            });
        }
        
        function createTracksProtools() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            showStatus('⏳ Erstelle ' + selected.length + ' Spuren in Pro Tools...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/create_tracks_protools', {
                method: 'POST',
                body: formData
            }).then(response => response.json()).then(data => {
                showStatus(data.error || data.success, !!data.error);
            });
        }
        
        function nameTracksProtools() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            showStatus('⏳ Benenne ' + selected.length + ' Spuren in Pro Tools...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/name_tracks_protools', {
                method: 'POST',
                body: formData
            }).then(response => response.json()).then(data => {
                showStatus(data.error || data.success, !!data.error);
            });
        }
        
        // Event-Listener für Live-Updates
        document.querySelectorAll('input[name="selected_tracks"]').forEach(cb => {
            cb.addEventListener('change', updateSelectedCount);
        });
    </script>
</body>
</html>
'''

def open_browser():
    time.sleep(1.5)
    webbrowser.open('http://127.0.0.1:5002')

if __name__ == "__main__":
    print("🎵 EINFACHER Multi-DAW Track Namer startet...")
    print("🌐 Öffne Browser auf: http://127.0.0.1:5002")
    print("⚠️  Zum Beenden: Strg+C drücken")
    
    threading.Thread(target=open_browser, daemon=True).start()
    
    try:
        app.run(host='127.0.0.1', port=5002, debug=False)
    except KeyboardInterrupt:
        print("\n🛑 Multi-DAW Track Namer beendet")