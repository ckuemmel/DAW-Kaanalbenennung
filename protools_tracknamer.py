import argparse
import csv
import sys
import time
from pathlib import Path
import threading

try:
    from pynput import keyboard
    from pynput.keyboard import Key, Controller
except Exception as e:
    print("Fehler: Das Paket 'pynput' ist nicht installiert oder konnte nicht geladen werden.")
    print("Installiere es mit: pip install pynput")
    sys.exit(1)

# Import Konfiguration
try:
    from config import REPO_ROOT, DATA_DIR, is_path_in_repo
except ImportError:
    # config.py ist optional für dieses Skript
    pass

def validate_path(path: Path) -> Path:
    """Validiert den Pfad und stellt sicher, dass er im Data-Verzeichnis liegt"""
    try:
        abs_path = path.resolve()
        if not DATA_DIR in abs_path.parents:
            print(f"Fehler: Die Datei {path} liegt außerhalb des Data-Verzeichnisses.")
            print(f"Bitte lege die Datei im {DATA_DIR} Verzeichnis ab.")
            sys.exit(1)
        if not abs_path.exists():
            print(f"Fehler: Die Datei {path} existiert nicht.")
            sys.exit(1)
        return abs_path
    except Exception as e:
        print(f"Fehler beim Zugriff auf die Datei {path}: {e}")
        sys.exit(1)


def find_column_index(header: list[str], column_name: str | None) -> int | None:
    if column_name is None:
        return None
    # Wenn column_name eine Zahl ist, als 0-basierten Index interpretieren
    try:
        if column_name.isdigit():
            idx = int(column_name)
            if 0 <= idx < len(header):
                return idx
    except (ValueError, AttributeError):
        pass
    # Ansonsten als Spaltenname behandeln
    try:
        return next(i for i, h in enumerate(header) if h.lower() == column_name.strip().lower())
    except StopIteration:
        print(f"Spaltenname '{column_name}' nicht gefunden. Verfügbare Spalten:")
        print(", ".join(header))
        sys.exit(1)

def read_names_from_csv(
    path: Path,
    column: int | None = 0,
    column_name: str | None = None,
    channel_column: str | None = None,
    mic_column: str | None = None,
    skip_header: bool = False
) -> list[tuple[str, str | None, str | None]]:
    names: list[tuple[str, str | None, str | None]] = []
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        header_checked = False
        header: list[str] | None = None
        channel_idx: int | None = None
        mic_idx: int | None = None
        
        for idx, row in enumerate(reader):
            if not header_checked:
                # Ersten Datensatz als möglichen Header auswerten
                header_checked = True
                if column_name is not None or channel_column is not None or mic_column is not None:
                    header = [str(x).strip() for x in row]
                    if column_name is not None:
                        column = find_column_index(header, column_name)
                    if channel_column is not None:
                        channel_idx = find_column_index(header, channel_column)
                    if mic_column is not None:
                        mic_idx = find_column_index(header, mic_column)
                    # Wenn wir nach Namen gemappt haben, ist dies die Headerzeile → überspringen
                    continue
                # Falls kein Spaltenname genutzt wird, kann optional die erste Zeile als Header übersprungen werden
                if skip_header:
                    continue
            if not row:
                continue
            assert column is not None
            if column < len(row):
                name = str(row[column]).strip()
                if name:
                    channel = str(row[channel_idx]).strip() if channel_idx is not None and channel_idx < len(row) else None
                    mic = str(row[mic_idx]).strip() if mic_idx is not None and mic_idx < len(row) else None
                    names.append((name, channel, mic))
    return names


def read_names_from_xlsx(
    path: Path,
    column: int | None = 0,
    column_name: str | None = None,
    channel_column: str | None = None,
    mic_column: str | None = None,
    skip_header: bool = False,
    sheet: str | None = None,
    sheet_index: int | None = None,
    header_row: int | None = None,
) -> list[tuple[str, str | None, str | None]]:
    # Debugging
    print("\nDEBUG: Lade Excel-Datei...")
    print(f"DEBUG: Pfad: {path}")
    print(f"DEBUG: Instrumente aus Spalte D (Index {column})")
    print(f"DEBUG: Kanalnummern aus Spalte A: {channel_column is not None}")
    print(f"DEBUG: Mikrofontypen aus Spalte E: {mic_column is not None}")
    print(f"DEBUG: Header-Zeile: {header_row}")
    
    try:
        import openpyxl  # type: ignore
    except Exception:
        print("Fehler: 'openpyxl' ist für XLSX erforderlich. Installiere es mit: pip install openpyxl")
        sys.exit(1)
        
    def has_content(row, ws, columns) -> bool:
        """Prüft ob eine Zeile in den relevanten Spalten Inhalte hat"""
        # Überprüfe die Hauptspalte (Name/Instrument)
        main_col = columns[0] if columns else None
        if main_col is not None:
            cell = ws.cell(row=row, column=main_col + 1)
            if cell.value is None or not str(cell.value).strip():
                return False  # Wenn kein Name vorhanden ist, hat die Zeile keinen Inhalt
            
        # Überprüfe die anderen Spalten nur wenn sie existieren
        for col in columns[1:]:  # Skip die Hauptspalte
            if col is None:
                continue
            cell = ws.cell(row=row, column=col + 1)
            if cell.value is not None and str(cell.value).strip():
                return True  # Mindestens eine andere Spalte hat Inhalt
        # Wenn wir hier sind, hat die Hauptspalte Inhalt
        return True

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Arbeitsblatt wählen
    if sheet is not None:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            matches = [s for s in wb.sheetnames if s.lower() == sheet.lower()]
            if matches:
                ws = wb[matches[0]]
            else:
                print("Blatt nicht gefunden. Verfügbare Blätter:")
                print(", ".join(wb.sheetnames))
                sys.exit(1)
    elif sheet_index is not None:
        if 0 <= sheet_index < len(wb.sheetnames):
            ws = wb[wb.sheetnames[sheet_index]]
        else:
            print(f"Blattindex {sheet_index} ist ungültig. Gültiger Bereich: 0..{len(wb.sheetnames)-1}")
            print("Verfügbare Blätter:")
            for i, name in enumerate(wb.sheetnames):
                print(f"  {i}: {name}")
            sys.exit(1)
    else:
        ws = wb.active
    names: list[tuple[str, str | None, str | None]] = []
    channel_idx: int | None = None
    mic_idx: int | None = None
    
    # Ermitteln der Headerzeile: explizit (--header-row) oder Suche in den ersten Zeilen
    def read_header_row(r: int) -> list[str]:
        vals: list[str] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            vals.append(str(v).strip() if v is not None else "")
        return vals

    # Spaltenindizes per Headername bestimmen, falls gewünscht
    if column_name is not None or channel_column is not None or mic_column is not None:
        headers: list[str]
        header_row_used: int
        search_limit = min(ws.max_row, 50)
        if header_row is not None and 1 <= header_row <= ws.max_row:
            headers = read_header_row(header_row)
            header_row_used = header_row
        else:
            # Erst Zeile 1 prüfen
            headers = read_header_row(1)
            header_row_used = 1
            # Wenn nicht gefunden, Suche in den ersten N Zeilen
            needed_columns = [c for c in [column_name, channel_column, mic_column] if c is not None]
            if any(all(h.lower() != col.strip().lower() for h in headers) for col in needed_columns):
                found_r: int | None = None
                for r in range(1, search_limit + 1):
                    hdr = read_header_row(r)
                    if any(any(h.lower() == col.strip().lower() for h in hdr) for col in needed_columns):
                        headers = hdr
                        found_r = r
                        break
                if found_r is not None:
                    header_row_used = found_r
                else:
                    print("Spaltennamen nicht gefunden. Tipp: --header-row N setzen.")
                    print("Erste Zeile (angenommener Header):")
                    print(", ".join(headers))
                    sys.exit(1)

        if column_name is not None:
            column = find_column_index(headers, column_name)
        if channel_column is not None:
            channel_idx = find_column_index(headers, channel_column)
        if mic_column is not None:
            mic_idx = find_column_index(headers, mic_column)

        start_row = header_row_used  # In diesem Fall ist die Header-Zeile auch die erste Datenzeile
    else:
        start_row = 2 if skip_header else 1
    # Finde die letzte Zeile mit Instrument in Spalte D (Index 3)
    last_instrument_row = start_row
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=4)  # Spalte D (3+1 für 1-basierte Excel-Indexierung)
        if cell.value is not None and str(cell.value).strip():
            last_instrument_row = r
    
    print(f"DEBUG: Letzte Zeile mit Instrument in Spalte D: {last_instrument_row}")
    row_count = 0
    
    # Durch die Zeilen gehen bis zur letzten Instrumentenzeile
    for r in range(start_row, last_instrument_row + 1):
        # Name/Instrument aus Spalte D lesen
        instrument_cell = ws.cell(row=r, column=4)  # Spalte D
        if instrument_cell.value is None:
            continue
            
        name = str(instrument_cell.value).strip()
        if not name:
            continue
            
        # Kanalnummer aus Spalte B lesen wenn gewünscht
        channel = None
        if channel_column is not None:
            channel_cell = ws.cell(row=r, column=2)  # Spalte B (nicht A!)
            channel_val = channel_cell.value
            if channel_val is not None:
                # Wenn es eine Zahl ist, als String formatieren
                if isinstance(channel_val, (int, float)):
                    channel = str(int(channel_val))
                else:
                    channel_str = str(channel_val).strip()
                    if channel_str:  # Nur wenn nicht leer
                        channel = channel_str
                
        # Mikrofon aus Spalte E lesen wenn gewünscht  
        mic = None
        if mic_column is not None:
            mic_cell = ws.cell(row=r, column=5)  # Spalte E
            if mic_cell.value is not None:
                mic = str(mic_cell.value).strip()
        
        print(f"DEBUG: Zeile {r} - Instrument: {name}, Kanal: {channel}, Mic: {mic}")
        names.append((name, channel, mic))
        row_count += 1
    
    print(f"DEBUG: {row_count} gültige Instrumente gefunden")
    return names


def load_names(
    path: Path,
    column: int | None,
    column_name: str | None,
    channel_column: str | None,
    mic_column: str | None,
    skip_header: bool,
    sheet: str | None,
    sheet_index: int | None,
    header_row: int | None
) -> list[tuple[str, str | None, str | None]]:
    suffix = path.suffix.lower()
    if suffix == '.csv':
        return read_names_from_csv(
            path,
            column=column,
            column_name=column_name,
            channel_column=channel_column,
            mic_column=mic_column,
            skip_header=skip_header
        )
    elif suffix in ('.xlsx', '.xlsm'):
        return read_names_from_xlsx(
            path,
            column=column,
            column_name=column_name,
            channel_column=channel_column,
            mic_column=mic_column,
            skip_header=skip_header,
            sheet=sheet,
            sheet_index=sheet_index,
            header_row=header_row,
        )
    else:
        print("Unterstützte Formate: .csv, .xlsx, .xlsm")
        sys.exit(1)


def get_headers_csv(path: Path) -> list[str]:
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        try:
            row = next(reader)
        except StopIteration:
            return []
        return [str(x).strip() for x in row]


def get_headers_xlsx(path: Path, sheet: str | None, sheet_index: int | None, header_row: int | None) -> tuple[list[str], str, int, int]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        print("Fehler: 'openpyxl' ist für XLSX erforderlich. Installiere es mit: pip install openpyxl")
        sys.exit(1)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Arbeitsblatt auswählen (gleiche Logik wie oben)
    if sheet is not None:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            sel_name = sheet
            sel_idx = wb.sheetnames.index(sel_name)
        else:
            matches = [s for s in wb.sheetnames if s.lower() == sheet.lower()]
            if matches:
                sel_name = matches[0]
                ws = wb[sel_name]
                sel_idx = wb.sheetnames.index(sel_name)
            else:
                print("Blatt nicht gefunden. Verfügbare Blätter:")
                print(", ".join(wb.sheetnames))
                sys.exit(1)
    elif sheet_index is not None:
        if 0 <= sheet_index < len(wb.sheetnames):
            sel_name = wb.sheetnames[sheet_index]
            ws = wb[sel_name]
            sel_idx = sheet_index
        else:
            print(f"Blattindex {sheet_index} ist ungültig. Gültiger Bereich: 0..{len(wb.sheetnames)-1}")
            print("Verfügbare Blätter:")
            for i, name in enumerate(wb.sheetnames):
                print(f"  {i}: {name}")
            sys.exit(1)
    else:
        ws = wb.active
        sel_name = ws.title
        sel_idx = wb.sheetnames.index(sel_name)
    row_used = 1 if header_row is None else max(1, min(header_row, ws.max_row))
    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=row_used, column=c).value
        headers.append(str(val).strip() if val is not None else "")
    return headers, sel_name, sel_idx, row_used


def main():
    parser = argparse.ArgumentParser(
        description="Überträgt Spur-Namen aus CSV/XLSX in Pro Tools per Hotkey (F8).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument('datei', type=str, help='Pfad zur CSV oder XLSX Datei (erste Spalte = Namen)')
    parser.add_argument('--spalte', type=int, default=None, help='Nullbasierter Spaltenindex (0 = erste Spalte)')
    parser.add_argument('--spaltenname', type=str, default=None, help="Spaltenüberschrift, z.B. 'instrument' (überschreibt --spalte)")
    parser.add_argument('--kanal-spalte', type=str, default=None, help="Name der Spalte mit Kanalnummern")
    parser.add_argument('--mikrofon-spalte', type=str, default=None, help="Name der Spalte mit Mikrofonbezeichnungen")
    parser.add_argument('--blatt', type=str, default=None, help='Tabellenblatt-Name (case-insensitive). Ohne Angabe wird das aktive Blatt verwendet.')
    parser.add_argument('--blatt-index', type=int, default=None, help='Tabellenblatt nach Index wählen (0 = erstes Blatt). Wird ignoriert, wenn --blatt gesetzt ist.')
    parser.add_argument('--header-row', type=int, default=None, help='Headerzeile (1-basiert). Nützlich, wenn Überschriften tiefer liegen oder zusammengeführt sind.')
    parser.add_argument('--skip-header', action='store_true', help='Erste Zeile überspringen (z.B. Überschrift)')
    parser.add_argument('--delay', type=float, default=0.5, help='Verzögerung vor dem Tippen (Sekunden)')
    parser.add_argument('--enter', action='store_true', help='Nach dem Namen automatisch Enter drücken (nicht mit --auto-next kombinieren)')
    parser.add_argument('--check', action='store_true', help='Nur Erkennung prüfen: Blatt, Header, Spalte und Vorschau anzeigen; keine Tasten senden')
    parser.add_argument('--preview', type=int, default=10, help='Anzahl anzuzeigender Namen in --check Modus')
    parser.add_argument('--auto-next', choices=['windows', 'mac'], default=None, help='Nach Bestätigen (Enter) automatisch zur nächsten Spur wechseln (Windows: Ctrl+Right, macOS: Cmd+Right)')
    parser.add_argument('--auto-run', action='store_true', help='Automatisch alle Namen in Folge übertragen (Start mit F8). Empfiehlt sich in Kombination mit --auto-next.')
    parser.add_argument('--next-delay', type=float, default=1.0, help='Verzögerung vor dem Spurwechsel (Sekunden)')
    args = parser.parse_args()

    path = Path(args.datei)
    if not path.exists():
        print(f"Datei nicht gefunden: {path}")
        sys.exit(1)

    names = load_names(
        path,
        column=args.spalte,
        column_name=args.spaltenname,
        channel_column=args.kanal_spalte,
        mic_column=args.mikrofon_spalte,
        skip_header=args.skip_header,
        sheet=args.blatt,
        sheet_index=args.blatt_index,
        header_row=args.header_row,
    )
    if not names:
        print("Keine Namen gefunden. Prüfe Datei/Spalte/Format.")
        sys.exit(1)
    
    # Format zur Anzeige bestimmen
    format_preview = "{name}"
    if args.kanal_spalte:
        format_preview = "{channel} {name}" if any(x[1] for x in names) else "{name}"
    if args.mikrofon_spalte:
        format_preview = format_preview + " {mic}" if any(x[2] for x in names) else format_preview

    if args.check:
        print("Erkennung:")
        if path.suffix.lower() in ('.xlsx', '.xlsm'):
            headers, sel_name, sel_idx, row_used = get_headers_xlsx(path, args.blatt, args.blatt_index, args.header_row)
            print(f"- Datei: {path.name} (XLSX/XLSM)")
            print(f"- Blatt: #{sel_idx} '{sel_name}'")
            print(f"- Headerzeile: {row_used}")
            if headers:
                print("- Header:")
                print("  " + ", ".join(f"[{i}] {h}" for i, h in enumerate(headers)))
            # Spaltenbestimmung für Anzeige
            col_used: str
            if args.spaltenname:
                try:
                    idx = next(i for i, h in enumerate(headers) if h.lower() == args.spaltenname.strip().lower())
                    col_used = f"Spaltenname '{args.spaltenname}' -> Index {idx}"
                except StopIteration:
                    col_used = f"Spaltenname '{args.spaltenname}' nicht in Header gefunden"
            else:
                col_idx = 0 if args.spalte is None else args.spalte
                col_used = f"Spaltenindex {col_idx}"
            print(f"- Spalte: {col_used}")
        else:
            print(f"- Datei: {path.name} (CSV)")
            headers = get_headers_csv(path)
            if headers:
                print("- Erste Zeile (Header/vermutet):")
                print("  " + ", ".join(f"[{i}] {h}" for i, h in enumerate(headers)))
            if args.spaltenname:
                try:
                    idx = next(i for i, h in enumerate(headers) if h.lower() == args.spaltenname.strip().lower())
                    print(f"- Spalte: Spaltenname '{args.spaltenname}' -> Index {idx}")
                except StopIteration:
                    print(f"- Spalte: Spaltenname '{args.spaltenname}' nicht in Header gefunden")
            else:
                col_idx = 0 if args.spalte is None else args.spalte
                print(f"- Spalte: Spaltenindex {col_idx}")

        print(f"- Gefundene Namen: {len(names)}")
        n = max(0, min(args.preview, len(names)))
        if n:
            print("- Vorschau:")
            for i in range(n):
                print(f"  {i}: {names[i]}")
        sys.exit(0)

    print("\nBereit! Anleitung:")
    print("1) Erstelle/Ordne die Spuren in Pro Tools.")
    print("2) Doppelklicke die erste Spurbezeichnung, damit der Umbenennen-Dialog aktiv ist.")
    print("3) Wechsle NICHT zurück zum Terminal.")
    print("4) Drücke F8 für jeden Namen. Mit F7 gehst du einen Namen zurück. ESC beendet.")
    print("   Tipp: Mit --enter drückt das Skript nach dem Tippen automatisch Enter.")
    print(f"Geladene Namen: {len(names)}. Starte mit Index 0.")

    idx = 0
    kb = Controller()
    paused = False
    auto_running = False
    stop_requested = False
    worker_thread: threading.Thread | None = None

    if args.auto_run and not args.auto_next and not args.enter:
        print("Hinweis: --auto-run ohne --auto-next oder --enter wird nach einem Namen stoppen. Empfohlen: --auto-next windows|mac.")

class TrackNamer:
    def __init__(self, excel_path, header_row=None, include_numbers=True, include_mics=True):
        print("DEBUG: TrackNamer initialization started")
        
        # Excel-Pfad als Path-Objekt speichern
        if isinstance(excel_path, str):
            self.excel_path = Path(excel_path)
        else:
            self.excel_path = excel_path
            
        self.header_row = header_row
        self.include_numbers = include_numbers
        self.include_mics = include_mics
        self.on_complete = None  # Callback wenn fertig
        
        # Namen aus Excel laden
        load_args = {
            'path': self.excel_path,
            'column': 3,  # Spalte D für Namen/Instrument (0-basiert: A=0, B=1, C=2, D=3)
            'column_name': None,
            'channel_column': "1" if include_numbers else None,  # Spalte B für Kanalnummern
            'mic_column': "4" if include_mics else None,  # Spalte E für Mikrofone
            'skip_header': False,
            'sheet': None,
            'sheet_index': None,
            'header_row': header_row,
        }
        
        print("\nDEBUG: Excel-Konfiguration:")
        print("- Instrumentenspalte: D (Index 3)")
        print("- Kanalspalte: B (Index 1)" if include_numbers else "- Keine Kanalnummern")
        print("- Mikrofonspalte: E (Index 4)" if include_mics else "- Keine Mikrofone")
        print("- Header-Zeile:", header_row)
        print("\nDEBUG: Loading names with args:", load_args)
        print("\nDEBUG: Lade Namen aus Excel...")
        self.names = load_names(**load_args)
        print(f"DEBUG: {len(self.names)} Namen geladen")
        
        if len(self.names) > 0:
            print("\nDEBUG: Erste 3 Namen zur Kontrolle:")
            for i, (name, channel, mic) in enumerate(self.names[:3]):
                formatted = []
                if channel:
                    formatted.append(f"Kanal: {channel}")
                formatted.append(f"Name: {name}")
                if mic:
                    formatted.append(f"Mic: {mic}")
                print(f"{i+1}. " + " | ".join(formatted))
        
        # Keyboard Controller und Status initialisieren
        self.kb = Controller()
        self.idx = 0
        self.auto_running = False
        self.stop_requested = False
        self.worker_thread = None
        
        # Konfiguration für das Verhalten
        self.delay = 0.01  # Verzögerung vor dem Tippen
        self.next_delay = 0.02  # Verzögerung zwischen Spuren
        self.auto_next = 'windows'  # Windows: Strg+Rechts für nächste Spur
        
    def release_all_keys(self):
        """Stellt sicher, dass alle Modifier-Tasten freigegeben sind"""
        try:
            # Alle möglicherweise hängenden Tasten freigeben
            self.kb.release(Key.ctrl)
            self.kb.release(Key.shift)
            self.kb.release(Key.alt)
            self.kb.release(Key.cmd)
            self.kb.release(Key.right)
            self.kb.release(Key.left)
            self.kb.release(Key.enter)
        except Exception as e:
            print(f"Warnung beim Freigeben der Tasten: {e}")
            pass

    def select_all(self):
        """Markiert den kompletten Text (Strg+A)"""
        try:
            if self.auto_next == 'mac':
                self.kb.press(Key.cmd)
                self.kb.press('a')
                self.kb.release('a')
                self.kb.release(Key.cmd)
            else:
                self.kb.press(Key.ctrl)
                self.kb.press('a')
                self.kb.release('a')
                self.kb.release(Key.ctrl)
        except Exception:
            print("DEBUG: Error in select_all")
            self.release_all_keys()

    def type_next_name(self):
        try:
            name, channel, mic = self.names[self.idx]
            print(f"\nDEBUG: Verarbeite Name {self.idx + 1}/{len(self.names)}:")
            print(f"- Name: {name}")
            if channel:
                print(f"- Kanal: {channel}")
            if mic:
                print(f"- Mikrofon: {mic}")
            
            self.idx += 1
            time.sleep(self.delay)
            self.select_all()
                    
            # Formatiere den Namen mit Unterstrichen zwischen den Spalten
            formatted_name = name.strip()
            if channel and self.include_numbers:  # Kanalnummer voranstellen wenn gewünscht
                formatted_name = f"{channel.strip()}_{formatted_name}"
            if mic and self.include_mics:  # Mikrofon anhängen wenn gewünscht
                formatted_name = f"{formatted_name}_{mic.strip()}"
                    
            print(f"DEBUG: Tippe formatierter Name: {formatted_name}")
            self.kb.type(formatted_name)
        except Exception as e:
            print(f"DEBUG: Fehler beim Tippen des Namens: {e}")
            return False
        
        # Prüfen ob dies der letzte Name ist
        is_last_name = self.idx >= len(self.names)
        
        # Beim letzten Namen mit Enter bestätigen und fertig
        if is_last_name:
            time.sleep(0.01)
            self.kb.press(Key.enter)
            time.sleep(0.01)
            self.kb.release(Key.enter)
            return

        # Ansonsten zur nächsten Spur navigieren
        if self.auto_next:
            print(f"DEBUG: Wechsele zur nächsten Spur ({self.auto_next})")
            try:
                if self.auto_next == 'windows':
                    # Sicherstellen dass alle Tasten losgelassen sind
                    self.release_all_keys()
                    time.sleep(0.01)
                    
                    # Strg+Rechts drücken um Namen zu bestätigen und zur nächsten Spur zu gehen
                    print("DEBUG: Drücke Strg+Rechts")
                    self.kb.press(Key.ctrl)
                    time.sleep(0.01)
                    self.kb.press(Key.right)
                    time.sleep(0.01)
                    self.kb.release(Key.right)
                    time.sleep(0.01)
                    self.kb.release(Key.ctrl)
                    time.sleep(0.01)  # Absolute Minimalpause nach der Navigation
                else:  # mac
                    print("DEBUG: Drücke Cmd+Rechts")
                    self.kb.press(Key.cmd)
                    self.kb.press(Key.right)
                    self.kb.release(Key.right)
                    self.kb.release(Key.cmd)
                print("DEBUG: Navigation zur nächsten Spur erfolgreich")
            except Exception as e:
                print(f"DEBUG: Fehler beim Wechsel zur nächsten Spur: {e}")
                return False
        return True

    def on_press(self, key):
        try:
            if key == Key.f9:  # F9 statt F8!
                print("DEBUG: F9 gedrückt, starte Benennung...")
                self.stop_requested = False
                self.run_all()
                return False  # Beende den Listener
            elif key == Key.esc:
                print("DEBUG: ESC gedrückt, breche ab...")
                self.stop_requested = True
                self.release_all_keys()
                if self.on_complete:
                    self.on_complete()
                return False
        except Exception as e:
            print(f"DEBUG: Fehler in on_press: {e}")
            self.release_all_keys()
            if self.on_complete:
                self.on_complete()
            return False
            # Bei Fehlern auch Tasten freigeben
            self.release_all_keys()
            print(f"Fehler beim Senden der Tasten: {e}")
            print("Alle Tasten wurden sicherheitshalber freigegeben.")

    def run_all(self):
        self.auto_running = True
        try:
            while self.idx < len(self.names) and not self.stop_requested:
                if not self.type_next_name():
                    break
                time.sleep(self.next_delay)
            
            if self.idx >= len(self.names):
                print("Fertig: Alle Namen wurden verwendet.")
            elif self.stop_requested:
                print("Abgebrochen durch Benutzer.")
        finally:
            self.auto_running = False

    def run(self):
        print("\nBereit zum Benennen!")
        print("1) Markiere die erste Spur in Pro Tools")
        print("2) Drücke F9 zum Starten")
        print("3) ESC zum Abbrechen")
        print(f"Es werden {len(self.names)} Spuren benannt.")
        
        try:
            with keyboard.Listener(on_press=self.on_press) as listener:
                listener.join()
        except Exception as e:
            print(f"Fehler: {e}")
            raise
        finally:
            self.release_all_keys()
            if self.on_complete:
                self.on_complete()


def main():
    parser = argparse.ArgumentParser(
        description="Überträgt Spur-Namen aus CSV/XLSX in Pro Tools per Hotkey (F8/F9).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument('datei', type=str, help='Pfad zur CSV oder XLSX Datei')
    parser.add_argument('--include-numbers', action='store_true', help='Kanalnummern aus Spalte A vor den Namen setzen')
    parser.add_argument('--include-mics', action='store_true', help='Mikrofontypen aus Spalte D hinter den Namen setzen')
    parser.add_argument('--header-row', type=int, default=None, help='Headerzeile (1-basiert)')
    args = parser.parse_args()

    path = Path(args.datei)
    if not path.exists():
        print(f"Datei nicht gefunden: {path}")
        sys.exit(1)

    namer = TrackNamer(
        excel_path=path,
        header_row=args.header_row,
        include_numbers=args.include_numbers,
        include_mics=args.include_mics
    )
    
    namer.run()


if __name__ == '__main__':
    main()
