#!/usr/bin/env python3
"""
Einfache Test-GUI um TreeView-Problem zu diagnostizieren
"""
import tkinter as tk
from tkinter import ttk
import openpyxl
from tkinter import filedialog, messagebox

class SimpleTestGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TreeView Test - Einfach")
        self.geometry("800x600")
        
        # Lade Button
        load_btn = tk.Button(self, text="Excel laden (TEST)", command=self.load_excel)
        load_btn.pack(pady=10)
        
        # Status Label
        self.status_label = tk.Label(self, text="Bereit zum Laden...", fg="blue")
        self.status_label.pack()
        
        # TreeView erstellen
        self.tree = ttk.Treeview(self, columns=("col1", "col2", "col3"), show="tree headings", height=20)
        self.tree.pack(expand=True, fill="both", padx=20, pady=20)
        
        # Spalten konfigurieren
        self.tree.heading("#0", text="Export")
        self.tree.heading("col1", text="Kanal")
        self.tree.heading("col2", text="Instrument")
        self.tree.heading("col3", text="Mikrofon")
        
        self.tree.column("#0", width=80)
        self.tree.column("col1", width=80)
        self.tree.column("col2", width=150)
        self.tree.column("col3", width=150)
        
        # Test-Daten einfügen
        self.add_test_data()
    
    def add_test_data(self):
        """Füge ein paar Test-Zeilen hinzu"""
        test_data = [
            ("☑", "1", "Violine 1", "MK 4"),
            ("☑", "2", "Violine 2", "MK 4"), 
            ("☐", "3", "Viola", "MK 4")
        ]
        
        for i, (check, kanal, inst, mic) in enumerate(test_data):
            self.tree.insert("", "end", text=check, values=(kanal, inst, mic))
        
        print("Test-Daten eingefügt")
    
    def load_excel(self):
        """Excel-Datei laden und anzeigen"""
        filetypes = [
            ("Excel files", ("*.xlsx", "*.xls", "*.xlsm")),
            ("All files", "*.*")
        ]
        filepath = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=filetypes)
        
        if not filepath:
            return
            
        try:
            # Alte Daten löschen (außer Test-Daten)
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            self.status_label.config(text=f"Lade: {filepath}")
            
            # Excel laden
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Layout B: Kanal=B(2), Instrument=D(4), Mikrofon=E(5)
            row_count = 0
            for row_num in range(7, sheet.max_row + 1):  # Ab Zeile 7
                kanal = sheet.cell(row=row_num, column=2).value
                instrument = sheet.cell(row=row_num, column=4).value
                mikrofon = sheet.cell(row=row_num, column=5).value
                
                if instrument:  # Nur wenn Instrument vorhanden
                    kanal_str = str(int(kanal)) if kanal else ""
                    inst_str = str(instrument) if instrument else ""
                    mic_str = str(mikrofon) if mikrofon else ""
                    
                    self.tree.insert("", "end", text="☑", values=(kanal_str, inst_str, mic_str))
                    row_count += 1
                    
                    if row_count < 5:  # Debug erste 5
                        print(f"Zeile {row_count}: {kanal_str} | {inst_str} | {mic_str}")
            
            workbook.close()
            self.status_label.config(text=f"✓ {row_count} Zeilen geladen", fg="green")
            print(f"Excel geladen: {row_count} Zeilen")
            
        except Exception as e:
            self.status_label.config(text=f"Fehler: {e}", fg="red")
            messagebox.showerror("Fehler", f"Excel konnte nicht geladen werden: {e}")

if __name__ == "__main__":
    app = SimpleTestGUI()
    app.mainloop()