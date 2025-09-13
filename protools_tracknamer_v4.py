import argparse
import sys
import time
from pathlib import Path
import threading
from pynput import keyboard
from pynput.keyboard import Key, Controller

class TrackNamer:
    def __init__(self, kb: Controller, names: list[tuple[str, str | None, str | None]], args):
        self.kb = kb
        self.names = names
        self.args = args
        self.idx = 0
        self.paused = False
        self.auto_running = False
        self.stop_requested = False
        self.worker_thread = None

    def select_all(self):
        try:
            if self.args.auto_next == 'mac':
                with self.kb.pressed(Key.cmd):
                    self.kb.press('a')
                    self.kb.release('a')
            else:
                with self.kb.pressed(Key.ctrl):
                    self.kb.press('a')
                    self.kb.release('a')
        except Exception:
            pass

    def activate_next_track(self):
        """Wechselt zur nächsten Spur und aktiviert diese"""
        print("DEBUG: Sende Strg+Rechts")
        # Sicherstellen dass keine Taste hängt
        self.kb.release(Key.ctrl)
        self.kb.release(Key.right)
        time.sleep(0.2)
        
        # Strg+Rechts senden
        with self.kb.pressed(Key.ctrl):
            self.kb.press(Key.right)
            time.sleep(0.1)
            self.kb.release(Key.right)
        time.sleep(0.2)

        # Spur aktivieren
        self.kb.press(Key.enter)
        self.kb.release(Key.enter)
        time.sleep(0.2)

    def type_name(self, name: str, channel: str | None, mic: str | None):
        """Tippt einen formatierten Namen"""
        formatted_name = name
        if channel and self.args.kanal_spalte:
            formatted_name = f"{channel}_{formatted_name}"
        if mic and self.args.mikrofon_spalte:
            formatted_name = f"{formatted_name}_{mic}"
        
        # Leerzeichen durch Unterstriche ersetzen
        formatted_name = formatted_name.replace(" ", "_")
        
        self.select_all()
        time.sleep(0.1)
        self.kb.type(formatted_name)
        print(f"Getippt: {formatted_name}")
        
        if self.args.auto_next:
            time.sleep(self.args.next_delay)
            self.activate_next_track()
        elif self.args.enter:
            self.kb.press(Key.enter)
            self.kb.release(Key.enter)

    def on_press(self, key):
        try:
            if key == Key.f9:
                if self.auto_running:
                    self.stop_requested = True
                    if self.worker_thread and self.worker_thread.is_alive():
                        self.worker_thread.join(timeout=1)
                    self.auto_running = False
                self.paused = not self.paused
                print(f"Pause: {self.paused}")
            elif key == Key.f7:
                if self.idx > 0:
                    self.idx -= 1
                print(f"Zurück. Nächster Index: {self.idx}")
            elif key == Key.f8 and not self.paused:
                if self.idx >= len(self.names):
                    print("Fertig: Alle Namen wurden verwendet.")
                    return
                name, channel, mic = self.names[self.idx]
                self.idx += 1
                time.sleep(self.args.delay)
                self.type_name(name, channel, mic)
            elif key == Key.esc:
                print("Beendet.")
                return False
        except Exception as e:
            print(f"Fehler beim Senden der Tasten: {e}")

def read_names_from_xlsx(path: Path, args) -> list[tuple[str, str | None, str | None]]:
    try:
        import openpyxl
    except ImportError:
        print("Fehler: Das Paket 'openpyxl' ist nicht installiert.")
        print("Installiere es mit: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[args.blatt] if args.blatt else wb.active
    names = []
    
    # Spalten finden
    name_col = None
    channel_col = None
    mic_col = None
    
    # Header-Zeile finden
    headers = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers.append(str(value).strip().lower())
            if value.strip().lower() == "instrument":
                name_col = col
            elif value.strip().lower() == "kanal":
                channel_col = col
            elif value.strip().lower() == "mikrofon":
                mic_col = col

    if not name_col:
        print("Fehler: Konnte die Instrument-Spalte nicht finden")
        print("Gefundene Spalten:", ", ".join(headers))
        sys.exit(1)

    # Namen einlesen (ab Zeile 2)
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        if name:
            name = str(name).strip()
            if not name:
                continue
                
            channel = None
            mic = None
            
            if channel_col:
                channel = ws.cell(row=row, column=channel_col).value
                if channel:
                    channel = str(channel).strip()
            
            if mic_col:
                mic = ws.cell(row=row, column=mic_col).value
                if mic:
                    mic = str(mic).strip()
                    
            names.append((name, channel, mic))
            
    return names

def main():
    parser = argparse.ArgumentParser(description="Pro Tools Track Namer")
    parser.add_argument("excel_file", help="Excel-Datei mit den Namen")
    parser.add_argument("--blatt", help="Name des Excel-Sheets")
    parser.add_argument("--kanal-spalte", action="store_true", help="Kanalnummer zum Namen hinzufügen")
    parser.add_argument("--mikrofon-spalte", action="store_true", help="Mikrofonname zum Namen hinzufügen")
    parser.add_argument("--delay", type=float, default=0.2, help="Verzögerung vor dem Tippen (Sekunden)")
    parser.add_argument("--next-delay", type=float, default=0.5, help="Verzögerung vor dem Spurwechsel (Sekunden)")
    parser.add_argument("--enter", action="store_true", help="Nach dem Namen Enter drücken")
    parser.add_argument("--auto-next", choices=['windows', 'mac'], default='windows', help="Automatisch zur nächsten Spur")
    args = parser.parse_args()

    path = Path(args.excel_file)
    if not path.exists():
        print(f"Fehler: Datei nicht gefunden: {path}")
        sys.exit(1)

    names = read_names_from_xlsx(path, args)
    if not names:
        print("Keine Namen gefunden!")
        sys.exit(1)

    print("Bereit! Anleitung:")
    print("1) Erstelle/Ordne die Spuren in Pro Tools")
    print("2) Doppelklicke die erste Spurbezeichnung")
    print("3) Drücke F8 für jeden Namen (F7 für zurück, ESC zum Beenden)")

    kb = Controller()
    namer = TrackNamer(kb, names, args)
    
    with keyboard.Listener(on_press=namer.on_press) as listener:
        listener.join()

if __name__ == "__main__":
    main()