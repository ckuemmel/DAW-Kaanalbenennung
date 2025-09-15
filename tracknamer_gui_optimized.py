import tkinter as tk
from tkinter import filedialog, ttk, messagebox
try:
    import openpyxl
except ImportError:
    messagebox.showerror("Fehler", "openpyxl Bibliothek nicht gefunden!\nBitte installieren: pip install openpyxl")
    exit(1)

class TrackNamerGUI(tk.Tk):
    def create_widgets(self):
        # Layout-Auswahl
        self.layout_frame = tk.Frame(self)
        self.layout_frame.pack(pady=5)
        tk.Label(self.layout_frame, text="Excel-Layout:").pack(side=tk.LEFT)
        
        self.layout_var = tk.StringVar(value="auto")
        layouts = [
            ("Automatisch erkennen", "auto"),
            ("Layout A: Instrument=C, Mikrofon=D", "layout_a"),
            ("Layout B: Instrument=D, Mikrofon=E", "layout_b")
        ]
        
        for text, value in layouts:
            rb = tk.Radiobutton(self.layout_frame, text=text, variable=self.layout_var, value=value)
            rb.pack(side=tk.LEFT, padx=5)
        
        # Excel laden Button
        load_btn = tk.Button(self, text="Excel laden", command=self.load_excel)
        load_btn.pack(pady=10)
        
        # Tabellen-Frame mit Scrollbars
        self.table_frame = tk.Frame(self)
        self.table_frame.pack(expand=True, fill="both", padx=10, pady=5)
        
        # Treeview mit Mehrfachauswahl
        self.tree = ttk.Treeview(self.table_frame, show="headings", selectmode="extended")
        
        # Scrollbars hinzufügen
        v_scrollbar = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout für Scrollbars
        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        self.table_frame.grid_rowconfigure(0, weight=1)
        self.table_frame.grid_columnconfigure(0, weight=1)
        
        # Bulk-Checkbox-Buttons
        self.bulk_frame = tk.Frame(self)
        self.bulk_frame.pack(pady=5)
        tk.Label(self.bulk_frame, text="Für ausgewählte Zeilen (blau markiert):").pack(side=tk.LEFT)
        tk.Button(self.bulk_frame, text="✓ Export AN", command=self.check_selected, bg="#e8f5e8").pack(side=tk.LEFT, padx=5)
        tk.Button(self.bulk_frame, text="✗ Export AUS", command=self.uncheck_selected, bg="#f5e8e8").pack(side=tk.LEFT, padx=5)
        tk.Button(self.bulk_frame, text="↕ Export umkehren", command=self.toggle_selected, bg="#e8f0f5").pack(side=tk.LEFT, padx=5)
        
        # Info-Frame für Anzeige der Selektion
        self.info_frame = tk.Frame(self)
        self.info_frame.pack(pady=5)
        self.selection_info = tk.Label(self.info_frame, text="", fg="blue")
        self.selection_info.pack()
        
        # Hilfe-Text
        help_text = "💡 Tipp: Bei Mehrfachauswahl schaltet ein Checkbox-Klick alle selektierten Zeilen synchron um"
        self.help_label = tk.Label(self.info_frame, text=help_text, fg="gray", font=("Arial", 8))
        self.help_label.pack()
        
        # Spaltenauswahl
        self.col_frame = tk.Frame(self)
        self.col_frame.pack(pady=10)
        self.col_label = tk.Label(self.col_frame, text="Spalten für Namensbildung auswählen:")
        self.col_label.pack(side=tk.LEFT)
        
        # Speicherort-Auswahl und Export-Button
        self.export_path = tk.StringVar()
        self.export_path_row = tk.Frame(self)
        self.export_path_row.pack(pady=10)
        self.export_path_label = tk.Label(self.export_path_row, text="Speicherort:")
        self.export_path_label.pack(side=tk.LEFT)
        self.export_path_entry = tk.Entry(self.export_path_row, textvariable=self.export_path, width=60)
        self.export_path_entry.pack(side=tk.LEFT)
        self.export_path_btn = tk.Button(self.export_path_row, text="...", command=self.choose_export_path)
        self.export_path_btn.pack(side=tk.LEFT)
        self.export_btn = tk.Button(self, text="Tracknamen speichern", command=self.export_names)
        self.export_btn.pack(pady=10)

    def check_selected(self):
        """Markiere alle selektierten Zeilen"""
        selected_items = self.tree.selection()
        for item_id in selected_items:
            if item_id in self.row_ids:
                row_index = self.row_ids.index(item_id)
                var = self.row_vars[row_index]
                var.set(1)
                # Anzeige aktualisieren
                current_values = list(self.tree.item(item_id, 'values'))
                current_values[0] = "☑"
                self.tree.item(item_id, values=current_values)
        self.update_selection_info()

    def uncheck_selected(self):
        """Entmarkiere alle selektierten Zeilen"""
        selected_items = self.tree.selection()
        for item_id in selected_items:
            if item_id in self.row_ids:
                row_index = self.row_ids.index(item_id)
                var = self.row_vars[row_index]
                var.set(0)
                # Anzeige aktualisieren
                current_values = list(self.tree.item(item_id, 'values'))
                current_values[0] = "☐"
                self.tree.item(item_id, values=current_values)
        self.update_selection_info()

    def toggle_selected(self):
        """Kehre die Markierung aller selektierten Zeilen um"""
        selected_items = self.tree.selection()
        for item_id in selected_items:
            if item_id in self.row_ids:
                row_index = self.row_ids.index(item_id)
                var = self.row_vars[row_index]
                new_value = 0 if var.get() else 1
                var.set(new_value)
                # Anzeige aktualisieren
                current_values = list(self.tree.item(item_id, 'values'))
                current_values[0] = "☑" if new_value else "☐"
                self.tree.item(item_id, values=current_values)
        self.update_selection_info()

    def update_selection_info(self):
        """Aktualisiere die Info über die aktuelle Selektion"""
        if not hasattr(self, 'selection_info'):
            return
            
        selected_items = self.tree.selection()
        selected_count = len(selected_items)
        
        # Zähle export-markierte Zeilen
        export_count = sum(1 for var in self.row_vars if var.get())
        
        info_text = f"Tabellen-Auswahl: {selected_count} Zeilen | Export-Markierung: {export_count} Zeilen"
        self.selection_info.config(text=info_text)

    def show_column_selectors(self):
        for widget in self.col_frame.winfo_children()[1:]:
            widget.destroy()
        self.col_vars = []
        for col in self.column_names:
            var = tk.IntVar(value=1)
            cb = tk.Checkbutton(self.col_frame, text=col, variable=var)
            cb.pack(side=tk.LEFT)
            self.col_vars.append((col, var))

    def export_names(self):
        selected_cols = [col for col, var in self.col_vars if var.get()]
        selected_rows = [int(self.row_ids[i]) for i, var in enumerate(self.row_vars) if var.get()]
        
        if not selected_cols:
            messagebox.showinfo("Hinweis", "Bitte wähle mindestens eine Spalte aus.")
            return
        if not selected_rows:
            messagebox.showinfo("Hinweis", "Bitte wähle mindestens eine Zeile aus.")
            return

        names = []
        for row_idx in selected_rows:
            row_data = self.data[row_idx]
            # Erste Spalte immer als int ohne Komma
            first = row_data[0]
            try:
                first = int(float(first)) if first else ""
            except:
                first = str(first).split('.')[0] if '.' in str(first) else str(first)
            
            # Nur gewählte Spalten verwenden
            name_parts = []
            for i, (col_name, var) in enumerate(self.col_vars):
                if var.get() and i < len(row_data):
                    if i == 0:  # Erste Spalte
                        name_parts.append(str(first))
                    else:
                        value = row_data[i]
                        if value:  # Nur nicht-leere Werte
                            name_parts.append(str(value))
            
            if name_parts:
                name = "_".join(name_parts)
                names.append(name)

        out_path = self.export_path.get()
        if not out_path:
            messagebox.showinfo("Hinweis", "Bitte wähle einen Zielpfad aus.")
            return

        # Sequoia-Header
        header = [
            "Sequoia - Track List",
            "Project: \"C:\\Users\\carst\\OneDrive\\Dokumente\\GitHub\\Namen Sequoia\\Data\\test_modified.vip\"",
            "*********************************************************",
            ""
        ]
        
        with open(out_path, "w", encoding="utf-8") as f:
            for line in header:
                f.write(line + "\n")
            for i, name in enumerate(names, 1):
                f.write(f"Spur {i}  [{name}]\n")
            # Master-Zeile am Ende
            f.write(f"Spur {len(names)+1}  [Master]  (unsichtbar)\n")
        
        messagebox.showinfo("Erfolg", f"{len(names)} Tracknamen im Sequoia-Format gespeichert!")

    def choose_export_path(self):
        path = filedialog.asksaveasfilename(title="Zielpfad wählen", defaultextension=".txt", filetypes=[("Textdatei", "*.txt")])
        if path:
            self.export_path.set(path)

    def __init__(self):
        super().__init__()
        self.title("Sequoia TrackNamer - Optimiert")
        self.geometry("1000x700")
        self.data = []
        self.column_names = []
        self.selected_rows = set()
        self.selected_columns = []
        self.col_vars = []
        self.row_vars = []
        self.row_ids = []
        self.create_widgets()

    def load_excel(self):
        filetypes = [("Excel files", "*.xlsx;*.xls;*.xlsm")]
        filepath = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=filetypes)
        if not filepath:
            return

        try:
            # Direkt mit openpyxl arbeiten (ohne pandas)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Header aus Zeile 6 (Index 5)
            header_row = 6
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                headers.append(str(cell_value) if cell_value else f"Spalte_{col}")
            
            # Layout bestimmen
            layout = self.layout_var.get()
            
            if layout == "auto":
                # Verbesserte automatische Erkennung
                
                # 1. Header-Analyse (Zeile 6)
                header_c = sheet.cell(row=6, column=3).value
                header_d = sheet.cell(row=6, column=4).value
                header_e = sheet.cell(row=6, column=5).value
                
                print(f"Header-Analyse: C='{header_c}', D='{header_d}', E='{header_e}'")
                
                # Prüfe Header für "Instrument" 
                if header_c and "instrument" in str(header_c).lower():
                    detected_layout = "layout_a"
                    print("→ Layout A: 'Instrument' in Header Spalte C gefunden")
                elif header_d and "instrument" in str(header_d).lower():
                    detected_layout = "layout_b"
                    print("→ Layout B: 'Instrument' in Header Spalte D gefunden")
                else:
                    # 2. Daten-Analyse mehrerer Zeilen
                    data_in_c = 0
                    data_in_d = 0
                    
                    for check_row in range(8, min(15, sheet.max_row + 1)):  # Zeilen 8-14
                        val_c = sheet.cell(row=check_row, column=3).value
                        val_d = sheet.cell(row=check_row, column=4).value
                        
                        # Prüfe auf nicht-numerische, nicht-leere Werte
                        if val_c and str(val_c).strip() and not str(val_c).replace('.','').isdigit():
                            data_in_c += 1
                        if val_d and str(val_d).strip() and not str(val_d).replace('.','').isdigit():
                            data_in_d += 1
                    
                    print(f"Daten-Analyse: {data_in_c} Instrument-Einträge in C, {data_in_d} in D")
                    
                    if data_in_c > 0 and data_in_d > 0:
                        detected_layout = "layout_a"  # Beide Spalten haben Daten = Layout A
                        print("→ Layout A: Daten in C und D gefunden")
                    elif data_in_d > data_in_c:
                        detected_layout = "layout_b"  # Mehr Daten in D = Layout B  
                        print("→ Layout B: Hauptsächlich Daten in D")
                    elif data_in_c > 0:
                        detected_layout = "layout_a"  # Nur C hat Daten = Layout A
                        print("→ Layout A: Hauptsächlich Daten in C")
                    else:
                        detected_layout = "layout_b"  # Fallback
                        print("→ Layout B: Fallback (keine eindeutigen Daten)")
                
                print(f"Erkanntes Layout: {detected_layout}")
            else:
                detected_layout = layout
            
            # Spalten-Indizes basierend auf Layout
            if detected_layout == "layout_a":
                # Layout A: Kanal=B(2), Instrument=C(3), Mikrofon=D(4)
                wanted_indices = [2, 3, 4]
                self.column_names = ['Kanal', 'Instrument', 'Mikrofon']
                instrument_col_index = 1  # Index im data array (0-basiert)
            else:
                # Layout B: Kanal=B(2), Instrument=D(4), Mikrofon=E(5)  
                wanted_indices = [2, 4, 5]
                self.column_names = ['Kanal', 'Instrument', 'Mikrofon']
                instrument_col_index = 1  # Index im data array (0-basiert)
            
            # Daten sammeln
            data = []
            last_instrument_row = -1
            
            for row_num in range(header_row + 1, sheet.max_row + 1):
                row_data = []
                for col_idx in wanted_indices:
                    cell_value = sheet.cell(row=row_num, column=col_idx).value
                    row_data.append(cell_value)
                
                # Prüfe ob Instrument-Spalte Daten hat
                if row_data[instrument_col_index]:  # Instrument nicht leer
                    last_instrument_row = len(data)
                
                data.append(row_data)
            
            workbook.close()
            
            # Nur bis zur letzten Instrument-Zeile
            if last_instrument_row >= 0:
                self.data = data[:last_instrument_row + 1]
            else:
                messagebox.showerror("Fehler", "Keine Daten in der Instrument-Spalte gefunden.")
                return
            
            print(f"Geladene Daten: {len(self.data)} Zeilen")
            print(f"Spalten: {self.column_names}")
            
            # Layout-Info anzeigen
            layout_info = {
                "layout_a": "Layout A (Instrument=C, Mikrofon=D)",
                "layout_b": "Layout B (Instrument=D, Mikrofon=E)"
            }
            self.detected_layout = detected_layout
            
            self.show_table()
            self.show_column_selectors()
            
            # Layout-Info in der GUI anzeigen
            if hasattr(self, 'layout_info_label'):
                self.layout_info_label.destroy()
            
            info_text = f"✓ Erkanntes Layout: {layout_info[detected_layout]} | {len(self.data)} Zeilen geladen"
            self.layout_info_label = tk.Label(self, text=info_text, fg="green", font=("Arial", 9))
            self.layout_info_label.pack()
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Excel konnte nicht geladen werden: {e}")
            return

    def show_table(self):
        # Alte Inhalte löschen
        self.tree.delete(*self.tree.get_children())
        
        # Spalten definieren (Auswahl + Datenspalten)
        columns = ["Export"] + self.column_names
        self.tree['columns'] = columns
        
        # Spaltenüberschriften setzen
        for col in columns:
            self.tree.heading(col, text=col)
            if col == "Export":
                self.tree.column(col, width=60, anchor='center')
            else:
                self.tree.column(col, width=120, anchor='w')
        
        # Zeilen-Checkboxen initialisieren
        self.row_vars = []
        self.row_ids = []
        
        # Daten in Tabelle einfügen
        for idx, row in enumerate(self.data):
            # Checkbox-Variable für diese Zeile
            # Standardmäßig ausgewählt nur wenn in der Instrument-Spalte Daten stehen
            has_instrument = row[1] and str(row[1]).strip() != ''
            var = tk.IntVar(value=1 if has_instrument else 0)
            self.row_vars.append(var)
            self.row_ids.append(str(idx))
            
            # Daten formatieren
            row_data = []
            for i, value in enumerate(row):
                if not value:
                    row_data.append("")
                elif i == 0 and isinstance(value, (float, int)):  # Erste Spalte (Kanal) als Integer
                    row_data.append(str(int(value)))
                else:
                    row_data.append(str(value))
            
            # Zeile mit Checkbox einfügen
            checkbox_text = "☑" if var.get() else "☐"
            values = [checkbox_text] + row_data
            self.tree.insert('', 'end', iid=str(idx), values=values)
        
        # Click-Handler für synchronisierte Checkbox-Klicks
        def toggle_synchronized_checkbox(event):
            # Identifiziere angeklickte Zeile und Spalte
            item_id = self.tree.identify_row(event.y)
            column = self.tree.identify_column(event.x)
            
            # Nur reagieren wenn erste Spalte (Export) angeklickt wurde
            if column == '#1' and item_id in self.row_ids:
                # Aktuelle Selektion in der Tabelle holen
                selected_items = self.tree.selection()
                
                # Wenn die geklickte Zeile Teil der Selektion ist, alle selektierten Zeilen umschalten
                if item_id in selected_items and len(selected_items) > 1:
                    # Multi-Selektion: Alle selektierten Zeilen synchron umschalten
                    # Bestimme neuen Wert basierend auf der geklickten Zeile
                    clicked_row_index = self.row_ids.index(item_id)
                    clicked_var = self.row_vars[clicked_row_index]
                    new_value = 0 if clicked_var.get() else 1
                    
                    # Alle selektierten Zeilen auf den neuen Wert setzen
                    for selected_id in selected_items:
                        if selected_id in self.row_ids:
                            row_index = self.row_ids.index(selected_id)
                            var = self.row_vars[row_index]
                            var.set(new_value)
                            
                            # Anzeige aktualisieren
                            current_values = list(self.tree.item(selected_id, 'values'))
                            current_values[0] = "☑" if new_value else "☐"
                            self.tree.item(selected_id, values=current_values)
                else:
                    # Einzel-Selektion: Nur die geklickte Zeile umschalten
                    row_index = self.row_ids.index(item_id)
                    var = self.row_vars[row_index]
                    
                    # Checkbox umschalten
                    new_value = 0 if var.get() else 1
                    var.set(new_value)
                    
                    # Anzeige aktualisieren
                    current_values = list(self.tree.item(item_id, 'values'))
                    current_values[0] = "☑" if new_value else "☐"
                    self.tree.item(item_id, values=current_values)
                
                # Info aktualisieren
                self.update_selection_info()
                
                return "break"  # Verhindere normale Selektion nur für Checkbox-Spalte
            # Für andere Spalten: Normale Selektion erlauben
            return None
        
        # Events binden
        self.tree.bind('<Button-1>', toggle_synchronized_checkbox)
        self.tree.bind('<<TreeviewSelect>>', lambda e: self.update_selection_info())

if __name__ == "__main__":
    app = TrackNamerGUI()
    app.mainloop()