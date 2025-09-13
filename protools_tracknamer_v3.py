import argparse
import sys
import time
from pathlib import Path
import threading
from pynput import keyboard
from pynput.keyboard import Key, Controller

def read_names_from_xlsx(path: Path, sheet: str | None = None) -> list[tuple[str, str | None, str | None]]:
    try:
        import openpyxl
    except ImportError:
        print("Fehler: Das Paket 'openpyxl' ist nicht installiert.")
        print("Installiere es mit: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    names = []
    
    # Spalten finden
    name_col = None
    channel_col = None
    mic_col = None
    
    # Suche in den ersten 10 Zeilen nach der Header-Zeile
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value:
                if isinstance(value, str):
                    if value.lower() == "instrument":
                        name_col = col
                    elif value.lower() == "kanal":
                        channel_col = col
                    elif value.lower() == "mikrofon":
                        mic_col = col

    if not name_col:
        print("Fehler: Konnte die Instrument-Spalte nicht finden")
        sys.exit(1)

    # Namen einlesen
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        if name:
            name = str(name).strip()
            if not name:
                continue
                
            channel = None
            mic = None
            
            if channel_col:
                channel = ws.cell(row=row, column=channel_col).value
                if channel:
                    channel = str(channel).strip()
            
            if mic_col:
                mic = ws.cell(row=row, column=mic_col).value
                if mic:
                    mic = str(mic).strip()
                    
            # Formatiere den Namen
            formatted_name = name
            if channel:
                formatted_name = f"{channel}_{formatted_name}"
            if mic:
                formatted_name = f"{formatted_name}_{mic}"
            
            # Ersetze Leerzeichen durch Unterstriche
            formatted_name = formatted_name.replace(" ", "_")
            names.append((formatted_name, channel, mic))
            
    return names

def main():
    parser = argparse.ArgumentParser(description="Pro Tools Track Namer")
    parser.add_argument("excel_file", help="Excel-Datei mit den Namen")
    parser.add_argument("--sheet", help="Name des Excel-Sheets")
    args = parser.parse_args()

    path = Path(args.excel_file)
    if not path.exists():
        print(f"Fehler: Datei nicht gefunden: {path}")
        sys.exit(1)

    names = read_names_from_xlsx(path, args.sheet)
    if not names:
        print("Keine Namen gefunden!")
        sys.exit(1)

    print("Bereit! Anleitung:")
    print("1) Erstelle/Ordne die Spuren in Pro Tools")
    print("2) Doppelklicke die erste Spurbezeichnung")
    print("3) Drücke F8 für jeden Namen (F7 für zurück, ESC zum Beenden)")

    kb = Controller()
    idx = 0

    def activate_track():
        """Aktiviert den Bearbeitungsmodus der aktuellen Spur"""
        time.sleep(0.2)  # Kurz warten
        kb.press(Key.enter)
        kb.release(Key.enter)
        time.sleep(0.2)  # Warten bis der Bearbeitungsmodus aktiv ist

    def next_track():
        """Wechselt zur nächsten Spur und aktiviert sie"""
        # Strg+Rechts für nächste Spur
        print("DEBUG: Sende Strg+Rechts")
        with kb.pressed(Key.ctrl):
            kb.press(Key.right)
            kb.release(Key.right)
        time.sleep(0.5)  # Längere Wartezeit für zuverlässigere Ausführung
        
        # Aktiviere die neue Spur
        activate_track()
        time.sleep(0.2)  # Zusätzliche Wartezeit für Stabilität

    def on_press(key):
        nonlocal idx
        try:
            if key == keyboard.Key.f8:
                if idx >= len(names):
                    print("Alle Namen verwendet!")
                    return

                name = names[idx][0]
                idx += 1

                # Warte kurz
                time.sleep(0.2)

                # Selektiere vorhandenen Text
                with kb.pressed(Key.ctrl):
                    kb.press('a')
                    kb.release('a')
                time.sleep(0.1)

                # Tippe den neuen Namen
                kb.type(name)
                print(f"Getippt: {name}")

                time.sleep(0.2)

                # Enter drücken um Namen zu bestätigen
                kb.press(Key.enter)
                kb.release(Key.enter)
                time.sleep(0.2)

                # Zur nächsten Spur und diese aktivieren
                next_track()

            elif key == keyboard.Key.f7:
                if idx > 0:
                    idx -= 1
                    print(f"Zurück zu Index {idx}")

            elif key == keyboard.Key.esc:
                print("Beendet.")
                return False

        except Exception as e:
            print(f"Fehler: {e}")

    with keyboard.Listener(on_press=on_press) as listener:
        listener.join()

if __name__ == "__main__":
    main()