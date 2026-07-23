#!/usr/bin/env python3
"""
Vereinfachte TrackNamer GUI - ohne komplexes Layout
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl

class SimpleTrackNamerGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Pro Tools TrackNamer - Vereinfacht")
        self.geometry("900x700")
        
        self.data = []
        self.column_names = ['Kanal', 'Instrument', 'Mikrofon']
        self.row_vars = []
        self.row_ids = []
        
        self.create_widgets()
    
    def create_widgets(self):
        # Layout-Auswahl
        layout_frame = tk.Frame(self)
        layout_frame.pack(pady=5)
        tk.Label(layout_frame, text="Layout:").pack(side=tk.LEFT)
        
        self.layout_var = tk.StringVar(value="auto")
        layouts = [
            ("Auto", "auto"),
            ("Layout A (Inst=C, Mic=D)", "layout_a"),
            ("Layout B (Inst=D, Mic=E)", "layout_b")
        ]
        
        for text, value in layouts:
            rb = tk.Radiobutton(layout_frame, text=text, variable=self.layout_var, value=value)
            rb.pack(side=tk.LEFT, padx=5)
        
        # Excel laden Button
        load_btn = tk.Button(self, text="Excel laden", command=self.load_excel, bg="lightblue")
        load_btn.pack(pady=10)
        
        # Status
        self.status_label = tk.Label(self, text="Bereit...", fg="blue")
        self.status_label.pack()
        
        # TreeView - EINFACH ohne komplexe Frames
        self.tree = ttk.Treeview(self, columns=("kanal", "instrument", "mikrofon"), show="tree headings", height=20)
        self.tree.pack(expand=True, fill="both", padx=20, pady=10)
        
        # Spalten
        self.tree.heading("#0", text="Export")
        self.tree.heading("kanal", text="Kanal")
        self.tree.heading("instrument", text="Instrument") 
        self.tree.heading("mikrofon", text="Mikrofon")
        
        self.tree.column("#0", width=80)
        self.tree.column("kanal", width=80)
        self.tree.column("instrument", width=200)
        self.tree.column("mikrofon", width=200)
        
        # Buttons
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="✓ Alle auswählen", command=self.select_all).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="✗ Alle abwählen", command=self.deselect_all).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Tracknamen exportieren", command=self.export_names, bg="lightgreen").pack(side=tk.LEFT, padx=10)
        
        # Export-Pfad
        path_frame = tk.Frame(self)
        path_frame.pack(pady=5)
        tk.Label(path_frame, text="Export:").pack(side=tk.LEFT)
        self.export_path = tk.StringVar(value="tracknamen.txt")
        tk.Entry(path_frame, textvariable=self.export_path, width=50).pack(side=tk.LEFT, padx=5)
        tk.Button(path_frame, text="...", command=self.choose_export_path).pack(side=tk.LEFT)
    
    def load_excel(self):
        filetypes = [
            ("Excel files", ("*.xlsx", "*.xls", "*.xlsm")),
            ("All files", "*.*")
        ]
        filepath = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=filetypes)
        
        if not filepath:
            return
        
        try:
            self.status_label.config(text=f"Lade: {filepath}", fg="blue")
            
            # TreeView leeren
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Excel laden
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Layout bestimmen
            layout = self.layout_var.get()
            if layout == "auto":
                # Einfache Auto-Erkennung
                header_d = sheet.cell(row=6, column=4).value
                if header_d and "instrument" in str(header_d).lower():
                    detected_layout = "layout_b"
                else:
                    detected_layout = "layout_a"
            else:
                detected_layout = layout
            
            # Spalten-Indizes
            if detected_layout == "layout_a":
                wanted_indices = [2, 3, 4]  # B, C, D
            else:
                wanted_indices = [2, 4, 5]  # B, D, E
            
            # Daten laden
            self.data = []
            self.row_vars = []
            self.row_ids = []
            
            for row_num in range(7, sheet.max_row + 1):  # Ab Zeile 7
                row_data = []
                for col_idx in wanted_indices:
                    cell_value = sheet.cell(row=row_num, column=col_idx).value
                    row_data.append(cell_value)
                
                # Nur Zeilen mit Instrument
                if row_data[1]:  # Instrument-Spalte nicht leer
                    self.data.append(row_data)
                    
                    # In TreeView einfügen
                    idx = len(self.data) - 1
                    self.row_ids.append(str(idx))
                    
                    # Daten formatieren
                    kanal_str = str(int(row_data[0])) if row_data[0] else ""
                    inst_str = str(row_data[1]) if row_data[1] else ""
                    mic_str = str(row_data[2]) if row_data[2] else ""
                    
                    # Variable für Checkbox
                    var = tk.IntVar(value=1)
                    self.row_vars.append(var)
                    
                    # TreeView-Eintrag
                    self.tree.insert("", "end", iid=str(idx), text="☑", values=(kanal_str, inst_str, mic_str))
            
            workbook.close()
            
            self.status_label.config(text=f"✓ {len(self.data)} Zeilen geladen ({detected_layout})", fg="green")
            print(f"Excel erfolgreich geladen: {len(self.data)} Zeilen")
            
        except Exception as e:
            self.status_label.config(text=f"Fehler: {e}", fg="red")
            messagebox.showerror("Fehler", f"Excel konnte nicht geladen werden: {e}")
    
    def select_all(self):
        for i, var in enumerate(self.row_vars):
            var.set(1)
            self.tree.item(str(i), text="☑")
    
    def deselect_all(self):
        for i, var in enumerate(self.row_vars):
            var.set(0)
            self.tree.item(str(i), text="☐")
    
    def choose_export_path(self):
        path = filedialog.asksaveasfilename(
            title="Speicherort wählen", 
            defaultextension=".txt",
            filetypes=[("Textdatei", "*.txt"), ("Alle Dateien", "*.*")]
        )
        if path:
            self.export_path.set(path)
    
    def export_names(self):
        selected_rows = [i for i, var in enumerate(self.row_vars) if var.get()]
        
        if not selected_rows:
            messagebox.showinfo("Hinweis", "Bitte wähle mindestens eine Zeile aus.")
            return
        
        try:
            names = []
            for row_idx in selected_rows:
                row_data = self.data[row_idx]
                # Format: Kanal_Instrument_Mikrofon
                kanal = str(int(row_data[0])) if row_data[0] else ""
                inst = str(row_data[1]) if row_data[1] else ""
                mic = str(row_data[2]) if row_data[2] else ""
                name = f"{kanal}_{inst}_{mic}"
                names.append(name)
            
            # Export
            out_path = self.export_path.get()
            if not out_path:
                messagebox.showinfo("Hinweis", "Bitte wähle einen Speicherpfad aus.")
                return
            
            with open(out_path, "w", encoding="utf-8") as f:
                for i, name in enumerate(names, 1):
                    f.write(f"Spur {i}  [{name}]\n")
            
            messagebox.showinfo("Erfolg", f"{len(names)} Tracknamen gespeichert!")
            self.status_label.config(text=f"✓ {len(names)} Namen exportiert", fg="green")
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Export fehlgeschlagen: {e}")

if __name__ == "__main__":
    app = SimpleTrackNamerGUI()
    app.mainloop()