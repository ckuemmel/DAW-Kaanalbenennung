#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

def test_layout_detection():
    files = [
        ("Data/test_layout_a.xlsx", "Layout A Test"),
        ("Data/Inputliste Blasorchester 2025 V2.xlsm", "Blasorchester"),
        ("Data/Inputliste GKO V1.xlsm", "GKO")
    ]
    
    for filepath, name in files:
        print(f"\n=== {name}: {filepath} ===")
        
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook.active
            
            # Header-Analyse
            header_c = sheet.cell(row=6, column=3).value
            header_d = sheet.cell(row=6, column=4).value
            header_e = sheet.cell(row=6, column=5).value
            
            print(f"Header: C='{header_c}', D='{header_d}', E='{header_e}'")
            
            # Layout-Erkennung wie in GUI
            if header_c and "instrument" in str(header_c).lower():
                detected = "layout_a"
                reason = f"'Instrument' in Header Spalte C"
            elif header_d and "instrument" in str(header_d).lower():
                detected = "layout_b" 
                reason = f"'Instrument' in Header Spalte D"
            else:
                # Daten-Analyse
                data_in_c = 0
                data_in_d = 0
                
                for check_row in range(8, min(15, sheet.max_row + 1)):
                    val_c = sheet.cell(row=check_row, column=3).value
                    val_d = sheet.cell(row=check_row, column=4).value
                    
                    if val_c and str(val_c).strip() and not str(val_c).replace('.','').isdigit():
                        data_in_c += 1
                    if val_d and str(val_d).strip() and not str(val_d).replace('.','').isdigit():
                        data_in_d += 1
                
                print(f"Daten: {data_in_c} in C, {data_in_d} in D")
                
                if data_in_c > 0 and data_in_d > 0:
                    detected = "layout_a"
                    reason = "Daten in C und D"
                elif data_in_d > data_in_c:
                    detected = "layout_b"
                    reason = "Mehr Daten in D"
                elif data_in_c > 0:
                    detected = "layout_a"  
                    reason = "Nur Daten in C"
                else:
                    detected = "layout_b"
                    reason = "Fallback"
            
            print(f"→ {detected.upper()}: {reason}")
            
            # Zeige beispiel Daten
            print("Beispiel-Daten (Zeile 8):")
            example_b = sheet.cell(row=8, column=2).value
            example_c = sheet.cell(row=8, column=3).value
            example_d = sheet.cell(row=8, column=4).value
            example_e = sheet.cell(row=8, column=5).value
            print(f"  B={example_b}, C={example_c}, D={example_d}, E={example_e}")
            
        except Exception as e:
            print(f"Fehler: {e}")

if __name__ == "__main__":
    test_layout_detection()