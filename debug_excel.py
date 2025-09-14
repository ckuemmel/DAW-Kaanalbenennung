import openpyxl
from pathlib import Path

# Excel-Datei öffnen
file_path = Path("Data/Inputliste GKO V1.xlsm")
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
ws = wb.active

print("DEBUG: Analysiere Excel-Datei...")
print("Erste 10 Zeilen, Spalten A bis F:")
print("Zeile | A      | B      | C      | D (Instr.) | E (Mikro.) | F      ")
print("------|--------|--------|--------|------------|------------|--------")

for r in range(8, 18):  # Zeilen 8-17
    col_a = ws.cell(row=r, column=1).value
    col_b = ws.cell(row=r, column=2).value
    col_c = ws.cell(row=r, column=3).value
    col_d = ws.cell(row=r, column=4).value
    col_e = ws.cell(row=r, column=5).value
    col_f = ws.cell(row=r, column=6).value
    print(f"{r:5} | {repr(col_a)[:6]:6} | {repr(col_b)[:6]:6} | {repr(col_c)[:6]:6} | {repr(col_d)[:9]:9} | {repr(col_e)[:9]:9} | {repr(col_f)[:6]:6}")

wb.close()