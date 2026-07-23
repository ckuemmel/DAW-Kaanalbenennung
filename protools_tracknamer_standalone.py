#!/usr/bin/env python3
"""
Pro Tools Track Namer - Standalone Version
Automatische Track-Erstellung und Benennung für Pro Tools
"""

import os
import sys
import tempfile
import webbrowser
import time
import socket
import subprocess
from threading import Timer
import signal

# Flask und andere Imports
try:
    from flask import Flask, render_template_string, request, jsonify, redirect, url_for, flash
    import openpyxl
    from pynput import keyboard
    from pynput.keyboard import Key, Controller as KeyboardController
except ImportError as e:
    print(f"❌ Fehler: Benötigte Bibliothek fehlt: {e}")
    print("🔧 Bitte installiere die Abhängigkeiten mit: pip install flask openpyxl pynput")
    input("Drücke Enter zum Beenden...")
    sys.exit(1)

# Globale Variablen
app = Flask(__name__)
app.secret_key = 'protools_track_namer_secret_key_2025'
current_data = []
current_layout = "auto"
keyboard_listener = None
server_port = 5000

def resource_path(relative_path):
    """Liefert den korrekten Pfad für normale und PyInstaller-Ausführung."""
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

def load_index_template():
    """Lädt das HTML-Template aus gebündelten Ressourcen, mit Fallback."""
    template_path = resource_path(os.path.join('templates', 'index.html'))
    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            return f.read()
    return create_template()

def find_available_port(preferred_port=5000, host='127.0.0.1'):
    """Nimmt bevorzugt Port 5000, weicht bei Belegung auf einen freien Port aus."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        if sock.connect_ex((host, preferred_port)) != 0:
            return preferred_port

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind((host, 0))
        return sock.getsockname()[1]

def activate_protools():
    """Versucht Pro Tools in den Vordergrund zu bringen und prüft das Ergebnis."""
    scripts = [
        'tell application "Pro Tools" to activate',
        (
            'tell application "System Events"\n'
            '  if exists (first application process whose name starts with "Pro Tools") then\n'
            '    set frontmost of first application process whose name starts with "Pro Tools" to true\n'
            '  end if\n'
            'end tell'
        ),
    ]

    errors = []
    for script in scripts:
        try:
            subprocess.run(['osascript', '-e', script], check=True, capture_output=True, text=True)
        except Exception as e:
            errors.append(str(e))

    front_ok, detail = ensure_protools_frontmost()
    if front_ok:
        return True, "Pro Tools wurde in den Vordergrund gebracht."

    if errors:
        return False, f"Pro Tools konnte nicht in den Vordergrund gebracht werden: {' | '.join(errors)}"
    return False, f"Pro Tools ist nicht im Vordergrund: {detail}"

def ensure_protools_frontmost():
    """Prüft, ob Pro Tools wirklich die aktive Frontmost-App ist."""
    script = (
        'tell application "System Events"\n'
        '  set frontName to name of first application process whose frontmost is true\n'
        '  return frontName\n'
        'end tell'
    )
    ok, out = _run_osascript(script)
    if not ok:
        return False, f'Frontmost-Check fehlgeschlagen: {out}'

    front_name = (out or '').strip()
    if front_name.startswith('Pro Tools'):
        return True, front_name
    return False, f'Aktive App ist "{front_name}" statt Pro Tools'

def _run_osascript(script):
    """Führt AppleScript aus und liefert (ok, stdout_or_error)."""
    try:
        result = subprocess.run(
            ['osascript', '-e', script],
            check=True,
            capture_output=True,
            text=True,
        )
        return True, (result.stdout or '').strip()
    except Exception as e:
        return False, str(e)

def _escape_applescript_text(text):
    return str(text).replace('\\', '\\\\').replace('"', '\\"')

def _get_protools_process_selector_script():
    return 'first application process whose name starts with "Pro Tools"'

def _open_new_track_dialog_ui():
    """Öffnet den New-Track-Dialog direkt über das Track-Menü (englisch/deutsch)."""
    process_selector = _get_protools_process_selector_script()
    script = (
        'tell application "System Events"\n'
        f'  if not (exists {process_selector}) then error "Pro Tools Prozess nicht gefunden"\n'
        f'  tell {process_selector}\n'
        '    set frontmost to true\n'
        '    tell menu bar item "Track" of menu bar 1\n'
        '      click\n'
        '      delay 0.1\n'
        '      if exists menu item "New..." of menu 1 then\n'
        '        click menu item "New..." of menu 1\n'
        '      else if exists menu item "Neu..." of menu 1 then\n'
        '        click menu item "Neu..." of menu 1\n'
        '      else\n'
        '        error "Menüpunkt Track > New.../Neu... nicht gefunden"\n'
        '      end if\n'
        '    end tell\n'
        '  end tell\n'
        'end tell'
    )
    return _run_osascript(script)

def _fill_and_submit_new_track_dialog_ui(track_count):
    """Setzt die Spuranzahl im New-Track-Dialog und bestätigt Create/Erstellen."""
    escaped_count = _escape_applescript_text(track_count)
    process_selector = _get_protools_process_selector_script()
    script = (
        'tell application "System Events"\n'
        f'  if not (exists {process_selector}) then error "Pro Tools Prozess nicht gefunden"\n'
        f'  tell {process_selector}\n'
        '    set frontmost to true\n'
        '    delay 0.2\n'
        '    set targetWindow to window 1\n'
        '    try\n'
        f'      set value of text field 1 of targetWindow to "{escaped_count}"\n'
        '    on error\n'
        '      keystroke "a" using {command down}\n'
        f'      keystroke "{escaped_count}"\n'
        '    end try\n'
        '    if exists button "Create" of targetWindow then\n'
        '      click button "Create" of targetWindow\n'
        '    else if exists button "Erstellen" of targetWindow then\n'
        '      click button "Erstellen" of targetWindow\n'
        '    else\n'
        '      key code 36\n'
        '    end if\n'
        '  end tell\n'
        'end tell'
    )
    return _run_osascript(script)

def _modifiers_to_applescript(modifiers):
    if not modifiers:
        return ''
    mapping = {
        'cmd': 'command down',
        'command': 'command down',
        'shift': 'shift down',
        'alt': 'option down',
        'option': 'option down',
        'ctrl': 'control down',
        'control': 'control down',
    }
    parts = [mapping[m] for m in modifiers if m in mapping]
    if not parts:
        return ''
    return ' using {' + ', '.join(parts) + '}'

def _send_applescript_keystroke(text, modifiers=None):
    escaped = _escape_applescript_text(text)
    mod_part = _modifiers_to_applescript(modifiers)
    script = (
        'tell application "System Events"\n'
        f'  keystroke "{escaped}"{mod_part}\n'
        'end tell'
    )
    try:
        subprocess.run(['osascript', '-e', script], check=True, capture_output=True, text=True)
        return True, ''
    except Exception as e:
        return False, str(e)

def _send_applescript_keycode(keycode, modifiers=None):
    mod_part = _modifiers_to_applescript(modifiers)
    script = (
        'tell application "System Events"\n'
        f'  key code {int(keycode)}{mod_part}\n'
        'end tell'
    )
    try:
        subprocess.run(['osascript', '-e', script], check=True, capture_output=True, text=True)
        return True, ''
    except Exception as e:
        return False, str(e)

def _send_keystroke_with_fallback(keyboard, text, modifiers=None):
    ok, err = _send_applescript_keystroke(text, modifiers)
    if ok:
        return True, 'AppleScript'

    try:
        modifiers = modifiers or []
        key_map = {
            'cmd': Key.cmd,
            'command': Key.cmd,
            'shift': Key.shift,
            'alt': Key.alt,
            'option': Key.alt,
            'ctrl': Key.ctrl,
            'control': Key.ctrl,
        }
        pressed = []
        for mod in modifiers:
            if mod in key_map:
                key_obj = key_map[mod]
                keyboard.press(key_obj)
                pressed.append(key_obj)

        if len(str(text)) == 1:
            keyboard.press(str(text))
            keyboard.release(str(text))
        else:
            keyboard.type(str(text))

        for key_obj in reversed(pressed):
            keyboard.release(key_obj)

        return True, 'pynput'
    except Exception as e:
        return False, f'AppleScript: {err} | pynput: {e}'

def _send_keycode_with_fallback(keyboard, keycode, modifiers=None):
    ok, err = _send_applescript_keycode(keycode, modifiers)
    if ok:
        return True, 'AppleScript'

    try:
        modifiers = modifiers or []
        key_map = {
            'cmd': Key.cmd,
            'command': Key.cmd,
            'shift': Key.shift,
            'alt': Key.alt,
            'option': Key.alt,
            'ctrl': Key.ctrl,
            'control': Key.ctrl,
        }
        pressed = []
        for mod in modifiers:
            if mod in key_map:
                key_obj = key_map[mod]
                keyboard.press(key_obj)
                pressed.append(key_obj)

        code_to_key = {
            36: Key.enter,
            124: Key.right,
        }
        key_obj = code_to_key.get(int(keycode))
        if key_obj is None:
            raise ValueError(f'Nicht unterstützter Fallback-Keycode: {keycode}')

        keyboard.press(key_obj)
        keyboard.release(key_obj)

        for mod_key in reversed(pressed):
            keyboard.release(mod_key)

        return True, 'pynput'
    except Exception as e:
        return False, f'AppleScript: {err} | pynput: {e}'

# === ALLE ROUTES UND FUNKTIONEN VON tracknamer_web.py ===

@app.route('/')
def index():
    return render_template_string(load_index_template(), 
                                data=current_data, 
                                current_layout=current_layout)

@app.route('/upload', methods=['POST'])
def upload_file():
    global current_data, current_layout
    
    if 'excel_file' not in request.files:
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('index'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('index'))
    
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        flash('Nur Excel-Dateien (.xlsx, .xls, .xlsm) erlaubt', 'error')
        return redirect(url_for('index'))
    
    try:
        current_layout = request.form.get('layout', 'auto')
        
        # Temporäre Datei erstellen
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            
            # Excel laden, bevorzugt den Tab "Inputliste" nutzen
            workbook = openpyxl.load_workbook(tmp_file.name)
            sheet = workbook['Inputliste'] if 'Inputliste' in workbook.sheetnames else workbook.active

            inputliste_header = str(sheet.cell(row=7, column=2).value or "").strip().lower()
            is_inputliste = sheet.title.strip().lower() == "inputliste" or (
                "signal" in inputliste_header and "instrument" in inputliste_header
            )

            # Layout bestimmen
            if is_inputliste:
                detected_layout = "inputliste"
            elif current_layout == "auto":
                header_d = sheet.cell(row=6, column=4).value
                detected_layout = "layout_b" if header_d and "instrument" in str(header_d).lower() else "layout_a"
            else:
                detected_layout = current_layout

            # Spalten-Indizes und Startzeile
            if is_inputliste:
                wanted_indices = [1, 2, 3]  # A=Kanal, B=Signal/Instrument, C=Mikrofon
                row_start = 8
            else:
                wanted_indices = [2, 3, 4] if detected_layout == "layout_a" else [2, 4, 5]
                row_start = 7
            
            # Daten laden
            current_data = []
            for row_num in range(row_start, min(sheet.max_row + 1, 200)):
                row_data = [sheet.cell(row=row_num, column=col_idx).value for col_idx in wanted_indices]

                if is_inputliste:
                    kanal = str(row_data[0]).strip() if row_data[0] is not None else ""
                    if kanal:
                        try:
                            kanal_float = float(kanal.replace(",", "."))
                            kanal = str(int(kanal_float)) if kanal_float.is_integer() else kanal
                        except ValueError:
                            pass
                    instrument = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else ""
                    mikrofon = str(row_data[2]).strip() if len(row_data) > 2 and row_data[2] else ""
                else:
                    if not row_data[1]:
                        continue

                    kanal_text = str(row_data[0]).strip() if row_data[0] is not None else ""
                    if kanal_text:
                        try:
                            kanal_float = float(kanal_text.replace(",", "."))
                            kanal = str(int(kanal_float)) if kanal_float.is_integer() else kanal_text
                        except ValueError:
                            kanal = kanal_text
                    else:
                        kanal = ""

                    instrument = str(row_data[1]).strip() if row_data[1] else ""
                    mikrofon = str(row_data[2]).strip() if row_data[2] else ""

                if instrument:
                    track_parts = [part for part in [kanal, instrument, mikrofon] if part]
                    trackname = "_".join(track_parts)
                    
                    current_data.append({
                        'id': len(current_data),
                        'kanal': kanal,
                        'instrument': instrument,
                        'mikrofon': mikrofon,
                        'trackname': trackname,
                        'selected': True
                    })
            
            current_layout = detected_layout
            os.unlink(tmp_file.name)
            
        flash(f'✅ Excel-Datei erfolgreich geladen! {len(current_data)} Zeilen gefunden.', 'success')
        
    except Exception as e:
        flash(f'❌ Fehler beim Laden der Excel-Datei: {str(e)}', 'error')
        
    return redirect(url_for('index'))

@app.route('/export', methods=['POST'])
def export_tracks():
    if not current_data:
        return "Keine Daten zum Exportieren", 404
        
    selected_ids = request.form.getlist('selected_tracks')
    selected_ids = [int(id_) for id_ in selected_ids]
    
    if not selected_ids:
        flash('Bitte mindestens eine Spur auswählen', 'error')
        return redirect(url_for('index'))

@app.route('/select_all')
def select_all():
    return jsonify({'action': 'select_all'})

@app.route('/select_none')
def select_none():
    return jsonify({'action': 'select_none'})

@app.route('/manual_instructions')
def manual_instructions():
    show_manual_instructions()
    return jsonify({'success': '📋 Manuelle Anleitung im Terminal angezeigt'})

@app.route('/protools_create_tracks', methods=['POST'])
def protools_create_tracks():
    """Pro Tools: Spuren erstellen - Browser-Button Version"""
    try:
        print("🔥 Browser-Button 'Spuren erstellen' geklickt!")
        
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            print("❌ Keine Spuren ausgewählt")
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        track_count = len(selected_ids)
        print(f"🎯 {track_count} Spuren werden erstellt (nur ausgewählte!)")
        
        # Wie in der alten funktionierenden Version: kein erzwungenes Aktivieren,
        # stattdessen kurze Zeit für manuellen Fokuswechsel.
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
        time.sleep(3)
        
        # Korrekte Anzahl verwenden (nur ausgewählte Spuren)
        ok, detail = create_tracks_correct(track_count)
        if not ok:
            return jsonify({'error': detail})
        
        return jsonify({
            'success': f'✅ {track_count} Spuren sollten erstellt sein! ({detail})'
        })
        
    except Exception as e:
        print(f"❌ Browser-Button Fehler: {e}")
        return jsonify({'error': f'Pro Tools Fehler: {e}'})

@app.route('/protools_name_tracks', methods=['POST'])
def protools_name_tracks():
    """Pro Tools: Spuren benennen - Browser-Button Version"""
    try:
        print("🔥 Browser-Button 'Spuren benennen' geklickt!")
        
        # Ausgewählte IDs sammeln
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            print("❌ Keine Spuren ausgewählt")
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        # Namensformat-Optionen sammeln
        include_channel = 'include_channel' in request.form
        include_instrument = 'include_instrument' in request.form  
        include_microphone = 'include_microphone' in request.form
        
        print(f"🎯 {len(selected_ids)} ausgewählte Spuren werden benannt")
        print(f"📝 Format: Kanal={include_channel}, Instrument={include_instrument}, Mikrofon={include_microphone}")
        
        # Wie in der alten funktionierenden Version: kein erzwungenes Aktivieren,
        # stattdessen kurze Zeit für manuellen Fokuswechsel.
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
        time.sleep(3)
        print("💡 Markieren Sie die erste Spur in Pro Tools!")
        
        # Korrekte Funktion verwenden
        ok, detail = name_tracks_correct(selected_ids, include_channel, include_instrument, include_microphone)
        if not ok:
            return jsonify({'error': detail})
        
        return jsonify({
            'success': f'✅ {len(selected_ids)} Spuren sollten benannt sein! ({detail})'
        })
        
    except Exception as e:
        print(f"❌ Browser-Button Fehler: {e}")
        return jsonify({'error': f'Pro Tools Fehler: {e}'})

# === ALLE HILFSFUNKTIONEN ===

def create_tracks_correct(track_count):
    """Spuren erstellen - Korrekte Anzahl basierend auf Auswahl"""
    try:
        if track_count <= 0:
            return False, 'Keine Spuren ausgewählt.'

        import time
        keyboard = KeyboardController()
        
        print(f"🏗️ Erstelle {track_count} Spuren in Pro Tools...")
        print("🎯 Schnelle Methode: Dialog öffnen + Anzahl sofort eingeben")
        
        # Bewusst wieder der alte, funktionierende Shortcut-Weg.
        keyboard.press(Key.cmd)
        keyboard.press(Key.shift)
        keyboard.press('n')
        keyboard.release('n')
        keyboard.release(Key.shift)
        keyboard.release(Key.cmd)
        
        # Minimal warten und sofort Anzahl eingeben
        time.sleep(0.6)  # Noch kürzer
        
        # Anzahl sofort eingeben (überschreibt Standard-Wert)
        track_str = str(track_count)
        print(f"   Eingabe sofort: {track_str}")
        
        keyboard.press(Key.cmd)
        keyboard.press('a')
        keyboard.release('a')
        keyboard.release(Key.cmd)

        keyboard.type(track_str)

        time.sleep(0.1)
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        
        print(f"✅ {track_count} Spuren sollten sofort erstellt sein!")
        print("💡 Optimiert: Minimale Wartezeit, sofortige Eingabe")
        return True, 'Siehe Pro Tools / New Track Dialog.'
        
    except Exception as e:
        error_text = str(e)
        print(f"❌ Fehler beim Erstellen der Spuren: {error_text}")
        if 'not trusted' in error_text.lower() or 'not permitted' in error_text.lower():
            return False, 'macOS blockiert Tastatursteuerung. Bitte Bedienungshilfen/Eingabeüberwachung für die App erlauben.'
        return False, f'Fehler beim Erstellen der Spuren: {error_text}'

def name_tracks_correct(selected_ids, include_channel, include_instrument, include_microphone):
    """Spuren benennen - Korrekt mit ausgewählten IDs und Namensformat"""
    try:
        if not current_data:
            print("❌ Keine Daten geladen")
            return False, 'Keine Daten geladen.'
        
        # Nur die wirklich ausgewählten Items holen
        selected_items = [item for item in current_data if item['id'] in selected_ids]
        
        if not selected_items:
            print("❌ Keine Spuren ausgewählt")
            return False, 'Keine Spuren ausgewählt.'
        
        import time
        keyboard = KeyboardController()
        
        print(f"🏷️ Benenne {len(selected_items)} Spuren in Pro Tools...")
        print("📋 ANLEITUNG:")
        print("   1. Öffne MANUELL den Namensgebungs-Dialog der ersten Spur")
        print("   2. Die App übernimmt dann automatisch!")
        print("⏳ 2 Sekunden Zeit zum Öffnen des ersten Dialogs...")
        time.sleep(2)
        
        # Jede Spur benennen mit korrektem Format
        for i, item in enumerate(selected_items):
            # Namen nach Auswahl zusammensetzen
            name_parts = []
            if include_channel and item['kanal']:
                name_parts.append(str(item['kanal']))
            if include_instrument and item['instrument']:
                name_parts.append(str(item['instrument']))
            if include_microphone and item['mikrofon']:
                name_parts.append(str(item['mikrofon']))
            
            # Namen zusammenbauen
            if name_parts:
                name = "_".join(name_parts)
            else:
                # Fallback: Wenn alle Felder leer sind oder nichts ausgewählt
                if item['instrument']:
                    name = str(item['instrument'])
                else:
                    name = f"Track_{i+1}"
            
            print(f"  🎯 Spur {i+1}/{len(selected_items)}: {name}")
            
            # Namen eingeben (überschreibt aktuellen Inhalt)
            keyboard.press(Key.cmd)
            keyboard.press('a')
            keyboard.release('a')
            keyboard.release(Key.cmd)
            time.sleep(0.03)  # Ultra kurz
            
            # Neuen Namen eingeben
            keyboard.type(name)
            time.sleep(0.05)  # Ultra kurz
            
            # Zur nächsten Spur wechseln (außer bei der letzten)
            if i < len(selected_items) - 1:
                keyboard.press(Key.cmd)
                keyboard.press(Key.right)
                keyboard.release(Key.right)
                keyboard.release(Key.cmd)
                time.sleep(0.08)  # Ultra kurz
        
        # Enter am Ende drücken um letzte Spur zu bestätigen
        print("  🎯 Bestätige letzte Spur mit Enter...")
        time.sleep(0.1)
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        
        print(f"✅ Alle {len(selected_items)} Spuren benannt und bestätigt!")
        print("🎉 Namensgebung komplett abgeschlossen!")
        return True, 'Namensgebung wurde an Pro Tools gesendet.'
        
    except Exception as e:
        error_text = str(e)
        print(f"❌ Fehler beim Benennen der Spuren: {error_text}")
        if 'not trusted' in error_text.lower() or 'not permitted' in error_text.lower():
            return False, 'macOS blockiert Tastatursteuerung. Bitte Bedienungshilfen/Eingabeüberwachung für die App erlauben.'
        return False, f'Fehler beim Benennen der Spuren: {error_text}'

def show_manual_instructions():
    """Zeige manuelle Anleitung für Pro Tools"""
    if not current_data:
        print("❌ Keine Daten geladen")
        return
        
    selected_count = sum(1 for item in current_data if item.get('selected', True))
    
    print("\\n" + "="*60)
    print("📋 MANUELLE PRO TOOLS ANLEITUNG")
    print("="*60)
    print(f"🎯 Zu erstellende Spuren: {selected_count}")
    
    print("\\n🏗️ SPUREN ERSTELLEN:")
    print("1. Gehe zu: Track → New...")
    print("2. Audio Tracks: Mono")
    print(f"3. Anzahl: {selected_count}")
    print("4. Klicke 'Create'")
    
    print("\\n🏷️ SPUREN BENENNEN:")
    for i, item in enumerate([item for item in current_data if item.get('selected', True)]):
        print(f"   Spur {i+1}: {item['trackname']}")
    
    print("\\n💡 Tipp: Erste Spur markieren, dann F9 für automatische Benennung")
    print("\\n⌨️ TASTENKOMBINATIONEN:")
    print("   F8 = Spuren erstellen")
    print("   F9 = Spuren benennen")
    print("   F10 = Diese Anleitung")
    print("="*60)

def create_template():
    """Fallback-Template, falls keine gebündelte Datei verfügbar ist."""
    
    html_content = '''<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🎵 Pro Tools Track Namer</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { text-align: center; margin-bottom: 30px; }
        .upload-section { background: #e3f2fd; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
        .layout-options { margin: 10px 0; }
        .layout-options input { margin-right: 5px; }
        .data-table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        .data-table th, .data-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        .data-table th { background-color: #f2f2f2; font-weight: bold; }
        .data-table tr:nth-child(even) { background-color: #f9f9f9; }
        .data-table tr.row-selected { background-color: #dff3ff !important; }
        .trackname { font-family: 'Courier New', monospace; font-weight: bold; color: #2196F3; }
        .buttons { margin: 20px 0; text-align: center; }
        .buttons button { margin: 0 10px; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; font-size: 14px; }
        .btn-primary { background-color: #2196F3; color: white; }
        .btn-success { background-color: #4CAF50; color: white; }
        .btn-warning { background-color: #FF9800; color: white; }
        .btn-danger { background-color: #f44336; color: white; }
        .flash-messages { margin: 20px 0; }
        .flash-success { background: #d4edda; color: #155724; padding: 10px; border-radius: 5px; border: 1px solid #c3e6cb; }
        .flash-error { background: #f8d7da; color: #721c24; padding: 10px; border-radius: 5px; border: 1px solid #f5c6cb; }
        .upload-btn { background-color: #2196F3; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }
        input[type="file"] { margin: 10px 0; }
        .stats { background: #e8f5e9; padding: 10px; border-radius: 5px; margin: 10px 0; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎵 Pro Tools Track Namer</h1>
            <p>Standalone Version - Automatische Spurerstellung und -benennung</p>
        </div>
        
        <!-- Flash Messages -->
        <div class="flash-messages">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="flash-{{ 'success' if category == 'success' else 'error' }}">
                            {{ message }}
                        </div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
        </div>
        
        <!-- Upload Section -->
        <div class="upload-section">
            <h3>📤 Excel-Datei hochladen</h3>
            <form method="post" action="/upload" enctype="multipart/form-data">
                <input type="file" name="excel_file" accept=".xlsx,.xls,.xlsm" required>
                
                <div class="layout-options">
                    <h4>Layout-Erkennung:</h4>
                    <label><input type="radio" name="layout" value="auto" checked> 🤖 Automatisch</label>
                    <label><input type="radio" name="layout" value="layout_a"> Layout A (B,C,D)</label>
                    <label><input type="radio" name="layout" value="layout_b"> Layout B (B,D,E)</label>
                </div>
                
                <button type="submit" class="upload-btn">📁 Excel-Datei laden</button>
            </form>
        </div>
        
        {% if data %}
        <div class="stats">
            <strong>📊 Geladene Daten:</strong> {{ data|length }} Zeilen | 
            <strong>Layout:</strong> {{ current_layout }}
        </div>
        
        <!-- Selection Buttons -->
        <div class="buttons">
            <button class="btn-success" onclick="selectAll()">✅ Alle auswählen</button>
            <button class="btn-warning" onclick="selectNone()">❌ Alle abwählen</button>
        </div>
        
        <!-- Data Table -->
        <form method="post" action="/export">
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Auswahl</th>
                        <th>Kanal</th>
                        <th>Instrument</th>
                        <th>Mikrofon</th>
                        <th>Trackname</th>
                    </tr>
                </thead>
                <tbody>
                    {% for row in data %}
                    <tr>
                        <td><input type="checkbox" name="selected_tracks" value="{{ row.id }}" {% if row.selected %}checked{% endif %}></td>
                        <td>{{ row.kanal }}</td>
                        <td>{{ row.instrument }}</td>
                        <td>{{ row.mikrofon }}</td>
                        <td class="trackname">{{ row.trackname }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
            
            <!-- Namensformat-Auswahl -->
            <div style="background: #e8f5e9; padding: 15px; border-radius: 8px; margin: 20px 0; border: 1px solid #4caf50;">
                <h3>🏷️ Namensformat wählen</h3>
                <div style="display: flex; gap: 20px; flex-wrap: wrap;">
                    <label><input type="checkbox" name="include_channel" checked> Spurnummer (z.B. "1_")</label>
                    <label><input type="checkbox" name="include_instrument" checked> Instrument (z.B. "Kick")</label>  
                    <label><input type="checkbox" name="include_microphone" checked> Mikrofon (z.B. "SM57")</label>
                </div>
                <p style="margin-top: 10px; font-size: 12px; color: #666;">
                    <strong>Beispiel:</strong> <span id="name-preview">1_Kick_SM57</span>
                </p>
            </div>
            
            <div class="buttons">
                <button type="submit" class="btn-primary">📄 Tracknamen als Textdatei exportieren</button>
            </div>
        </form>
        
        <!-- Pro Tools Integration -->
        <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0; border: 1px solid #ffeaa7;">
            <h3>🎵 Pro Tools Integration</h3>
            <p><strong>Workflow:</strong> Erst Spuren erstellen, dann benennen</p>
            <p><strong>⌨️ Tastenkombinationen:</strong> F8 = Erstellen | F9 = Benennen | F10 = Anleitung</p>
            
            <div class="buttons">
                <button class="btn-success" onclick="protoolsCreateTracks()" id="createBtn">
                    🏗️ Spuren erstellen
                </button>
                <button class="btn-warning" onclick="protoolsNameTracks()" id="nameBtn">
                    🏷️ Spuren benennen
                </button>
                <button class="btn-primary" onclick="showManualInstructions()" id="manualBtn">
                    📋 Anleitung
                </button>
            </div>
            
            <div id="protools-status" style="margin-top: 10px; padding: 10px; border-radius: 5px; display: none;"></div>
        </div>
        {% else %}
        <div class="stats">
            <strong>ℹ️ Anleitung:</strong> Laden Sie eine Excel-Datei (.xlsm, .xlsx, .xls), um zu beginnen.
        </div>
        {% endif %}
    </div>
    
    <script>
        let lastTrackAnchorIndex = -1;
        let trackSelectionInitialized = false;

        function getTrackCheckboxes() {
            return Array.from(document.querySelectorAll('input[name="selected_tracks"]'));
        }

        function updateSelectedRowStyles() {
            getTrackCheckboxes().forEach(cb => {
                const row = cb.closest('tr');
                if (row) {
                    row.classList.toggle('row-selected', cb.checked);
                }
            });
        }

        function applyRangeSelection(checkboxes, startIndex, endIndex, shouldCheck) {
            const from = Math.min(startIndex, endIndex);
            const to = Math.max(startIndex, endIndex);
            for (let i = from; i <= to; i++) {
                checkboxes[i].checked = shouldCheck;
            }
        }

        function initializeTrackSelection() {
            if (trackSelectionInitialized) {
                updateSelectedRowStyles();
                return;
            }

            const tbody = document.querySelector('.data-table tbody');
            if (!tbody) {
                return;
            }

            trackSelectionInitialized = true;

            tbody.addEventListener('click', event => {
                const row = event.target.closest('tr');
                if (!row) {
                    return;
                }

                let checkbox = event.target.closest('input[name="selected_tracks"]');

                if (!checkbox) {
                    if (event.target.closest('button,a,input,select,textarea,label')) {
                        return;
                    }
                    checkbox = row.querySelector('input[name="selected_tracks"]');
                    if (!checkbox) {
                        return;
                    }
                    checkbox.checked = !checkbox.checked;
                }

                const checkboxes = getTrackCheckboxes();
                const currentIndex = checkboxes.indexOf(checkbox);
                if (currentIndex === -1) {
                    return;
                }

                if (event.shiftKey && lastTrackAnchorIndex !== -1) {
                    applyRangeSelection(checkboxes, lastTrackAnchorIndex, currentIndex, checkbox.checked);
                }

                lastTrackAnchorIndex = currentIndex;
                updateSelectedRowStyles();
            });

            updateSelectedRowStyles();
        }

        function selectAll() {
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]');
            checkboxes.forEach(cb => cb.checked = true);
            updateSelectedRowStyles();
        }
        
        function selectNone() {
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]');
            checkboxes.forEach(cb => cb.checked = false);
            updateSelectedRowStyles();
        }
        
        function getSelectedTracks() {
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]:checked');
            return Array.from(checkboxes).map(cb => cb.value);
        }
        
        function protoolsCreateTracks() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showProtoolsStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showProtoolsStatus('⏳ Erstelle ' + selected.length + ' Spuren in Pro Tools...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/protools_create_tracks', {
                method: 'POST', 
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showProtoolsStatus(data.error, true);
                } else {
                    showProtoolsStatus(data.success);
                }
            })
            .catch(error => {
                showProtoolsStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function protoolsNameTracks() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showProtoolsStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showProtoolsStatus('⏳ Benenne ' + selected.length + ' Spuren in Pro Tools...');
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            // Namensformat-Checkboxen hinzufügen
            const includeChannel = document.querySelector('input[name="include_channel"]').checked;
            const includeInstrument = document.querySelector('input[name="include_instrument"]').checked;
            const includeMicrophone = document.querySelector('input[name="include_microphone"]').checked;
            
            if (includeChannel) formData.append('include_channel', 'on');
            if (includeInstrument) formData.append('include_instrument', 'on');
            if (includeMicrophone) formData.append('include_microphone', 'on');
            
            fetch('/protools_name_tracks', {
                method: 'POST', 
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showProtoolsStatus(data.error, true);
                } else {
                    showProtoolsStatus(data.success);
                }
            })
            .catch(error => {
                showProtoolsStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function showManualInstructions() {
            showProtoolsStatus('📋 Manuelle Anleitung im Terminal angezeigt - schauen Sie ins Terminal!');
            
            // Einfacher GET-Request um F10-Funktion auszulösen
            fetch('/manual_instructions', { method: 'GET' })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        showProtoolsStatus(data.success);
                    }
                })
                .catch(error => {
                    showProtoolsStatus('❌ Netzwerk-Fehler: ' + error, true);
                });
        }

        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', initializeTrackSelection);
        } else {
            initializeTrackSelection();
        }
        
        function showProtoolsStatus(message, isError = false) {
            const statusDiv = document.getElementById('protools-status');
            statusDiv.textContent = message;
            statusDiv.style.display = 'block';
            statusDiv.style.backgroundColor = isError ? '#f8d7da' : '#d4edda';
            statusDiv.style.color = isError ? '#721c24' : '#155724';
            statusDiv.style.border = isError ? '1px solid #f5c6cb' : '1px solid #c3e6cb';
            
            // Status nach 5 Sekunden ausblenden
            setTimeout(() => {
                statusDiv.style.display = 'none';
            }, 5000);
        }
    </script>
</body>
</html>'''
    
    return html_content

# === STANDALONE APP FUNKTIONEN ===

def open_browser():
    """Öffnet den Browser nach kurzer Verzögerung"""
    webbrowser.open(f'http://127.0.0.1:{server_port}')
    print(f"🌐 Browser sollte sich automatisch öffnen: http://127.0.0.1:{server_port}")

def signal_handler(sig, frame):
    """Behandelt Ctrl+C für sauberes Beenden"""
    print("\\n🛑 Pro Tools Track Namer wird beendet...")
    stop_keyboard_listener()
    sys.exit(0)

# === TASTENKOMBINATIONS-FUNKTIONEN ===

def on_key_press(key):
    """Behandelt Tastendruck-Events"""
    try:
        if hasattr(key, 'name'):
            if key.name == 'f8':
                print("\\n🏗️ F8 gedrückt: Erstelle Spuren in Pro Tools...")
                create_tracks_from_hotkey()
            elif key.name == 'f9':
                print("\\n🏷️ F9 gedrückt: Benenne Spuren in Pro Tools...")
                name_tracks_from_hotkey()
            elif key.name == 'f10':
                print("\\n📋 F10 gedrückt: Zeige manuelle Anleitung...")
                show_manual_instructions()
    except AttributeError:
        pass

def create_tracks_from_hotkey():
    """Erstellt Spuren über Tastenkombination F8"""
    if not current_data:
        print("❌ Keine Daten geladen! Bitte zuerst Excel-Datei laden.")
        return
    
    # Alle als ausgewählt markierte Spuren zählen
    selected_items = [item for item in current_data if item.get('selected', True)]
    track_count = len(selected_items)
    
    if track_count == 0:
        print("❌ Keine Spuren zum Erstellen gefunden!")
        return
    
    print(f"🎯 {track_count} Spuren werden erstellt")
    print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
    time.sleep(3)
    
    create_tracks_correct(track_count)

def name_tracks_from_hotkey():
    """Benennt Spuren über Tastenkombination F9"""
    if not current_data:
        print("❌ Keine Daten geladen! Bitte zuerst Excel-Datei laden.")
        return
    
    # Alle als ausgewählt markierte Spuren sammeln
    selected_items = [item for item in current_data if item.get('selected', True)]
    
    if not selected_items:
        print("❌ Keine Spuren zum Benennen gefunden!")
        return
    
    # Standard-Format: Alle Komponenten einbeziehen
    selected_ids = [item['id'] for item in selected_items]
    
    print(f"🎯 {len(selected_items)} Spuren werden benannt")
    print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
    time.sleep(3)
    
    # Standard-Namensformat: Kanal + Instrument + Mikrofon
    name_tracks_correct(selected_ids, True, True, True)

def start_keyboard_listener():
    """Startet den Keyboard-Listener für Hotkeys"""
    global keyboard_listener
    try:
        keyboard_listener = keyboard.Listener(on_press=on_key_press)
        keyboard_listener.start()
        print("⌨️ Tastenkombinationen aktiviert:")
        print("   F8 = Spuren erstellen")
        print("   F9 = Spuren benennen") 
        print("   F10 = Manuelle Anleitung")
    except Exception as e:
        print(f"⚠️ Keyboard-Listener konnte nicht gestartet werden: {e}")

def stop_keyboard_listener():
    """Stoppt den Keyboard-Listener"""
    global keyboard_listener
    if keyboard_listener:
        keyboard_listener.stop()
        keyboard_listener = None

def main():
    """Hauptfunktion für standalone App"""
    global server_port, current_data

    print("🎵 Pro Tools Track Namer - Standalone Version")
    print("=" * 50)
    print("🚀 Starte Server...")

    # Immer mit leerem Zustand starten
    current_data = []

    server_port = find_available_port(5000)
    if server_port != 5000:
        print(f"⚠️ Port 5000 ist belegt. Nutze stattdessen Port {server_port}.")
    
    # Ctrl+C Handler
    signal.signal(signal.SIGINT, signal_handler)
    
    # Keyboard-Listener starten
    start_keyboard_listener()
    
    # Browser nach 2 Sekunden öffnen
    Timer(2.0, open_browser).start()
    
    # Server starten
    try:
        app.run(host='127.0.0.1', port=server_port, debug=False)
    except Exception as e:
        print(f"❌ Fehler beim Starten: {e}")
        input("Drücke Enter zum Beenden...")
    finally:
        stop_keyboard_listener()

if __name__ == "__main__":
    main()