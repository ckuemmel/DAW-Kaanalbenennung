#!/usr/bin/env python3
"""
Multi-DAW Track Namer - Unterstützt Pro Tools und Reaper
Web-basierte Version mit DAW-Auswahl
"""
import os
import time
from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify
import openpyxl
import tempfile
from werkzeug.utils import secure_filename
import webbrowser
import threading

app = Flask(__name__)
app.secret_key = 'tracknamer_multiDAW_secret_key'

# Globale Variablen
current_data = []
current_layout = "auto"
selected_daw = "protools"  # Standard: Pro Tools
automation_lock = threading.Lock()
hotkeys_lock = threading.Lock()
hotkey_listener = None


def build_selected_items(selected_ids, include_channel=True, include_instrument=True, include_microphone=True):
    """Erzeuge ausgewählte Track-Items mit stabilem Namensformat."""
    selected_items = []

    for track_id in selected_ids:
        if track_id < 0 or track_id >= len(current_data):
            continue

        track_data = current_data[track_id]

        parts = []
        if include_channel and track_data.get('kanal'):
            parts.append(str(track_data['kanal']))
        if include_instrument and track_data.get('instrument'):
            parts.append(str(track_data['instrument']))
        if include_microphone and track_data.get('mikrofon'):
            parts.append(str(track_data['mikrofon']))

        custom_trackname = ' '.join(parts) if parts else f"Track {track_id + 1}"

        selected_items.append({
            'id': track_id,
            'kanal': track_data.get('kanal', ''),
            'instrument': track_data.get('instrument', ''),
            'mikrofon': track_data.get('mikrofon', ''),
            'trackname': custom_trackname,
        })

    return selected_items

@app.route('/')
def index():
    return render_template('index.html', data=current_data, selected_daw=selected_daw)

@app.route('/upload', methods=['POST'])
def upload_excel():
    global current_data, current_layout
    
    if 'excel_file' not in request.files:
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('index'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('index'))
    
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        flash('Nur Excel-Dateien (.xlsx, .xls, .xlsm) erlaubt', 'error')
        return redirect(url_for('index'))
    
    try:
        current_layout = request.form.get('layout', 'auto')
        
        # Temporäre Datei erstellen und sofort schließen (wichtig für Windows)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_path = tmp_file.name
            file.save(tmp_path)
        # with-Block ist hier beendet – File-Handle ist geschlossen
        
        try:
            # Excel laden, bevorzugt den Tab "Inputliste" nutzen
            workbook = openpyxl.load_workbook(tmp_path)
            sheet = workbook['Inputliste'] if 'Inputliste' in workbook.sheetnames else workbook.active

            inputliste_header = str(sheet.cell(row=7, column=2).value or "").strip().lower()
            is_inputliste = sheet.title.strip().lower() == "inputliste" or (
                "signal" in inputliste_header and "instrument" in inputliste_header
            )

            # Layout bestimmen
            if is_inputliste:
                detected_layout = "inputliste"
            elif current_layout == "auto":
                header_d = sheet.cell(row=6, column=4).value
                detected_layout = "layout_b" if header_d and "instrument" in str(header_d).lower() else "layout_a"
            else:
                detected_layout = current_layout

            # Spalten-Indizes und Startzeile
            if is_inputliste:
                wanted_indices = [1, 2, 3]  # A=Kanal, B=Signal/Instrument, C=Mikrofon
                row_start = 8
            else:
                wanted_indices = [2, 3, 4] if detected_layout == "layout_a" else [2, 4, 5]
                row_start = 7
            
            # Daten laden
            current_data = []
            for row_num in range(row_start, min(sheet.max_row + 1, 200)):
                row_data = [sheet.cell(row=row_num, column=col_idx).value for col_idx in wanted_indices]

                if is_inputliste:
                    kanal = str(row_data[0]).strip() if row_data[0] is not None else ""
                    if kanal:
                        try:
                            kanal_float = float(kanal.replace(",", "."))
                            kanal = str(int(kanal_float)) if kanal_float.is_integer() else kanal
                        except ValueError:
                            pass
                    instrument = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else ""
                    mikrofon = str(row_data[2]).strip() if len(row_data) > 2 and row_data[2] else ""
                else:
                    if not row_data[1]:
                        continue

                    kanal_text = str(row_data[0]).strip() if row_data[0] is not None else ""
                    if kanal_text:
                        try:
                            kanal_float = float(kanal_text.replace(",", "."))
                            kanal = str(int(kanal_float)) if kanal_float.is_integer() else kanal_text
                        except ValueError:
                            kanal = kanal_text
                    else:
                        kanal = ""

                    # Bereinige Instrument und Mikrofon von führenden/nachfolgenden Leerzeichen
                    instrument = str(row_data[1]).strip() if row_data[1] else ""
                    mikrofon = str(row_data[2]).strip() if row_data[2] else ""
                    
                    # Doppelte Einträge und beschädigte Daten erkennen und bereinigen
                    def clean_text(text):
                        if not text:
                            return text
                        # Entferne übermäßige Wiederholungen
                        words = text.split()
                        if len(words) > 10:  # Zu viele Wörter, wahrscheinlich beschädigt
                            # Nimm nur die ersten sinnvollen Wörter
                            return ' '.join(words[:3])
                        return text
                    
                    instrument = clean_text(instrument)
                    mikrofon = clean_text(mikrofon)
                    
                    # Sonderzeichen-Unterstützung: Bindestriche, Schrägstriche, etc. bleiben erhalten
                    track_parts = [part for part in [kanal, instrument, mikrofon] if part]
                    trackname = "_".join(track_parts)
                    
                    current_data.append({
                        'id': len(current_data),
                        'kanal': kanal,
                        'instrument': instrument,
                        'mikrofon': mikrofon,
                        'trackname': trackname,
                        'selected': True
                    })
            
            workbook.close()
            flash(f'✅ {len(current_data)} Zeilen geladen ({detected_layout})', 'success')
        finally:
            # Temporäre Datei löschen (auch im Fehlerfall)
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            
    except Exception as e:
        flash(f'Fehler beim Laden: {e}', 'error')
    
    return redirect(url_for('index'))

@app.route('/select_daw', methods=['POST'])
def select_daw():
    """DAW auswählen"""
    global selected_daw
    selected_daw = request.form.get('daw', 'protools')
    print(f"🔍 DEBUG: DAW gewechselt zu: '{selected_daw}'")  # DEBUG
    return jsonify({'success': f'DAW gewechselt zu: {selected_daw.upper()}'})

@app.route('/create_tracks', methods=['POST'])
def create_tracks():
    """Spuren erstellen - Multi-DAW"""
    try:
        print(f"🔍 DEBUG: selected_daw = '{selected_daw}'")  # DEBUG
        
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        track_count = len(selected_ids)
        
        if selected_daw == "protools":
            return create_tracks_protools(track_count)
        elif selected_daw == "reaper":
            return create_tracks_reaper(track_count)
        else:
            return jsonify({'error': f'Unbekannte DAW: {selected_daw}'})
            
    except Exception as e:
        return jsonify({'error': f'Fehler: {str(e)}'})

@app.route('/name_tracks', methods=['POST'])
def name_tracks():
    """Spuren benennen - Multi-DAW"""
    try:
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        # Namensformat-Optionen auslesen
        include_channel = request.form.get('include_channel') == 'on'
        include_instrument = request.form.get('include_instrument') == 'on'
        include_microphone = request.form.get('include_microphone') == 'on'
        
        # Ausgewählte Track-Namen sammeln und anpassen
        selected_items = []
        for item in current_data:
            if item['id'] in selected_ids:
                # Angepassten Tracknamen erstellen
                name_parts = []
                if include_channel:
                    name_parts.append(item['kanal'])
                if include_instrument:
                    name_parts.append(item['instrument'])
                if include_microphone:
                    name_parts.append(item['mikrofon'])
                
                # Trennzeichen: Leerzeichen zwischen allen Teilen
                custom_trackname = " ".join(name_parts) if name_parts else item['trackname']
                
                selected_items.append({
                    'id': item['id'],
                    'kanal': item['kanal'],
                    'instrument': item['instrument'],
                    'mikrofon': item['mikrofon'],
                    'trackname': custom_trackname
                })
        
        if selected_daw == "protools":
            return name_tracks_protools(selected_items)
        elif selected_daw == "reaper":
            return name_tracks_reaper(selected_items)
        else:
            return jsonify({'error': f'Unbekannte DAW: {selected_daw}'})
            
    except Exception as e:
        return jsonify({'error': f'Fehler: {str(e)}'})

# Neue direkte DAW-Routes
@app.route('/create_tracks_protools', methods=['POST'])
def create_tracks_protools_route():
    """Pro Tools: Spuren erstellen"""
    try:
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        track_count = len(selected_ids)
        return create_tracks_protools(track_count)
            
    except Exception as e:
        return jsonify({'error': f'Fehler: {str(e)}'})

@app.route('/create_tracks_reaper', methods=['POST'])
def create_tracks_reaper_route():
    """Reaper: Spuren erstellen"""
    try:
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})
        
        track_count = len(selected_ids)
        return create_tracks_reaper(track_count)
            
    except Exception as e:
        return jsonify({'error': f'Fehler: {str(e)}'})

@app.route('/name_tracks_protools', methods=['POST'])
def name_tracks_protools_route():
    """Pro Tools: Spuren benennen"""
    try:
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})

        include_channel = request.form.get('include_channel') == 'on'
        include_instrument = request.form.get('include_instrument') == 'on'
        include_microphone = request.form.get('include_microphone') == 'on'

        # Wenn keine Optionen übergeben wurden (Direktbutton), nutze alle Komponenten.
        if (
            'include_channel' not in request.form
            and 'include_instrument' not in request.form
            and 'include_microphone' not in request.form
        ):
            include_channel = True
            include_instrument = True
            include_microphone = True

        selected_items = build_selected_items(
            selected_ids,
            include_channel=include_channel,
            include_instrument=include_instrument,
            include_microphone=include_microphone,
        )

        if not selected_items:
            return jsonify({'error': 'Keine gültigen Spuren ausgewählt'})
        
        return name_tracks_protools(selected_items)
            
    except Exception as e:
        return jsonify({'error': f'Fehler: {str(e)}'})

@app.route('/name_tracks_reaper', methods=['POST'])
def name_tracks_reaper_route():
    """Reaper: Spuren benennen"""
    try:
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'})

        include_channel = request.form.get('include_channel') == 'on'
        include_instrument = request.form.get('include_instrument') == 'on'
        include_microphone = request.form.get('include_microphone') == 'on'

        # Wenn keine Optionen übergeben wurden (Direktbutton), nutze alle Komponenten.
        if (
            'include_channel' not in request.form
            and 'include_instrument' not in request.form
            and 'include_microphone' not in request.form
        ):
            include_channel = True
            include_instrument = True
            include_microphone = True

        selected_items = build_selected_items(
            selected_ids,
            include_channel=include_channel,
            include_instrument=include_instrument,
            include_microphone=include_microphone,
        )

        if not selected_items:
            return jsonify({'error': 'Keine gültigen Spuren ausgewählt'})
        
        return name_tracks_reaper(selected_items)
            
    except Exception as e:
        return jsonify({'error': f'Fehler: {str(e)}'})

def create_tracks_protools(track_count):
    """Pro Tools: Spuren erstellen"""
    if not automation_lock.acquire(blocking=False):
        return jsonify({'error': '⚠️ Ein Makro läuft bereits. Bitte warten, bis es fertig ist.'})

    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        print(f"🎵 Erstelle {track_count} Spuren in Pro Tools...")
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Pro Tools...")
        time.sleep(3)
        
        # Pro Tools New Track Dialog: Cmd+Shift+N (macOS) oder Ctrl+Shift+N (Windows)
        keyboard.press(Key.cmd)
        keyboard.press(Key.shift)
        keyboard.press('n')
        keyboard.release('n')
        keyboard.release(Key.shift)
        keyboard.release(Key.cmd)
        
        time.sleep(1.5)
        
        # Anzahl eingeben
        keyboard.type(str(track_count))
        time.sleep(0.5)
        
        # Enter drücken
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        
        return jsonify({'success': f'✅ {track_count} Pro Tools Spuren erstellt!'})
        
    except Exception as e:
        return jsonify({'error': f'Pro Tools Fehler: {e}'})
    finally:
        automation_lock.release()

def create_tracks_reaper(track_count):
    """Reaper: Spuren erstellen"""
    if not automation_lock.acquire(blocking=False):
        return jsonify({'error': '⚠️ Ein Makro läuft bereits. Bitte warten, bis es fertig ist.'})

    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        print(f"🎧 Erstelle {track_count} Spuren in Reaper...")
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Reaper...")
        time.sleep(3)
        
        # Reaper: Spuren einzeln erstellen
        for i in range(track_count):
            print(f"🔧 Erstelle Spur {i+1}/{track_count}")
            
            # Einfach: Cmd+T (der ursprünglich funktionierende Code)
            keyboard.press(Key.cmd)
            keyboard.press('t')
            keyboard.release('t')
            keyboard.release(Key.cmd)
            time.sleep(0.3)  # Pause zwischen Spuren
            
            # Falls das nicht funktioniert, versuche Menü-Zugang
            # Alternative: Track-Menü öffnen und "Insert new track" wählen
            # keyboard.press(Key.alt)
            # keyboard.press('t')  # Track menu
            # keyboard.release('t')
            # keyboard.release(Key.alt)
            # time.sleep(0.2)
            # keyboard.press('i')  # Insert new track
            # keyboard.release('i')
            # time.sleep(0.3)
        
        return jsonify({'success': f'✅ {track_count} Reaper Spuren erstellt (Cmd+T)!'})
        
    except Exception as e:
        return jsonify({'error': f'Reaper Fehler: {e}'})
    finally:
        automation_lock.release()

def name_tracks_protools(selected_items):
    """Pro Tools: Spuren benennen"""
    if not automation_lock.acquire(blocking=False):
        return jsonify({'error': '⚠️ Ein Makro läuft bereits. Bitte warten, bis es fertig ist.'})

    try:
        from pynput.keyboard import Key, Controller as KeyboardController
        keyboard = KeyboardController()
        
        print(f"🏷️ Benenne {len(selected_items)} Spuren in Pro Tools...")
        print("📋 ANLEITUNG:")
        print("   1. Öffne MANUELL den Namensgebungs-Dialog der ersten Spur")
        print("   2. Danach übernimmt die App automatisch alle weiteren Spuren")
        print("⏳ 2 Sekunden Zeit zum Öffnen des ersten Dialogs...")
        time.sleep(2)
        
        for i, item in enumerate(selected_items):
            trackname = item['trackname']
            print(f"🏷️ Spur {i+1}: '{trackname}'")

            # Bestehenden Namen markieren und überschreiben
            keyboard.press(Key.cmd)
            keyboard.press('a')
            keyboard.release('a')
            keyboard.release(Key.cmd)
            time.sleep(0.03)

            keyboard.type(trackname)
            time.sleep(0.05)

            # Zur nächsten Spur wechseln (außer bei der letzten)
            if i < len(selected_items) - 1:
                keyboard.press(Key.cmd)
                keyboard.press(Key.right)
                keyboard.release(Key.right)
                keyboard.release(Key.cmd)
                time.sleep(0.08)

        # Letzte Eingabe bestätigen
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        
        return jsonify({'success': f'✅ {len(selected_items)} Pro Tools Spuren benannt!'})
        
    except Exception as e:
        return jsonify({'error': f'Pro Tools Fehler: {e}'})
    finally:
        automation_lock.release()

def name_tracks_reaper(selected_items):
    """Reaper: Spuren benennen - Analog zu Pro Tools"""
    if not automation_lock.acquire(blocking=False):
        return jsonify({'error': '⚠️ Ein Makro läuft bereits. Bitte warten, bis es fertig ist.'})

    try:
        from pynput.keyboard import Key, Controller
        keyboard = Controller()
        
        print(f"🏷️ Benenne {len(selected_items)} Spuren in Reaper...")
        print("⏳ 3 Sekunden Zeit zum Wechseln zu Reaper...")
        print("💡 Markieren Sie die erste Spur in Reaper!")
        time.sleep(3)
        
        for i, item in enumerate(selected_items):
            trackname = item['trackname']
            print(f"🏷️ Spur {i+1}: '{trackname}'")
            
            # F2 drücken für Track Name Edit (analog zu Return in Pro Tools)
            keyboard.press(Key.f2)
            keyboard.release(Key.f2)
            time.sleep(0.3)
            
            # Trackname eingeben
            keyboard.type(trackname)
            time.sleep(0.2)
            
            # Enter bestätigen
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(0.2)
            
            # Tab zur nächsten Spur (analog zu Pfeil runter in Pro Tools)
            keyboard.press(Key.tab)
            keyboard.release(Key.tab)
            time.sleep(0.3)
        
        return jsonify({'success': f'✅ {len(selected_items)} Reaper Spuren benannt!'})
        
    except Exception as e:
        return jsonify({'error': f'Reaper Fehler: {e}'})
    finally:
        automation_lock.release()

@app.route('/export', methods=['POST'])
def export_tracks():
    try:
        selected_ids = request.form.getlist('selected_tracks')
        selected_ids = [int(id_) for id_ in selected_ids]
        
        if not selected_ids:
            return jsonify({'error': 'Bitte mindestens eine Spur auswählen'}), 400
        
        include_channel    = request.form.get('include_channel') == 'on'
        include_instrument = request.form.get('include_instrument') == 'on'
        include_microphone = request.form.get('include_microphone') == 'on'
        
        # Tracknamen dynamisch zusammenbauen
        names = []
        for item in current_data:
            if item['id'] in selected_ids:
                parts = []
                if include_channel    and item.get('kanal'):    parts.append(item['kanal'])
                if include_instrument and item.get('instrument'): parts.append(item['instrument'])
                if include_microphone and item.get('mikrofon'):  parts.append(item['mikrofon'])
                names.append('_'.join(parts) if parts else item['trackname'])
        
        content = ''.join(f"Spur {i}  [{name}]\n" for i, name in enumerate(names, 1))
        
        from flask import Response
        return Response(
            content,
            mimetype='text/plain',
            headers={'Content-Disposition': 'attachment; filename=tracknamen.txt'}
        )
        
    except Exception as e:
        return jsonify({'error': f'Export-Fehler: {e}'}), 500

if not os.path.exists('templates'):
    os.makedirs('templates')

def create_template():
    """HTML Template erstellen"""
    html_content = '''<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🎵 Multi-DAW Track Namer</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { text-align: center; margin-bottom: 30px; }
        .upload-section { background: #e3f2fd; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
        .layout-options { margin: 10px 0; }
        .layout-options input { margin-right: 5px; }
        .data-table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        .data-table th, .data-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        .data-table th { background-color: #f2f2f2; font-weight: bold; }
        .data-table tr:nth-child(even) { background-color: #f9f9f9; }
        .trackname { font-family: 'Courier New', monospace; font-weight: bold; color: #2196F3; }
        .buttons { margin: 20px 0; text-align: center; }
        .buttons button { margin: 0 10px; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; font-size: 14px; }
        .btn-primary { background-color: #2196F3; color: white; }
        .btn-success { background-color: #4CAF50; color: white; }
        .btn-warning { background-color: #FF9800; color: white; }
        .btn-danger { background-color: #f44336; color: white; }
        .flash-messages { margin: 20px 0; }
        .flash-success { background: #d4edda; color: #155724; padding: 10px; border-radius: 5px; border: 1px solid #c3e6cb; }
        .flash-error { background: #f8d7da; color: #721c24; padding: 10px; border-radius: 5px; border: 1px solid #f5c6cb; }
        .upload-btn { background-color: #2196F3; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }
        input[type="file"] { margin: 10px 0; }
        .stats { background: #e8f5e9; padding: 10px; border-radius: 5px; margin: 10px 0; }
        .status-box { background: #fff3cd; border: 1px solid #ffeaa7; padding: 10px; border-radius: 5px; margin: 10px 0; min-height: 30px; }
        .protools-section { background: #f0f8ff; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎵 Multi-DAW Track Namer</h1>
            <p>Unterstützt Pro Tools und Reaper - Web-basierte Version</p>
        </div>
        <!-- Flash Messages -->
        <div class="flash-messages">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="flash-{{ 'success' if category == 'success' else 'error' }}">
                            {{ message }}
                        </div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
        </div>
        
        <!-- Upload Section -->
        <div class="upload-section">
            <h3>📁 Excel-Datei laden</h3>
            <form method="POST" action="/upload" enctype="multipart/form-data">
                <input type="file" name="excel_file" accept=".xlsx,.xls,.xlsm" required>
                <div class="layout-options">
                    <label><input type="radio" name="layout" value="auto" checked> Auto-Erkennung</label>
                    <label><input type="radio" name="layout" value="layout_a"> Layout A (B/C/D)</label>
                    <label><input type="radio" name="layout" value="layout_b"> Layout B (B/D/E)</label>
                </div>
                <button type="submit" class="upload-btn">📁 Excel laden</button>
            </form>
        </div>
        
        <!-- Namensformat-Optionen - IMMER sichtbar -->
        <div class="upload-section">
            <h3>🏷️ Namensformat anpassen</h3>
            <div class="layout-options">
                <label><input type="checkbox" name="include_channel" checked> Kanalnummer (103, 79, 44...)</label><br>
                <label><input type="checkbox" name="include_instrument" checked> Instrument (TB Broadcast, Bugler 3, Tom 3...)</label><br>
                <label><input type="checkbox" name="include_microphone" checked> Mikrofon (Switchable, WL 185, SM 98...)</label>
            </div>
            <p><strong>Beispiel:</strong> <span id="name-preview">103_TB Broadcast_Switchable</span></p>
        </div>
        
        {% if data %}
        <div class="stats">
            📊 <strong>{{ data|length }}</strong> Spuren geladen | 
            <span id="selected-count">{{ data|length }}</span> ausgewählt
        </div>
        
        <!-- Auswahl-Buttons -->
        <div class="buttons">
            <button class="btn-primary" onclick="selectAll()">✅ Alle auswählen</button>
            <button class="btn-primary" onclick="selectNone()">❌ Alle abwählen</button>
        </div>
        
        <!-- Daten-Tabelle -->
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">✓</th>
                    <th>Kanal</th>
                    <th>Instrument</th>
                    <th>Mikrofon</th>
                    <th>Track-Name</th>
                </tr>
            </thead>
            <tbody>
                {% for item in data %}
                <tr>
                    <td>
                        <input type="checkbox" name="selected_tracks" value="{{ item.id }}" 
                               {% if item.selected %}checked{% endif %}>
                    </td>
                    <td>{{ item.kanal }}</td>
                    <td>{{ item.instrument }}</td>
                    <td>{{ item.mikrofon }}</td>
                    <td class="trackname">{{ item.trackname }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        
        <!-- DAW Integration -->
        <div class="protools-section">
            <h3>🎵 Multi-DAW Integration</h3>
            <div class="status-box" id="status-box">
                Bereit für Pro Tools und Reaper...
            </div>
            
            <!-- Pro Tools Buttons -->
            <div style="margin-bottom: 15px;">
                <h4>🎹 Pro Tools</h4>
                <div class="buttons">
                    <button class="btn-success" onclick="createTracksProtools()">
                        🏗️ Pro Tools - Spuren erstellen
                    </button>
                    <button class="btn-warning" onclick="nameTracksProtools()">
                        🏷️ Pro Tools - Spuren benennen
                    </button>
                </div>
            </div>
            
            <!-- Reaper Buttons -->
            <div>
                <h4>🎧 Reaper</h4>
                <div class="buttons">
                    <button class="btn-success" onclick="createTracksReaper()">
                        🏗️ Reaper - Spuren erstellen
                    </button>
                    <button class="btn-warning" onclick="nameTracksReaper()">
                        🏷️ Reaper - Spuren benennen
                    </button>
                </div>
            </div>
        </div>
        
        <!-- Export -->
        <div class="buttons">
            <button type="button" class="btn-primary" onclick="exportTracks()">
                💾 Export als TXT
            </button>
        </div>
        {% endif %}
    </div>

    <script>
        let isProtoolsNamingRunning = false;
        let isReaperNamingRunning = false;

        // Basis-Funktionen
        function updateSelectedCount() {
            const checked = document.querySelectorAll('input[name="selected_tracks"]:checked');
            if (document.getElementById('selected-count')) {
                document.getElementById('selected-count').textContent = checked.length;
            }
        }
        
        function selectAll() {
            console.log('🔍 DEBUG: selectAll() aufgerufen');
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]');
            console.log('🔍 DEBUG: Gefundene Checkboxen:', checkboxes.length);
            checkboxes.forEach(cb => cb.checked = true);
            updateSelectedCount();
        }
        
        function selectNone() {
            console.log('🔍 DEBUG: selectNone() aufgerufen');
            const checkboxes = document.querySelectorAll('input[name="selected_tracks"]');
            console.log('🔍 DEBUG: Gefundene Checkboxen:', checkboxes.length);
            checkboxes.forEach(cb => cb.checked = false);
            updateSelectedCount();
        }
        
        function getSelectedTracks() {
            const checked = document.querySelectorAll('input[name="selected_tracks"]:checked');
            return Array.from(checked).map(cb => cb.value);
        }
        
        function showStatus(message, isError = false) {
            const statusBox = document.getElementById('status-box');
            if (statusBox) {
                statusBox.textContent = message;
                statusBox.style.backgroundColor = isError ? '#f8d7da' : '#d4edda';
                statusBox.style.color = isError ? '#721c24' : '#155724';
            }
        }
        
        function updateNamePreview() {
            const channelCheckbox = document.querySelector('input[name="include_channel"]');
            const instrumentCheckbox = document.querySelector('input[name="include_instrument"]');
            const microphoneCheckbox = document.querySelector('input[name="include_microphone"]');
            
            if (!channelCheckbox || !instrumentCheckbox || !microphoneCheckbox) {
                return; // Checkboxen nicht gefunden
            }
            
            const parts = [];
            if (channelCheckbox.checked) parts.push('103');
            if (instrumentCheckbox.checked) parts.push('TB Broadcast');
            if (microphoneCheckbox.checked) parts.push('Switchable');
            
            const preview = parts.length > 0 ? parts.join(' ') : 'Keine Auswahl';
            const previewElement = document.getElementById('name-preview');
            if (previewElement) {
                previewElement.textContent = preview;
            }
        }
        
        function setupTrackCheckboxListeners() {
            const checkboxes = Array.from(document.querySelectorAll('input[name="selected_tracks"]'));
            let lastClickedTrackCheckbox = null;

            checkboxes.forEach(cb => {
                cb.addEventListener('change', updateSelectedCount);
                cb.addEventListener('click', event => {
                    if (!event.shiftKey || !lastClickedTrackCheckbox) {
                        lastClickedTrackCheckbox = cb;
                        return;
                    }

                    const start = checkboxes.indexOf(lastClickedTrackCheckbox);
                    const end = checkboxes.indexOf(cb);
                    if (start === -1 || end === -1) {
                        lastClickedTrackCheckbox = cb;
                        return;
                    }

                    const shouldCheck = cb.checked;
                    const from = Math.min(start, end);
                    const to = Math.max(start, end);

                    for (let i = from; i <= to; i++) {
                        checkboxes[i].checked = shouldCheck;
                    }

                    updateSelectedCount();
                    lastClickedTrackCheckbox = cb;
                });
            });
        }
        
        function setupGlobalHotkeys() {
            try {
                fetch('/setup_hotkeys', { method: 'POST' });
                console.log('🎹 Globale Hotkeys aktiviert');
            } catch (e) {
                console.log('⚠️ Hotkeys nicht verfügbar');
            }
        }
        
        // Pro Tools Funktionen
        function createTracksProtools() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showStatus(`⏳ Erstelle ${selected.length} Spuren in Pro Tools...`);
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/create_tracks_protools', {
                method: 'POST', 
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showStatus(data.error, true);
                } else {
                    showStatus(data.success);
                }
            })
            .catch(error => {
                showStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function nameTracksProtools() {
            if (isProtoolsNamingRunning) {
                showStatus('⚠️ Pro Tools Benennung läuft bereits. Bitte warten...');
                return;
            }

            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }

            isProtoolsNamingRunning = true;
            const protoolsNameBtn = document.querySelector("button[onclick='nameTracksProtools()']");
            if (protoolsNameBtn) {
                protoolsNameBtn.disabled = true;
                protoolsNameBtn.style.opacity = '0.6';
                protoolsNameBtn.style.cursor = 'not-allowed';
            }
            
            showStatus(`⏳ Benenne ${selected.length} Spuren in Pro Tools...`);
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            const includeChannel = document.querySelector('input[name="include_channel"]');
            const includeInstrument = document.querySelector('input[name="include_instrument"]');
            const includeMicrophone = document.querySelector('input[name="include_microphone"]');
            formData.append('include_channel', includeChannel && includeChannel.checked ? 'on' : 'off');
            formData.append('include_instrument', includeInstrument && includeInstrument.checked ? 'on' : 'off');
            formData.append('include_microphone', includeMicrophone && includeMicrophone.checked ? 'on' : 'off');
            
            fetch('/name_tracks_protools', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showStatus(data.error, true);
                } else {
                    showStatus(data.success);
                }
            })
            .catch(error => {
                showStatus('❌ Netzwerk-Fehler: ' + error, true);
            })
            .finally(() => {
                isProtoolsNamingRunning = false;
                if (protoolsNameBtn) {
                    protoolsNameBtn.disabled = false;
                    protoolsNameBtn.style.opacity = '';
                    protoolsNameBtn.style.cursor = '';
                }
            });
        }
        
        // Reaper Funktionen
        function createTracksReaper() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            
            showStatus(`⏳ Erstelle ${selected.length} Spuren in Reaper...`);
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            
            fetch('/create_tracks_reaper', {
                method: 'POST', 
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showStatus(data.error, true);
                } else {
                    showStatus(data.success);
                }
            })
            .catch(error => {
                showStatus('❌ Netzwerk-Fehler: ' + error, true);
            });
        }
        
        function nameTracksReaper() {
            if (isReaperNamingRunning) {
                showStatus('⚠️ Reaper Benennung läuft bereits. Bitte warten...');
                return;
            }

            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }

            isReaperNamingRunning = true;
            const reaperNameBtn = document.querySelector("button[onclick='nameTracksReaper()']");
            if (reaperNameBtn) {
                reaperNameBtn.disabled = true;
                reaperNameBtn.style.opacity = '0.6';
                reaperNameBtn.style.cursor = 'not-allowed';
            }
            
            showStatus(`⏳ Benenne ${selected.length} Spuren in Reaper...`);
            
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            const includeChannel = document.querySelector('input[name="include_channel"]');
            const includeInstrument = document.querySelector('input[name="include_instrument"]');
            const includeMicrophone = document.querySelector('input[name="include_microphone"]');
            formData.append('include_channel', includeChannel && includeChannel.checked ? 'on' : 'off');
            formData.append('include_instrument', includeInstrument && includeInstrument.checked ? 'on' : 'off');
            formData.append('include_microphone', includeMicrophone && includeMicrophone.checked ? 'on' : 'off');
            
            fetch('/name_tracks_reaper', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.error) {
                    showStatus(data.error, true);
                } else {
                    showStatus(data.success);
                }
            })
            .catch(error => {
                showStatus('❌ Netzwerk-Fehler: ' + error, true);
            })
            .finally(() => {
                isReaperNamingRunning = false;
                if (reaperNameBtn) {
                    reaperNameBtn.disabled = false;
                    reaperNameBtn.style.opacity = '';
                    reaperNameBtn.style.cursor = '';
                }
            });
        }
        
        async function exportTracks() {
            const selected = getSelectedTracks();
            if (selected.length === 0) {
                showStatus('❌ Bitte mindestens eine Spur auswählen', true);
                return;
            }
            const formData = new FormData();
            selected.forEach(id => formData.append('selected_tracks', id));
            // Namensformat-Checkboxen mitsenden
            const inclChannel = document.querySelector('input[name="include_channel"]');
            const inclInstrument = document.querySelector('input[name="include_instrument"]');
            const inclMicrophone = document.querySelector('input[name="include_microphone"]');
            if (inclChannel && inclChannel.checked) formData.append('include_channel', 'on');
            if (inclInstrument && inclInstrument.checked) formData.append('include_instrument', 'on');
            if (inclMicrophone && inclMicrophone.checked) formData.append('include_microphone', 'on');
            try {
                const response = await fetch('/export', { method: 'POST', body: formData });
                if (!response.ok) throw new Error('Server-Fehler');
                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = 'tracknamen.txt';
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                window.URL.revokeObjectURL(url);
                showStatus(`✅ ${selected.length} Tracknamen exportiert`);
            } catch (e) {
                showStatus('❌ Export-Fehler: ' + e, true);
            }
        }
        
        // Globale Hotkeys Info
        function setupGlobalHotkeys() {
            try {
                fetch('/setup_hotkeys', { method: 'POST' });
                console.log('🎹 Globale Hotkeys aktiviert');
            } catch (e) {
                console.log('⚠️ Hotkeys nicht verfügbar');
            }
        }
        
        // Function um Event-Listener zu setzen (wird nach Excel-Upload aufgerufen)
        function setupTrackCheckboxListeners() {
            const checkboxes = Array.from(document.querySelectorAll('input[name="selected_tracks"]'));
            let lastClickedTrackCheckbox = null;

            checkboxes.forEach(cb => {
                cb.addEventListener('change', updateSelectedCount);
                cb.addEventListener('click', event => {
                    if (!event.shiftKey || !lastClickedTrackCheckbox) {
                        lastClickedTrackCheckbox = cb;
                        return;
                    }

                    const start = checkboxes.indexOf(lastClickedTrackCheckbox);
                    const end = checkboxes.indexOf(cb);
                    if (start === -1 || end === -1) {
                        lastClickedTrackCheckbox = cb;
                        return;
                    }

                    const shouldCheck = cb.checked;
                    const from = Math.min(start, end);
                    const to = Math.max(start, end);

                    for (let i = from; i <= to; i++) {
                        checkboxes[i].checked = shouldCheck;
                    }

                    updateSelectedCount();
                    lastClickedTrackCheckbox = cb;
                });
            });
        }
        
        // Initial setup - wird beim Laden der Seite ausgeführt
        function initializePage() {
            updateSelectedCount();
            updateNamePreview();
            setupTrackCheckboxListeners();
            setupNamePreviewListeners();
            setupGlobalHotkeys();
        }
        
        function setupNamePreviewListeners() {
            const checkboxes = document.querySelectorAll('input[name="include_channel"], input[name="include_instrument"], input[name="include_microphone"]');
            checkboxes.forEach(cb => {
                cb.addEventListener('change', updateNamePreview);
            });
        }
        
        // Seite initialisieren wenn DOM geladen ist
        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', initializePage);
        } else {
            initializePage();
        }
    </script>
</body>
</html>'''
    
    with open('templates/index.html', 'w', encoding='utf-8') as f:
        f.write(html_content)

def setup_global_hotkeys():
    """Globale Hotkeys für beide DAWs"""
    global hotkey_listener

    if hotkey_listener is not None:
        return

    try:
        from pynput import keyboard
        
        def on_hotkey_create():
            """Ctrl+Alt+C: Spuren erstellen"""
            print(f"🔥 Ctrl+Alt+C gedrückt - Erstelle Spuren in {selected_daw.upper()}...")
            # Hier könnten wir die Backend-Funktionen aufrufen
        
        def on_hotkey_name():
            """Ctrl+Alt+N: Spuren benennen"""
            print(f"🔥 Ctrl+Alt+N gedrückt - Benenne Spuren in {selected_daw.upper()}...")
            # Hier könnten wir die Backend-Funktionen aufrufen
            
        def on_hotkey_switch():
            """Ctrl+Alt+S: DAW wechseln"""
            global selected_daw
            selected_daw = "reaper" if selected_daw == "protools" else "protools"
            print(f"🔄 DAW gewechselt zu: {selected_daw.upper()}")
        
        with hotkeys_lock:
            if hotkey_listener is not None:
                return

            # Globale Hotkeys genau einmal registrieren
            hotkey_listener = keyboard.GlobalHotKeys({
                '<ctrl>+<alt>+c': on_hotkey_create,
                '<ctrl>+<alt>+n': on_hotkey_name,
                '<ctrl>+<alt>+s': on_hotkey_switch,
            })
            hotkey_listener.start()
        
        print("🎹 Multi-DAW Hotkeys registriert:")
        print("   Ctrl+Alt+C (Spuren erstellen), Ctrl+Alt+N (Spuren benennen), Ctrl+Alt+S (DAW wechseln)")
        
    except Exception as e:
        print(f"⚠️ Hotkey-Fehler: {e}")

@app.route('/setup_hotkeys', methods=['POST'])
def setup_hotkeys_route():
    """Hotkeys über Web-Interface aktivieren"""
    setup_global_hotkeys()
    return jsonify({'success': 'Hotkeys aktiviert'})

def open_browser():
    """Browser automatisch öffnen"""
    time.sleep(1.5)
    webbrowser.open('http://127.0.0.1:5001')

if __name__ == "__main__":
    print("🎵 Multi-DAW Track Namer startet...")
    print("🌐 Öffne Browser auf: http://127.0.0.1:5001")
    print("🎹 Hotkeys: Ctrl+Alt+C (Erstellen), Ctrl+Alt+N (Benennen), Ctrl+Alt+S (DAW wechseln)")
    print("⚠️  Zum Beenden: Strg+C drücken")
    
    # Template erstellen
    create_template()
    
    # Globale Hotkeys einrichten
    setup_global_hotkeys()
    
    # Browser öffnen
    threading.Thread(target=open_browser, daemon=True).start()
    
    # Flask-App starten
    try:
        app.run(host='127.0.0.1', port=5001, debug=False)
    except KeyboardInterrupt:
        print("\n🛑 Multi-DAW Track Namer beendet")