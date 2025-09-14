import argparse
import sys
import time
from pathlib import Path
import threading
from pynput import keyboard
from pynput.keyboard import Key, Controller

class TrackNamer:
    def __init__(self, kb: Controller, names: list[tuple[str, str | None, str | None]], args):
        self.kb = kb
        self.names = names
        self.args = args
        self.idx = 0
        self.paused = False
        self.auto_running = False
        self.stop_requested = False
        self.worker_thread = None

    def select_all(self):
        try:
            if self.args.auto_next == 'mac':
                with self.kb.pressed(Key.cmd):
                    self.kb.press('a')
                    self.kb.release('a')
            else:
                with self.kb.pressed(Key.ctrl):
                    self.kb.press('a')
                    self.kb.release('a')
        except Exception:
            pass

    def activate_next_track(self):
        """Wechselt zur nächsten Spur und aktiviert diese"""
        print("DEBUG: Sende Strg+Rechts")
        # Sicherstellen dass keine Taste hängt
        self.kb.release(Key.ctrl)
        self.kb.release(Key.right)
        time.sleep(0.2)
        
        # Strg+Rechts senden
        with self.kb.pressed(Key.ctrl):
            self.kb.press(Key.right)
            time.sleep(0.1)
            self.kb.release(Key.right)
        time.sleep(0.2)

        # Spur aktivieren
        self.kb.press(Key.enter)
        self.kb.release(Key.enter)
        time.sleep(0.2)

    def type_name(self, name: str, channel: str | None, mic: str | None):
        """Tippt einen formatierten Namen"""
        formatted_name = name
        if channel and self.args.use_kanal:
            formatted_name = f"{channel}_{formatted_name}"
        if mic and self.args.use_mic:
            formatted_name = f"{formatted_name}_{mic}"
        
        # Leerzeichen durch Unterstriche ersetzen
        formatted_name = formatted_name.replace(" ", "_")
        
        self.select_all()
        time.sleep(0.1)
        self.kb.type(formatted_name)
        print(f"Getippt: {formatted_name}")
        
        if self.args.auto_next:
            time.sleep(self.args.next_delay)
            self.activate_next_track()
        elif self.args.enter:
            self.kb.press(Key.enter)
            self.kb.release(Key.enter)

    def on_press(self, key):
        try:
            if key == Key.f9:
                if self.auto_running:
                    self.stop_requested = True
                    if self.worker_thread and self.worker_thread.is_alive():
                        self.worker_thread.join(timeout=1)
                    self.auto_running = False
                self.paused = not self.paused
                print(f"Pause: {self.paused}")
            elif key == Key.f7:
                if self.idx > 0:
                    self.idx -= 1
                print(f"Zurück. Nächster Index: {self.idx}")
            elif key == Key.f8 and not self.paused:
                if self.idx >= len(self.names):
                    print("Fertig: Alle Namen wurden verwendet.")
                    return
                name, channel, mic = self.names[self.idx]
                self.idx += 1
                time.sleep(self.args.delay)
                self.type_name(name, channel, mic)
            elif key == Key.esc:
                print("Beendet.")
                return False
        except Exception as e:
            print(f"Fehler beim Senden der Tasten: {e}")

def read_cell(ws, row: int, col: int) -> str | None:
    """Liest eine Zelle und gibt den Wert als String zurück, oder None wenn leer"""
    try:
        val = ws.cell(row=row, column=col).value
        if val is not None:
            val = str(val).strip()
            if val:
                return val
    except Exception as e:
        print(f"DEBUG: Fehler beim Lesen von Zelle R{row}C{col}: {e}")
    return None

def read_names_from_xlsx(path: Path, args) -> list[tuple[str, str | None, str | None]]:
    """Liest Namen aus einer Excel-Datei mit konfigurierbaren Spalten"""
    try:
        import openpyxl
    except ImportError:
        print("Fehler: Das Paket 'openpyxl' ist nicht installiert.")
        print("Installiere es mit: pip install openpyxl")
        sys.exit(1)

    print(f"DEBUG: Öffne Excel-Datei: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    # Arbeitsblatt auswählen
    if args.blatt:
        if args.blatt not in wb.sheetnames:
            print("Verfügbare Arbeitsblätter:")
            for i, name in enumerate(wb.sheetnames):
                print(f"  {i}: {name}")
            sys.exit(1)
        ws = wb[args.blatt]
    else:
        ws = wb.active
    print(f"DEBUG: Verwende Arbeitsblatt: {ws.title}")

    names = []
    
    print(f"\nDEBUG: Spalten-Konfiguration:")
    print(f"  Namen: Spalte {args.name_spalte}")
    print(f"  Kanäle: Spalte {args.kanal_spalte}")
    print(f"  Mikrofone: Spalte {args.mic_spalte}")
    
    # Zeige erste Zeile zur Kontrolle
    print("\nDEBUG: Erste Zeile (Header):")
    for col in [args.name_spalte, args.kanal_spalte, args.mic_spalte]:
        val = read_cell(ws, 1, col)
        print(f"  Spalte {col}: {val}")
    
    # Namen ab der konfigurierten Zeile einlesen
    print("\nDEBUG: Lese Namen...")
    start_row = args.start_zeile if args.start_zeile else 2
    for row in range(start_row, ws.max_row + 1):
        name = read_cell(ws, row, args.name_spalte)
        if name:
            channel = read_cell(ws, row, args.kanal_spalte) if args.use_kanal else None
            mic = read_cell(ws, row, args.mic_spalte) if args.use_mic else None
            names.append((name, channel, mic))
            if args.debug and len(names) <= 3:
                print(f"\nDEBUG: Zeile {row}:")
                print(f"  Name ({args.name_spalte}): {name}")
                print(f"  Kanal ({args.kanal_spalte}): {channel}")
                print(f"  Mikrofon ({args.mic_spalte}): {mic}")
    
    print(f"\nINFO: {len(names)} Namen gefunden")
    if args.debug and names:
        print("Erste 3 Namen zur Kontrolle:")
        for i, (name, channel, mic) in enumerate(names[:3]):
            print(f"  {i+1}. Name: {name}, Kanal: {channel}, Mikrofon: {mic}")
            
    return names

def main():
    parser = argparse.ArgumentParser(description="Pro Tools Track Namer")
    parser.add_argument("excel_file", help="Excel-Datei mit den Namen")
    parser.add_argument("--blatt", help="Name des Excel-Sheets")
    parser.add_argument("--name-spalte", type=int, default=1, help="Spaltennummer für Namen (1=A, 2=B, ...)")
    parser.add_argument("--kanal-spalte", type=int, default=2, help="Spaltennummer für Kanäle (1=A, 2=B, ...)")
    parser.add_argument("--mic-spalte", type=int, default=5, help="Spaltennummer für Mikrofone (1=A, 2=B, ...)")
    parser.add_argument("--start-zeile", type=int, help="Erste Zeile mit Daten (Standard: 2)")
    parser.add_argument("--use-kanal", action="store_true", help="Kanalnummer zum Namen hinzufügen")
    parser.add_argument("--use-mic", action="store_true", help="Mikrofonname zum Namen hinzufügen")
    parser.add_argument("--delay", type=float, default=0.2, help="Verzögerung vor dem Tippen (Sekunden)")
    parser.add_argument("--next-delay", type=float, default=0.5, help="Verzögerung vor dem Spurwechsel (Sekunden)")
    parser.add_argument("--enter", action="store_true", help="Nach dem Namen Enter drücken")
    parser.add_argument("--auto-next", choices=['windows', 'mac'], default='windows', help="Automatisch zur nächsten Spur")
    parser.add_argument("--debug", action="store_true", help="Debug-Ausgaben aktivieren")
    args = parser.parse_args()

    path = Path(args.excel_file)
    if not path.exists():
        print(f"Fehler: Datei nicht gefunden: {path}")
        sys.exit(1)

    names = read_names_from_xlsx(path, args)
    if not names:
        print("Keine Namen gefunden!")
        sys.exit(1)

    print("\nBereit! Anleitung:")
    print("1) Erstelle/Ordne die Spuren in Pro Tools")
    print("2) Doppelklicke die erste Spurbezeichnung")
    print("3) Drücke F8 für jeden Namen (F7 für zurück, ESC zum Beenden)")

    kb = Controller()
    namer = TrackNamer(kb, names, args)
    
    with keyboard.Listener(on_press=namer.on_press) as listener:
        listener.join()

if __name__ == "__main__":
    main()